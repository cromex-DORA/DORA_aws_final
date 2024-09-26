# -*- coding: utf-8 -*-
import os
import pandas as pd
import io
from app.DORApy.classes.modules import connect_path
from openpyxl import load_workbook
pd.set_option("display.max_rows", None, "display.max_columns", None,'display.max_colwidth',None)
pd.options.mode.chained_assignment = None  # default='warn'

###Production
environment = os.getenv('ENVIRONMENT')

#####################################################################################################
#DORA
#####################################################################################################
def creation_dicts_config():
    filename = ("config\\DORA\\DORA_config.csv")
    df_config_obj = connect_path.conv_s3_obj_vers_python_obj("config",filename)
    df_config = pd.read_csv(df_config_obj, index_col=0)
    df_config=df_config.dropna(axis=0)
    df_config_espace = df_config.apply(pd.to_numeric, errors='coerce')
    df_config_espace=df_config_espace.dropna(axis=0)
    df_config_police = df_config.loc[~df_config.index.isin(list(df_config_espace.index.values))]
    dict_config_espace = df_config_espace.to_dict('index')
    dict_config_police = df_config_police.to_dict('index')
    return dict_config_espace,dict_config_police

def import_csv_placement_icone_bloc_icone():
    filename = ("config\\DORA\\placement icone bloc icone n parmi n x.csv")

    df_csv_x_icone_bloc_icone = connect_path.conv_s3_obj_vers_python_obj("config",filename)
    df_csv_x_icone_bloc_icone = pd.read_csv(df_csv_x_icone_bloc_icone)

    filename = ("config\\DORA\\placement icone bloc icone n parmi n y.csv")
    df_csv_y_icone_bloc_icone = connect_path.conv_s3_obj_vers_python_obj("config",filename)
    df_csv_y_icone_bloc_icone = pd.read_csv(df_csv_y_icone_bloc_icone)
    return df_csv_x_icone_bloc_icone,df_csv_y_icone_bloc_icone

def creation_dict_dict_config_df_actions_MIA():
    dict_dict_config_df_actions_MIA = {}
    filename = ("config\\DORA\\dict_changement_nom_et_type_colonne_tableau_actions_MIA_MO.csv")
    df_chgmt_nom_et_type_col_tableau_actions_MIA_MO = connect_path.conv_s3_obj_vers_python_obj("config",filename)  
    df_chgmt_nom_et_type_col_tableau_actions_MIA_MO = pd.read_csv(df_chgmt_nom_et_type_col_tableau_actions_MIA_MO)  
    df_chgmt_nom_et_type_col_tableau_actions_MIA_MO=df_chgmt_nom_et_type_col_tableau_actions_MIA_MO.rename(columns=lambda x: x.strip())
    
    dict_nom_col_MO = {}
    for nom_col in list(df_chgmt_nom_et_type_col_tableau_actions_MIA_MO):
        dict_nom_col_MO[nom_col] = df_chgmt_nom_et_type_col_tableau_actions_MIA_MO[nom_col].iloc[0]
    dict_nom_col_MO = {k:v for k,v in dict_nom_col_MO.items() if v==v}
    dict_dict_config_df_actions_MIA['dict_conv_nom_col_DORA_MO'] = dict_nom_col_MO
    dict_inv = {v: k for k, v in dict_nom_col_MO.items()}
    
    dict_type_col_MO = {}
    for nom_col in list(list(dict_nom_col_MO.values())):
        dict_type_col_MO[nom_col] = df_chgmt_nom_et_type_col_tableau_actions_MIA_MO[dict_inv[nom_col]].iloc[1]
    dict_dict_config_df_actions_MIA['dict_conv_type_col_DORA_MO'] = dict_type_col_MO

    dict_nb_chiffres_col_DORA_MO = {}
    for nom_col in list(list(dict_nom_col_MO.values())):
        dict_nb_chiffres_col_DORA_MO[nom_col] = df_chgmt_nom_et_type_col_tableau_actions_MIA_MO[dict_inv[nom_col]].iloc[2]
    dict_dict_config_df_actions_MIA['dict_nb_chiffres_col_DORA_MO'] = dict_nb_chiffres_col_DORA_MO
    
    filename = ("config\\DORA\\dict_changement_nom_et_type_colonne_tableau_actions_MIA_DEP.csv")
    df_chgmt_nom_et_type_col_tableau_actions_MIA_DEP = connect_path.conv_s3_obj_vers_python_obj("config",filename)   
    df_chgmt_nom_et_type_col_tableau_actions_MIA_DEP = pd.read_csv(df_chgmt_nom_et_type_col_tableau_actions_MIA_DEP)     
    df_chgmt_nom_et_type_col_tableau_actions_MIA_DEP=df_chgmt_nom_et_type_col_tableau_actions_MIA_DEP.rename(columns=lambda x: x.strip())
    
    dict_nom_col_DEP = {}
    for nom_col in list(df_chgmt_nom_et_type_col_tableau_actions_MIA_DEP):
        dict_nom_col_DEP[nom_col] = df_chgmt_nom_et_type_col_tableau_actions_MIA_DEP[nom_col].iloc[0]
    dict_nom_col_DEP = {k:v for k,v in dict_nom_col_DEP.items() if v==v}
    dict_dict_config_df_actions_MIA['dict_conv_nom_col_DORA_DEP'] = dict_nom_col_DEP
    dict_inv = {v: k for k, v in dict_nom_col_DEP.items()}
    
    dict_type_col_DEP = {}
    for nom_col in list(list(dict_nom_col_DEP.values())):
        dict_type_col_DEP[nom_col] = df_chgmt_nom_et_type_col_tableau_actions_MIA_DEP[dict_inv[nom_col]].iloc[1]
    dict_dict_config_df_actions_MIA['dict_conv_type_col_DORA_DEP'] = dict_type_col_DEP

    dict_nb_chiffres_col_DORA_DEP = {}
    for nom_col in list(list(dict_nom_col_DEP.values())):
        dict_nb_chiffres_col_DORA_DEP[nom_col] = df_chgmt_nom_et_type_col_tableau_actions_MIA_DEP[dict_inv[nom_col]].iloc[2]
    dict_dict_config_df_actions_MIA['dict_nb_chiffres_col_DORA_DEP'] = dict_nb_chiffres_col_DORA_DEP    

    filename = ("config\\Osmose\\dict_changement_nom_et_type_colonne_BDD_Osmose.csv")
    df_Osmose_vierge = connect_path.conv_s3_obj_vers_python_obj("config",filename)  
    df_Osmose_vierge = pd.read_csv(df_Osmose_vierge) 
    df_Osmose_vierge=df_Osmose_vierge.rename(columns=lambda x: x.strip())
    
    dict_nom_col_BDD_Osmose = {}
    for nom_col in list(df_Osmose_vierge):
        dict_nom_col_BDD_Osmose[nom_col] = df_Osmose_vierge[nom_col].iloc[0]
    dict_nom_col_BDD_Osmose = {k:v for k,v in dict_nom_col_BDD_Osmose.items() if v==v}
    dict_dict_config_df_actions_MIA['dict_conv_nom_col_Osmose'] = dict_nom_col_BDD_Osmose
    dict_inv = {v: k for k, v in dict_nom_col_BDD_Osmose.items()}
    
    dict_type_col_BDD_Osmose = {}
    for nom_col in list(list(dict_nom_col_BDD_Osmose.values())):
        dict_type_col_BDD_Osmose[nom_col] = df_Osmose_vierge[dict_inv[nom_col]].iloc[1]
    dict_dict_config_df_actions_MIA['dict_conv_type_col_Osmose'] = dict_type_col_BDD_Osmose

    dict_nb_chiffres_col_BDD_Osmose = {}
    for nom_col in list(list(dict_nom_col_BDD_Osmose.values())):
        dict_nb_chiffres_col_BDD_Osmose[nom_col] = df_Osmose_vierge[dict_inv[nom_col]].iloc[2]
    dict_dict_config_df_actions_MIA['dict_nb_chiffres_col_Osmose'] = dict_nb_chiffres_col_BDD_Osmose
    
    return dict_dict_config_df_actions_MIA

def creation_dict_reduction_nom_col_bloc_et_boite():
    dict_reduction_nom_col_bloc_et_boite = {}
    filename = ('config\\DORA\\fichier_reduction_nom_colonne_df_contour.csv')
    dict_reduction_nom_col_bloc_et_boite['df_contour'] = connect_path.conv_s3_obj_vers_python_obj("config",filename)  
    dict_reduction_nom_col_bloc_et_boite['df_contour'] = pd.read_csv(dict_reduction_nom_col_bloc_et_boite['df_contour'])

    filename = ('config\\DORA\\fichier_reduction_nom_colonne_df_texte_simple.csv')
    dict_reduction_nom_col_bloc_et_boite['df_texte_simple'] = connect_path.conv_s3_obj_vers_python_obj("config",filename)  
    dict_reduction_nom_col_bloc_et_boite['df_texte_simple'] = pd.read_csv(dict_reduction_nom_col_bloc_et_boite['df_texte_simple'])

    filename = ('config\\DORA\\fichier_reduction_nom_colonne_df_icone.csv')
    dict_reduction_nom_col_bloc_et_boite['df_icone'] = connect_path.conv_s3_obj_vers_python_obj("config",filename)
    dict_reduction_nom_col_bloc_et_boite['df_icone'] = pd.read_csv(dict_reduction_nom_col_bloc_et_boite['df_icone'])      

    filename = ('config\\DORA\\fichier_reduction_nom_colonne_df_bloc_lignes_multiples.csv')
    dict_reduction_nom_col_bloc_et_boite['df_lm'] = connect_path.conv_s3_obj_vers_python_obj("config",filename) 
    dict_reduction_nom_col_bloc_et_boite['df_lm'] = pd.read_csv(dict_reduction_nom_col_bloc_et_boite['df_lm']) 
    return dict_reduction_nom_col_bloc_et_boite

def import_dict_config_actions_MIA_DORA():
    dict_config_DORA_actions_MIA = {}
    filename = ("config\\DORA\\config_liste_action_DORA.ods")
    obj_df_config_DORA = connect_path.conv_s3_obj_vers_python_obj("config",filename)

    # Charger le contenu ODS en une liste de données
    df = pd.read_excel(io.BytesIO(obj_df_config_DORA), engine='odf')
    df_config_DORA = df.rename(columns=lambda x: x.strip())
    df_config_DORA = df_config_DORA.rename({'Type_action':"NOM_TYPE_ACTION_DORA",'Intégration possible du PAOT':'compatible_PAOT','Code Osmose associé':"CODE_TYPE_ACTION_OSMOSE"},axis=1)


    liste_col_commune_a_garder = [x for x in list(df_config_DORA) if not any(x.startswith(s) for s in ['TECH',"ELU"])]
    list_col_ELU = [x for x in list(df_config_DORA) if x.startswith("ELU")]
    list_col_TECH = [x for x in list(df_config_DORA) if x.startswith("TECH")]

    #Elu
    df_DORA_actions_MIA_elu = df_config_DORA[liste_col_commune_a_garder + list_col_ELU]
    df_DORA_actions_MIA_elu.columns = [colonne.replace("ELU-", '') for colonne in list(df_DORA_actions_MIA_elu)]

    #Tech
    df_DORA_actions_MIA_tech = df_config_DORA[liste_col_commune_a_garder + list_col_TECH]
    df_DORA_actions_MIA_elu.columns = [colonne.replace("TECH-", '') for colonne in list(df_DORA_actions_MIA_elu)]
    
    #Convertisseur Osmose
    df_DORA_actions_MIA_conv_Osmose = df_config_DORA[liste_col_commune_a_garder]
    df_DORA_actions_MIA_conv_Osmose.loc[~df_DORA_actions_MIA_conv_Osmose["NOM_TYPE_ACTION_DORA"].isnull()]


    dict_config_DORA_actions_MIA["df_actions_MIA_elu"] = df_DORA_actions_MIA_elu
    dict_config_DORA_actions_MIA["df_actions_MIA_tech"] = df_DORA_actions_MIA_tech
    dict_config_DORA_actions_MIA["df_DORA_actions_MIA_conv_Osmose"] = df_DORA_actions_MIA_conv_Osmose
    
    return dict_config_DORA_actions_MIA


def import_df_BDD_DORA_vierge():
    filename = ("config\\DORA\\dict_type_col_BDD_DORA.csv")
    filename = connect_path.conv_s3_obj_vers_python_obj("config",filename)       
    df_BDD_DORA_vierge = pd.read_csv(filename)
    return df_BDD_DORA_vierge

def import_dict_config_col_BDD_DORA_vierge():
    df_BDD_DORA_vierge = import_df_BDD_DORA_vierge()
    df_BDD_DORA_vierge=df_BDD_DORA_vierge.rename(columns=lambda x: x.strip())
    
    dict_config_col_BDD_DORA_vierge={}
    
    dict_type_col_BDD_DORA = {}
    for nom_col in list(df_BDD_DORA_vierge):
        dict_type_col_BDD_DORA[nom_col] = df_BDD_DORA_vierge[nom_col].iloc[0]
    dict_config_col_BDD_DORA_vierge['type_col'] = dict_type_col_BDD_DORA

    dict_nb_chiffres_col_BDD_DORA = {}
    for nom_col in list(df_BDD_DORA_vierge):
        dict_nb_chiffres_col_BDD_DORA[nom_col] = df_BDD_DORA_vierge[nom_col].iloc[1]    
    dict_config_col_BDD_DORA_vierge['nb_chiffre'] = dict_nb_chiffres_col_BDD_DORA
    
    return dict_config_col_BDD_DORA_vierge

#####################################################################################################
#Continuité
#####################################################################################################
def import_tableau_attributs_Osmose_continuite():
    df_attributs_Osmose = import_tableau_attributs_Osmose()
    df_attributs_Osmose = df_attributs_Osmose.loc[(~df_attributs_Osmose["Obligatoire à partir"].isnull())|
                                                  (df_attributs_Osmose["Attribut"]=="Code ROE")|
                                                  (df_attributs_Osmose["Attribut"]=="Action concernant un ROE")|
                                                  (df_attributs_Osmose["Attribut"]=="Solution technique retenue")]
    df_attributs_Osmose["chiffre_obligatoire_a_partir"] = df_attributs_Osmose["Obligatoire à partir"].map({"Engagée":4})
    return df_attributs_Osmose

def import_tableau_etapes_Osmose():
    filename = ("config\\Osmose\\Configuration_Etapes_Type_action.csv")
    df_config = connect_path.conv_s3_obj_vers_python_obj("config",filename)     
    df_etapes_Osmose = pd.read_csv(filename)
    return df_etapes_Osmose

def import_tableau_etapes_Osmose_continuite():
    df_etapes_Osmose = import_tableau_etapes_Osmose()
    df_etapes_Osmose = df_etapes_Osmose[[x for x in list(df_etapes_Osmose) if not x.startswith("Unnamed")]]
    df_etapes_Osmose["Type d'action"] = [x[:7] for x in df_etapes_Osmose["Type d'action"].to_list()]
    df_etapes_Osmose = df_etapes_Osmose.rename({"Type d'action":"CODE_TYPE_ACTION_OSMOSE","Ordre N°":"Num étape","Intitulé de l'étape":"Avancement_spe_conti","Niveau d'avancement":"Avancement"},axis=1)
    return df_etapes_Osmose

#####################################################################################################
#Osmose
#####################################################################################################
def import_df_vierge_Osmose_onglet_action():
    filename = ("config\\Osmose\\df_actions_Osmose_vierge.csv")
    df_config = connect_path.conv_s3_obj_vers_python_obj("config",filename)
    df_vierge_Osmose_onglet_action = pd.read_csv(filename)
    return df_vierge_Osmose_onglet_action

def import_dict_renommage_df_Osmose_final():
    df_vierge_Osmose_onglet_action = import_df_vierge_Osmose_onglet_action()
    dict_tempo = {}
    for nom_col in list(df_vierge_Osmose_onglet_action):
        if df_vierge_Osmose_onglet_action[nom_col].iloc[0]==df_vierge_Osmose_onglet_action[nom_col].iloc[0]:
            dict_tempo[nom_col] = df_vierge_Osmose_onglet_action[nom_col].iloc[0]
    dict_renommage_df_Osmose_final = {v: k for k, v in dict_tempo.items()}      
    return dict_renommage_df_Osmose_final


def recuperation_excel_vierge_Osmose():
    filename = ("config\\Osmose\\TABLEAU_CREATION_OSMOSE_VIERGE.xlsx")
    filename = connect_path.conv_s3_obj_vers_python_obj("config",filename)   
    workbook = load_workbook(filename, read_only=False)
    return workbook

def recuperation_excel_MIA_MO_vierge_DORA():
    filename = ("config\\DORA\\Tableau_suivi_MIA_MO_vierge.xlsx")
    filename = connect_path.conv_s3_obj_vers_python_obj("config",filename)
    data = io.BytesIO(filename)
    workbook = load_workbook(data, read_only=False)
    return workbook

def recuperation_df_info_PPG_vierge_DORA():
    filename = ("config\\DORA\\Tableau_suivi_MIA_MO_vierge.xlsx")
    filename = connect_path.conv_s3_obj_vers_python_obj("config",filename)     
    df_info_PPG_vierge = pd.read_excel(filename, sheet_name="info PPG")
    return df_info_PPG_vierge

def recuperation_excel_MIA_DEP_vierge_DORA():
    filename = ("config\\DORA\\Tableau_suivi_MIA_DEP_vierge.xlsx")
    filename = connect_path.conv_s3_obj_vers_python_obj("config",filename) 
    workbook = load_workbook(filename, read_only=False)
    return workbook

def import_onglet_vierge_blocage_Osmose():
    filename = ("config\\Osmose\\tableau import onglet blocage.csv")
    filename = connect_path.conv_s3_obj_vers_python_obj("config",filename) 
    df_vierge_Osmose_onglet_blocage = pd.read_csv(filename)
    return df_vierge_Osmose_onglet_blocage

def import_onglet_vierge_etapes_Osmose():
    filename = ("config\\Osmose\\tableau import onglet etapes.csv")
    filename = connect_path.conv_s3_obj_vers_python_obj("config",filename) 
    df_vierge_Osmose_onglet_etapes = pd.read_csv(filename)
    return df_vierge_Osmose_onglet_etapes

def import_onglet_vierge_financeurs_Osmose():
    filename = ("config\\Osmose\\tableau import onglet financeurs.csv")
    filename = connect_path.conv_s3_obj_vers_python_obj("config",filename) 
    df_vierge_Osmose_onglet_financeurs = pd.read_csv(filename)
    return df_vierge_Osmose_onglet_financeurs

def recuperation_df_info_complementaire_info_CODE_Osmose():
    filename = ("config\\Osmose\\Osmose_code_action_sous_domaine.csv")
    filename = connect_path.conv_s3_obj_vers_python_obj("config",filename) 
    df_info_complementaire_info_CODE_Osmose = pd.read_csv(filename)
    liste_osmose = []
    liste_osmose2 = []
    liste_nomenclature = []
    liste_sous_domaine = []
    liste_domaine = []
    liste_osmose = df_info_complementaire_info_CODE_Osmose["Code type d'action OSMOSE*"].to_list()
    liste_nomenclature = [x[-7:] for x in liste_osmose]
    df_info_complementaire_info_CODE_Osmose['CODE_TYPE_ACTION_OSMOSE']=liste_nomenclature
    liste_osmose2 = [".".join(x.split(".", 2)[:2]) for x in liste_osmose]
    liste_nomenclature = [x[:5] for x in liste_nomenclature]
    liste_sous_domaine = list(zip(liste_osmose2,liste_nomenclature))
    liste_sous_domaine = ["-".join(x) for x in liste_sous_domaine]
    df_info_complementaire_info_CODE_Osmose["code sous domaine OSMOSE*"]=liste_sous_domaine
    liste_nomenclature = [x[:3] for x in liste_nomenclature]
    liste_osmose2 = [x[0] for x in liste_osmose2]
    liste_domaine = list(zip(liste_osmose2,liste_nomenclature))
    liste_domaine = ["-".join(x) for x in liste_domaine]
    df_info_complementaire_info_CODE_Osmose['Code domaine*']=liste_domaine
    return df_info_complementaire_info_CODE_Osmose

def import_tableau_dict_renommage_conv_maj_Osmose_basique():
    filename = ("config\\Osmose\\dict_changement_nom_et_type_colonne_tableau_Osmose.csv")
    df_config = connect_path.conv_s3_obj_vers_python_obj("config",filename) 
    df_Osmose_vierge = pd.read_csv(filename)
    dict_tempo = {}
    for nom_col in list(df_Osmose_vierge):
        if df_Osmose_vierge[nom_col].iloc[0]==df_Osmose_vierge[nom_col].iloc[0]:
            dict_tempo[nom_col] = df_Osmose_vierge[nom_col].iloc[0]
    dict_renommage_conv_maj_Osmose_basique = {k: v for k, v in dict_tempo.items()}  
    return dict_renommage_conv_maj_Osmose_basique

def import_tableau_attributs_Osmose():
    filename = ("config\\Osmose\\Configuration_Attributs_Type_action.csv")
    df_config = connect_path.conv_s3_obj_vers_python_obj("config",filename) 
    df_attributs_Osmose = pd.read_csv(filename)
    df_attributs_Osmose["Type d'action"] = [x[:7] for x in df_attributs_Osmose["Type d'action"].to_list()]
    dict_map_avancement_obligatoire_chiffre = {"Engagée":4}
    df_attributs_Osmose["obligatoire_chiffre_avancement"] = df_attributs_Osmose["Obligatoire à partir"].map(dict_map_avancement_obligatoire_chiffre)
    df_attributs_Osmose = df_attributs_Osmose.rename({"Type d'action":"CODE_TYPE_ACTION_OSMOSE"},axis=1)
    df_attributs_Osmose = df_attributs_Osmose[[x for x in list(df_attributs_Osmose) if x!="Valeurs"]]
    return df_attributs_Osmose

def import_dict_num_dep_nom_dep():
    filename = ("G:\\travail\\carto\\couches de bases\\dep\\LIEN numero dep to dep.csv")
    df_config = connect_path.conv_s3_obj_vers_python_obj("config",filename) 
    df_num_dep_nom_dep = pd.read_csv(filename)
    dict_num_dep_nom_dep = dict(zip(df_num_dep_nom_dep['Département Pilote PAOT'].to_list(),df_num_dep_nom_dep['Intitulé du service pilote*'].to_list()))
    return dict_num_dep_nom_dep

def import_dict_TYPE_MO_type_MO_Osmose():
    filename = ("config\\Osmose\\conversion_TYPE_MO_type_MO_Osmose.csv")
    df_config = connect_path.conv_s3_obj_vers_python_obj("config",filename) 
    df_num_dep_nom_dep = pd.read_csv(filename)
    dict_TYPE_MO_type_MO_Osmose = dict(zip(df_num_dep_nom_dep["TYPE_MO"].to_list(),df_num_dep_nom_dep["TYPE_MO_Osmose"].to_list()))
    return dict_TYPE_MO_type_MO_Osmose

def import_dict_config_actions_MIA_Osmose():
    dict_config_DORA_actions_MIA = {}
    filename = ("config\\Osmose\\config_liste_action_Osmose.ods")
    df_config = connect_path.conv_s3_obj_vers_python_obj("config",filename) 
    df_config_DORA = pd.read_excel(filename, engine="odf", header=[0, 1])
    df_config_DORA.columns=pd.MultiIndex.from_tuples([{('commun', 'Type_action'):('commun', 'CODE_TYPE_ACTION_OSMOSE')}.get(x, x) for x in df_config_DORA.columns])
    df_config_DORA = df_config_DORA.rename(columns=lambda x: x.strip())

    #Elu
    df_DORA_actions_MIA_elu = df_config_DORA.iloc[:, (df_config_DORA.columns.get_level_values(1)=="CODE_TYPE_ACTION_OSMOSE")|(df_config_DORA.columns.get_level_values(0)=='Elu')]
    col_df_elu = [b for (a,b) in list(df_DORA_actions_MIA_elu)]
    df_DORA_actions_MIA_elu.columns = col_df_elu
    df_DORA_actions_MIA_elu.loc[~df_DORA_actions_MIA_elu["CODE_TYPE_ACTION_OSMOSE"].isnull()]
    df_DORA_actions_MIA_elu["CODE_TYPE_ACTION_OSMOSE"] = df_DORA_actions_MIA_elu["CODE_TYPE_ACTION_OSMOSE"].apply(lambda x:x.split("-")[0])

    #Tech
    df_DORA_actions_MIA_tech = df_config_DORA.iloc[:, (df_config_DORA.columns.get_level_values(1)=="CODE_TYPE_ACTION_OSMOSE")|(df_config_DORA.columns.get_level_values(0)=='Tech')]
    col_df_tech = [b for (a,b) in list(df_DORA_actions_MIA_tech)]
    df_DORA_actions_MIA_tech.columns = col_df_tech
    df_DORA_actions_MIA_tech.loc[~df_DORA_actions_MIA_tech["CODE_TYPE_ACTION_OSMOSE"].isnull()]
    df_DORA_actions_MIA_tech["CODE_TYPE_ACTION_OSMOSE"] = df_DORA_actions_MIA_tech["CODE_TYPE_ACTION_OSMOSE"].apply(lambda x:x.split("-")[0])

    dict_config_DORA_actions_MIA["df_actions_MIA_elu"] = df_DORA_actions_MIA_elu
    dict_config_DORA_actions_MIA["df_actions_MIA_tech"] = df_DORA_actions_MIA_tech
    
    return dict_config_DORA_actions_MIA




