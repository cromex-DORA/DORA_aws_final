# -*- coding: utf-8 -*-
import pandas as pd
import geopandas as gpd
import glob
import os
import geopandas as gpd
from app.DORApy.classes.modules import dataframe
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

from app.DORApy.classes.modules import config_DORA

def ajustement_taille_col(ws):
    for col in ws.columns:
         max_length = 0
         column = col[0].column_letter # Get the column name
         for cell in col:
             try: # Necessary to avoid error on empty cells
                 if (str(cell.value)) > max_length:
                     max_length = (str(cell.value))
             except:
                 pass
         adjusted_width = (max_length + 2) * 1.2
         ws.column_dimensions[column].width = adjusted_width
    return ws


def ajout_onglet_AIDE_liste_ME(excel_modif,CODE_custom,dict_relation_shp_liste,dict_dict_info_REF):
    ws = excel_modif['AIDE Liste ME']
    #Pour le tableau, on prend uniquement les pressions hydromorphologies
    liste_CODE_ME_par_custom = dict_relation_shp_liste['dict_liste_ME_par_custom'][CODE_custom]
    df_info_ME_boosted = dict_dict_info_REF['df_info_ME'].loc[dict_dict_info_REF['df_info_ME']['CODE_ME'].isin(liste_CODE_ME_par_custom)]
    for r in dataframe_to_rows(df_info_ME_boosted, index=False, header=True):
        ws.append(r)
    ws = ajustement_taille_col(ws)
    return excel_modif


def ajout_onglet_AIDE_liste_SME(excel_modif,CODE_custom,dict_relation_shp_liste,dict_dict_info_REF):
    ws = excel_modif['AIDE Liste SME']
    liste_CODE_SME_par_custom = dict_relation_shp_liste['dict_liste_SME_par_custom'][CODE_custom]
    df_info_SME_boosted = dict_dict_info_REF['df_info_SME'].loc[dict_dict_info_REF['df_info_SME']['CODE_SME'].isin(liste_CODE_SME_par_custom)]
    for r in dataframe_to_rows(df_info_SME_boosted, index=False, header=True):
        ws.append(r)
    ws = ajustement_taille_col(ws)
    return excel_modif

def ajout_onglet_AIDE_liste_REF(projet,excel_modif,nom_custom,dict_relation_shp_liste,dict_dict_info_REF,contenu_custom):
    def ajout_MO_generique_dep(self,CODE_custom):
        numero_dep = CODE_custom
        df_NOM_MO_generique = dataframe.recuperation_df_liste_MO_generique()
        liste_MO_dep = df_NOM_MO_generique['MO_dep'].to_list()
        liste_MO_generique_dep = [x + ' ' + numero_dep for x in liste_MO_dep]
        liste_MO_generique_global = df_NOM_MO_generique['MO_global'].to_list()
        df_generique_dep = pd.DataFrame(liste_MO_generique_dep,columns =['NOM_MO'])
        df_generique_global = pd.DataFrame(liste_MO_generique_global,columns =['NOM_MO'])
        self = pd.concat([self,df_generique_dep,df_generique_global])
        return self
    liste_REF = contenu_custom.liste_REF
    liste_REF.append('ME')
    liste_REF.append(contenu_custom.echelle_base_REF)
    liste_REF = list(set(liste_REF))
    for REF in liste_REF:
        ws = excel_modif['AIDE Liste ' + REF]
        if "dict_liste_" + REF + "_par_custom" in dict_relation_shp_liste:
            if nom_custom in dict_relation_shp_liste["dict_liste_" + REF + "_par_custom"]:
                liste_REF = dict_relation_shp_liste["dict_liste_" + REF + "_par_custom"][nom_custom]
                df_info_tempo_REF = dict_dict_info_REF['df_info_'+REF]
                df_info_tempo_REF = df_info_tempo_REF.loc[df_info_tempo_REF['CODE_'+REF].isin(liste_REF)]
                for column in list(df_info_tempo_REF):
                    df_info_tempo_REF[column] = df_info_tempo_REF[column].apply(lambda x: ','.join(map(str, x)) if isinstance(x,list) else x)
                col_CODE_ME = df_info_tempo_REF.pop("NOM_"+REF)
                df_info_tempo_REF.insert(0,'NOM_'+REF,col_CODE_ME)
            if nom_custom not in dict_relation_shp_liste["dict_liste_" + REF + "_par_custom"]:
                df_info_tempo_REF = pd.DataFrame([])
        if "dict_liste_" + REF + "_par_custom" not in dict_relation_shp_liste:
            df_info_tempo_REF = pd.DataFrame([])
        if REF=="MO" and contenu_custom.echelle_REF=="DEP":
            df_info_tempo_REF = ajout_MO_generique_dep(df_info_tempo_REF,contenu_custom.CODE_custom)
        #Menage sur les colonnes !
        if REF=="MO":
            df_info_tempo_REF = df_info_tempo_REF[["NOM_MO","CODE_MO","TYPE_MO","CODE_SIRET","CODE_SIRET_SANDRE"]]
        if REF=="BVG":
            df_info_tempo_REF = df_info_tempo_REF[["NOM_BVG","CODE_BVG"]]            
        for r in dataframe_to_rows(df_info_tempo_REF, index=False, header=True):
            ws.append(r)
    return excel_modif


def ajout_onglet_info_PPG(excel_modif,nom_custom,dict_relation_shp_liste,dict_dict_info_REF):
    ws = excel_modif['info PPG']
    if "dict_liste_PPG_par_MO" in dict_relation_shp_liste:
        if nom_custom in dict_relation_shp_liste["dict_liste_PPG_par_MO"]:
            df_info_PPG_custom = dict_dict_info_REF["df_info_PPG"].loc[dict_dict_info_REF["df_info_PPG"]["CODE_PPG"].isin(dict_relation_shp_liste["dict_liste_PPG_par_MO"][nom_custom])]
            onglet_info_PPG_vierge = config_DORA.recuperation_df_info_PPG_vierge_DORA()
            dict_renommage = {'NOM_MO_gemapi':list(onglet_info_PPG_vierge)[0],"CODE_SIRET":"code SIRET","NOM_PPG":"Nom_PPG","CODE_PPG":'CODE_PPG',"debut_PPG":"Année Début_PPG","fin_PPG":"Année_Fin_PPG","debut_DIG":"Année Début_DIG","fin_DIG":"Année_Fin_DIG"}
            df_info_PPG_custom = df_info_PPG_custom.rename(dict_renommage,axis=1)
            df_info_PPG_custom = df_info_PPG_custom[list(dict_renommage.values())]
            for colonne_onglet_vierge in list(onglet_info_PPG_vierge):
                if colonne_onglet_vierge not in dict_renommage.values():
                    df_info_PPG_custom[colonne_onglet_vierge] = ""
            df_info_PPG_custom = df_info_PPG_custom[list(onglet_info_PPG_vierge)]

            for r in dataframe_to_rows(df_info_PPG_custom, index=False, header=False):
                ws.append(r)          
    return excel_modif

def ajout_onglet_Lien_ROE_CODE_ROE(excel_modif,nom_custom,dict_relation_shp_liste,dict_dict_info_REF):
    ws = excel_modif['Lien_ROE_CODE_ROE']
    dict_CODE_ROE_NOM_ROE =  dict_dict_info_REF["df_info_ROE"].dict_CODE_NOM
    
    if "dict_liste_ROE_par_custom" in dict_relation_shp_liste:
        if nom_custom in dict_relation_shp_liste["dict_liste_ROE_par_custom"]:
            dict_ROE_par_custom = {CODE_ME:dict_CODE_ROE_NOM_ROE[CODE_ME] for CODE_ME in dict_relation_shp_liste["dict_liste_ROE_par_custom"][nom_custom]}
            df_NOM_ROE_CODE_ROE = pd.DataFrame(data=dict_ROE_par_custom, index=["NOM_ROE","CODE_ROE"])
            df_NOM_ROE_CODE_ROE = df_NOM_ROE_CODE_ROE.T
            df_NOM_ROE_CODE_ROE['CODE_ROE'] = df_NOM_ROE_CODE_ROE.index
            for r in dataframe_to_rows(df_NOM_ROE_CODE_ROE, index=False, header=True):
                ws.append(r)
    return excel_modif

def ajout_onglet_Lien_REF_ME(projet,excel_modif,CODE_custom,dict_relation_shp_liste,dict_dict_info_REF,contenu_custom):
    def ajout_liste_MO_generique_avec_liste_toutes_ME_dep(self,contenu_custom,dict_relation_shp_liste):
        numero_dep = contenu_custom.CODE_custom
        df_NOM_MO_generique = dataframe.recuperation_df_liste_MO_generique()
        liste_MO_dep = df_NOM_MO_generique['MO_dep'].to_list()
        liste_MO_generique_dep = [x + ' ' + numero_dep for x in liste_MO_dep]
        liste_MO_generique_global = df_NOM_MO_generique['MO_global'].to_list()
        liste_toute_ME_dep = dict_relation_shp_liste['dict_liste_ME_par_custom'][numero_dep]
        for MO_generique in liste_MO_generique_dep:
            self[MO_generique] = liste_toute_ME_dep
        for MO_generique in liste_MO_generique_global:
            self[MO_generique] = liste_toute_ME_dep
        return self

    if contenu_custom.echelle_base_REF=="ME":
        liste_echelle_base = ["ME"]
    if contenu_custom.echelle_base_REF=="SME":
        liste_echelle_base = ["ME","SME"]
    
    liste_onglet_pour_lien = [x for x in excel_modif.sheetnames if x.startswith("Pour lien ")]
    if contenu_custom.echelle_base_REF =="ME":
        liste_onglet_pour_lien = [x for x in liste_onglet_pour_lien if not "SME" in x]
    for onglet in liste_onglet_pour_lien:

        ws = excel_modif[onglet]
        REF1 = onglet.split(" ")[-2]
        REF2 = onglet.split(" ")[-1]
        dict_CODE_NOM_REF1 = dict(zip(dict_dict_info_REF['df_info_'+REF1]['CODE_'+REF1].to_list(),dict_dict_info_REF['df_info_'+REF1]['NOM_'+REF1].to_list()))
        dict_CODE_NOM_REF2 = dict(zip(dict_dict_info_REF['df_info_'+REF2]['CODE_'+REF2].to_list(),dict_dict_info_REF['df_info_'+REF2]['NOM_'+REF2].to_list()))
        
        dict_CODE_REF_list_CODE_REF_base = dict_relation_shp_liste["dict_liste_" + REF2  + "_par_" + REF1]
        dict_CODE_REF_list_CODE_REF_base = {k:v for k,v in dict_CODE_REF_list_CODE_REF_base.items() if k in dict_relation_shp_liste["dict_liste_" + REF1  + "_par_custom"][CODE_custom]}
        dict_CODE_REF_list_CODE_REF_base = {dict_CODE_NOM_REF1[k]:v for k,v in dict_CODE_REF_list_CODE_REF_base.items()}
        if onglet=="Pour lien MO ME":
            dict_CODE_REF_list_CODE_REF_base = ajout_liste_MO_generique_avec_liste_toutes_ME_dep(dict_CODE_REF_list_CODE_REF_base,contenu_custom,dict_relation_shp_liste)
        dict_CODE_REF_list_CODE_REF_base = {k:[dict_CODE_NOM_REF2[x] for x in v] for k,v in dict_CODE_REF_list_CODE_REF_base.items()}
        
        df_ME_par_REF_par_custom=pd.DataFrame.from_dict(dict_CODE_REF_list_CODE_REF_base,orient='index')
        df_ME_par_REF_par_custom['NOM_'+REF1] = df_ME_par_REF_par_custom.index
        col_NOM_REF = df_ME_par_REF_par_custom.pop('NOM_'+REF1)
        df_ME_par_REF_par_custom.insert(0,'NOM_'+REF1,col_NOM_REF)
        df_ME_par_REF_par_custom = df_ME_par_REF_par_custom.sort_values('NOM_'+REF1)
        for r in dataframe_to_rows(df_ME_par_REF_par_custom, index=False, header=True):
            ws.append(r)

    return excel_modif


def ajout_onglet_Lien_REF_SME(projet,excel_modif,CODE_custom,dict_relation_shp_liste,dict_dict_info_REF):
    dict_CODE_NOM_SME = dict(zip(dict_dict_info_REF['df_info_SME']['CODE_SME'].to_list(),dict_dict_info_REF['df_info_SME']['NOM_SME'].to_list()))
    if projet.type_rendu=='tableau_vierge' and projet.type_donnees=='action' and projet.public_cible=='MO':
        liste_REF = ['PPG','ROE']
    for REF in liste_REF:
        if REF == 'PPG':
            ws = excel_modif['Pour lien ' + REF + ' SME']
        if REF == 'ROE':
            ws = excel_modif['Pour lien SME ' + REF]        

        dict_CODE_REF_list_CODE_SME = {}
        if REF=='PPG':
            if "dict_liste_" + REF + "_par_custom" in dict_relation_shp_liste:
                dict_CODE_REF_list_CODE_SME = dict_relation_shp_liste["dict_liste_SME_par_" + REF]
                dict_CODE_REF_list_CODE_SME = {dict_dict_info_REF['df_info_'+REF].dict_CODE_NOM[k]:v for k,v in dict_CODE_REF_list_CODE_SME.items()}
                dict_NOM_REF_list_CODE_SME = {k:[dict_CODE_NOM_SME[x] for x in v] for k,v in dict_CODE_REF_list_CODE_SME.items()}
                nom_col_maitre = REF
        if REF=='ROE':
            dict_CODE_REF_list_CODE_SME = dict_relation_shp_liste["dict_liste_" + REF + "_par_SME"]
            dict_CODE_REF_list_CODE_SME = {dict_CODE_NOM_SME[k]:v for k,v in dict_CODE_REF_list_CODE_SME.items()}
            dict_NOM_REF_list_CODE_SME = {k:[dict_dict_info_REF['df_info_'+REF].dict_CODE_NOM[x] for x in v] for k,v in dict_CODE_REF_list_CODE_SME.items()}
            nom_col_maitre = "SME"
            
        df_SME_par_REF_par_custom=pd.DataFrame.from_dict(dict_NOM_REF_list_CODE_SME,orient='index')
        df_SME_par_REF_par_custom['NOM_'+nom_col_maitre] = df_SME_par_REF_par_custom.index
        col_NOM_REF = df_SME_par_REF_par_custom.pop('NOM_'+nom_col_maitre)
        df_SME_par_REF_par_custom.insert(0,'NOM_'+nom_col_maitre,col_NOM_REF)
        df_SME_par_REF_par_custom = df_SME_par_REF_par_custom.sort_values('NOM_'+nom_col_maitre)
        for r in dataframe_to_rows(df_SME_par_REF_par_custom, index=False, header=True):
            ws.append(r)

    return excel_modif

def ajout_onglet_Lien_SME_ROE(excel_modif,nom_custom,dict_relation_shp_liste,dict_dict_info_REF,dict_CODE_REF_NOM_REF):
    #Lien ME ROE
    ws = excel_modif['Pour lien SME ROE']
    liste_ME = dict_relation_shp_liste["dict_liste_SME_par_custom"][nom_custom]
    dict_CODE_SME_list_CODE_ROE = {k:v for k,v in dict_relation_shp_liste["dict_liste_ROE_par_SME"].items() if k in liste_ME}
    dict_CODE_SME_list_CODE_ROE = {k:sorted(v) for k,v in dict_CODE_SME_list_CODE_ROE.items()}
    df_ROE_par_SME_par_custom=pd.DataFrame.from_dict(dict_CODE_SME_list_CODE_ROE,orient='index')
    df_ROE_par_SME_par_custom['CODE_SME'] = df_ROE_par_SME_par_custom.index
    df_ROE_par_SME_par_custom['CODE_SME'] = df_ROE_par_SME_par_custom['CODE_SME']
    col_CODE_SME = df_ROE_par_SME_par_custom.pop('CODE_SME')
    df_ROE_par_SME_par_custom.insert(0,'CODE_SME',col_CODE_SME)
    df_ROE_par_SME_par_custom = df_ROE_par_SME_par_custom.replace(dict_CODE_REF_NOM_REF['ROE'])
    for r in dataframe_to_rows(df_ROE_par_SME_par_custom, index=False, header=True):
        ws.append(r)
    return excel_modif

def export_fichier_excel_recherche_SANDRE(df_recherche,BDD_SIRET,df_info_GEMAPI_sans_CODE_SIRET):
    excel_modif = load_workbook("/mnt/g/travail/carto/couches de bases/syndicats GEMAPI/recherche_SANDRE.xlsx", read_only=False)
    '''excel_modif.save("/mnt/g/travail/carto/couches de bases/syndicats GEMAPI/recherche_SANDRE_0.xlsx")
    excel_modif = load_workbook("/mnt/g/travail/carto/couches de bases/syndicats GEMAPI/recherche_SANDRE_0.xlsx", read_only=False)'''
    ws = excel_modif['BDD_SIRET']
    for r in dataframe_to_rows(BDD_SIRET, index=False, header=True):
        ws.append(r)
    ws = excel_modif['Pour lien MO prop']
    for r in dataframe_to_rows(df_recherche, index=False, header=True):
        ws.append(r)
    ws = excel_modif['DF_GEMAPI_original']
    for r in dataframe_to_rows(df_info_GEMAPI_sans_CODE_SIRET[['NOM_MO','CODE_SIRET']], index=False, header=False):
        ws.append(r)
    print("allo")
    excel_modif.save("/mnt/g/travail/carto/couches de bases/syndicats GEMAPI/recherche_SANDRE_0.xlsx")
        
    

##############################
#MAJ Osmose
##############################
def ajout_onglet_actions(excel_modif,df):
    ws = excel_modif['Actions']
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)
    return excel_modif

def ajout_onglet_attributs(excel_modif,df):
    ws = excel_modif['+Attributs']
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)
    return excel_modif

