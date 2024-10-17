import pandas as pd
import geopandas as gpd
from app.DORApy.classes.modules import dataframe,BDD
from os import path
import numpy as np
from datetime import datetime
import sys
import re
import copy
from itertools import compress
import glob
import os
from openpyxl import load_workbook
from app.DORApy.classes.modules import connect_path
from app.DORApy.classes.Class_DfTableauxActionsMIA import DfTableauxActionsMIA,DfBDDOsmose
from app.DORApy.classes.Class_dictGdfCompletREF import dictGdfCompletREF
from app.DORApy.classes import Class_Folder
from app.DORApy.classes.modules import config_DORA


bucket_files_common = os.getenv('S3_BUCKET_COMMON_FILES')
bucket_files_CUSTOM = os.getenv('S3_BUCKET_USERS_FILES')
dict_config_espace,dict_config_police = config_DORA.creation_dicts_config()
dict_dict_config_df_actions_MIA = config_DORA.creation_dict_dict_config_df_actions_MIA()

##########################################################################################
#DictDFTableauxActionsMIA
##########################################################################################
class DictDFTableauxActionsMIA(dict):
    @property
    def _constructor(self):
        return DictDFTableauxActionsMIA
    
    def recuperation_dict_tableaux_actions_MIA(self):
        dict_dict_df_actions_originaux ={}
        for nom_CUSTOM,dict_CUSTOM in self.items():
            chemin_dossier = Class_Folder.get_path_dossier_CUSTOM(dict_CUSTOM.echelle_REF)
            CODE_REF = dict_CUSTOM.CODE_CUSTOM
            if self.type_rendu == "tableau_DORA_vers_BDD":
                nom_fichier = "tableau_proposition_" + dict_CUSTOM.NOM_CUSTOM + ".xlsx"
            if self.type_rendu == "verif_tableau_DORA":
                nom_fichier = "tableau_rempli_" + dict_CUSTOM.NOM_CUSTOM + ".xlsx"
            self.complet_path = connect_path.get_file_path_racine(os.path.join(chemin_dossier,CODE_REF,nom_fichier))
            self.relative_path = os.path.join(chemin_dossier,CODE_REF,nom_fichier)
            self.folder_path = os.path.join(chemin_dossier,CODE_REF)
            dict_dict_df_actions_originaux[dict_CUSTOM.NOM_CUSTOM] = DfTableauxActionsMIA(self.relative_path)
        return dict_dict_df_actions_originaux
        
    def recuperation_dict_tableaux_actions_MIA_bis(dict_dict_info_REF=None,self=None,forme="remplis"):
        #Obliger de passer par de l'extraction de nom de colonne d'un tableau vierge, car on ne pas écrire manuellement les noms des colonnes depuis un csv a cause des caractéres : ' ?
        if liste_echelle_CUSTOM==None:
            liste_echelle_CUSTOM = ['MO']

        liste_df_actions = []
        liste_nom_fichier = []
        liste_nom_CUSTOM = []
        liste_echelle = []
        liste_numero_dep = []
        liste_annee_remplissage = []
        liste_fichier_brut = []
        liste_fichier_excel_ou_ods = []
        liste_dict_nom_col = []
        liste_dict_type_col = []
        liste_dict_nb_col = []
        
        list_type_fichier = ['xlsx','xlsm','ods']
        
        for type_CUSTOM in liste_echelle_CUSTOM:
            chemin_fichier_tableau = "/mnt/g/travail/carto/projets basiques/PAOT global 5.0/Tableau actions/MIA/" + type_CUSTOM + "/tableaux_" + forme
            for type_fichier in list_type_fichier:
                if type_fichier == 'xlsm' or type_fichier == 'xlsx':
                    liste_fichier_excel_ou_ods = [load_workbook(filename=f) for f in glob.glob(chemin_fichier_tableau + "/*." + type_fichier)]
                    liste_df_actions_originaux = [pd.read_excel(f,sheet_name="tableau a remplir",skiprows = [0,1,2,4]) for f in glob.glob(chemin_fichier_tableau + "/*." + type_fichier)]
                    liste_nom_fichier_actions = [x["tableau a remplir"].cell(1,1).value for x in liste_fichier_excel_ou_ods]
                    liste_numero_dep_temp = [x["tableau a remplir"].cell(3,2).value for x in liste_fichier_excel_ou_ods]
                    liste_annee_temp = [x["tableau a remplir"].cell(3,1).value for x in liste_fichier_excel_ou_ods]
                    #Il faut fermer les fichier workbook sinon toutes les formules affichent des NaN
                    for fichier_excel in liste_fichier_excel_ou_ods:
                        fichier_excel.close()
                if type_fichier == 'ods':
                    liste_df_actions_originaux = [pd.read_excel(f,sheet_name="tableau a remplir",skiprows = [0,1,2,4], engine="odf") for f in glob.glob(chemin_fichier_tableau + "/*.ods")]
                    liste_nom_fichier_actions = [pd.read_excel(f,header=None,sheet_name="tableau a remplir", engine="odf").iat[0,0] for f in glob.glob(chemin_fichier_tableau + "/*.ods")]
                    liste_numero_dep_temp = [pd.read_excel(f,header=None,sheet_name="tableau a remplir", engine="odf").iat[2,1] for f in glob.glob(chemin_fichier_tableau + "/*.ods")]
                    liste_annee_temp = [pd.read_excel(f,header=None,sheet_name="tableau a remplir", engine="odf").iat[2,0] for f in glob.glob(chemin_fichier_tableau + "/*.ods")]
                list_bool = [True if x.startswith('Tableau DORA actions MIA ') else False for x in liste_nom_fichier_actions]
                liste_fichier_excel_ou_ods = list(compress(liste_fichier_excel_ou_ods, list_bool)) 
                liste_df_actions_originaux = list(compress(liste_df_actions_originaux, list_bool))
                liste_nom_fichier_actions = list(compress(liste_nom_fichier_actions, list_bool))
                liste_numero_dep_temp = list(compress(liste_numero_dep_temp, list_bool))
                liste_annee_temp = list(compress(liste_annee_temp, list_bool))
                
                liste_df_actions_originaux = [x for x in liste_df_actions_originaux if len(x)>0]
                if forme=="remplis":
                    liste_nom_CUSTOM_temp = [x.split("Tableau DORA actions MIA "+type_CUSTOM+" ")[-1] for x in liste_nom_fichier_actions]
                if forme=="final":
                    liste_nom_CUSTOM_temp = [x.split("Tableau DORA actions MIA "+type_CUSTOM+" final ")[-1] for x in liste_nom_fichier_actions]                    
                if type_CUSTOM!="DEP":
                    liste_numero_dep_temp = [re.findall(r'[0-9]{2}', x)[0] if len(re.findall(r'[0-9]{2}', x))>0 else 33 for x in liste_numero_dep_temp]
                if type_CUSTOM=="DEP":
                    liste_numero_dep_temp = [dict_dict_info_REF['df_info_DEP'].dict_NOM_CODE[x] for x in liste_nom_CUSTOM_temp]

                liste_annee_temp = [re.findall(r'[0-9]{4}', x)[0] if len(re.findall(r'[0-9]{4}', x))>0 else str(datetime.today().year) for x in liste_annee_temp]
                liste_echelle_temp = [type_CUSTOM for x in liste_nom_fichier_actions]
                liste_df_actions.extend(liste_df_actions_originaux)
                liste_nom_fichier.extend(liste_nom_fichier_actions)
                liste_nom_CUSTOM.extend(liste_nom_CUSTOM_temp)
                liste_echelle.extend(liste_echelle_temp)
                liste_numero_dep.extend(liste_numero_dep_temp)
                liste_annee_remplissage.extend(liste_annee_temp)
                liste_fichier_brut.extend(liste_fichier_excel_ou_ods)
        for echelle_CUSTOM in liste_echelle:
            liste_dict_nom_col.append(config_DORA.creation_dict_dict_config_df_actions_MIA()['dict_conv_nom_col_DORA_'+echelle_CUSTOM])
            liste_dict_type_col.append(config_DORA.creation_dict_dict_config_df_actions_MIA()['dict_conv_type_col_DORA_'+echelle_CUSTOM])
            liste_dict_nb_col.append(config_DORA.creation_dict_dict_config_df_actions_MIA()['dict_nb_chiffres_col_DORA_'+echelle_CUSTOM])

                    
        dict_dict_df_actions_originaux = {k:DictDFTableauxActionsMIA({}) for k in liste_nom_CUSTOM}
        for numero_tableau,nom_tableau in enumerate(liste_nom_fichier):
            dict_dict_df_actions_originaux[liste_nom_CUSTOM[numero_tableau]].df = DfTableauxActionsMIA(liste_df_actions[numero_tableau])
            dict_dict_df_actions_originaux[liste_nom_CUSTOM[numero_tableau]].echelle_df = liste_echelle[numero_tableau]
            echelle_df =dict_dict_df_actions_originaux[liste_nom_CUSTOM[numero_tableau]].echelle_df
            dict_dict_df_actions_originaux[liste_nom_CUSTOM[numero_tableau]].dict_nom_col = liste_dict_nom_col[numero_tableau]
            dict_dict_df_actions_originaux[liste_nom_CUSTOM[numero_tableau]].dict_type_col = liste_dict_type_col[numero_tableau]
            dict_dict_df_actions_originaux[liste_nom_CUSTOM[numero_tableau]].dict_nb_chiffres = liste_dict_nb_col[numero_tableau]
            dict_dict_df_actions_originaux[liste_nom_CUSTOM[numero_tableau]].numero_dep = liste_numero_dep[numero_tableau]
            dict_dict_df_actions_originaux[liste_nom_CUSTOM[numero_tableau]].annee_remplissage = liste_annee_remplissage[numero_tableau]
            dict_dict_df_actions_originaux[liste_nom_CUSTOM[numero_tableau]].fichier_brut = liste_fichier_brut[numero_tableau]
            
        if dict_dict_info_REF!=None:
            for nom_CUSTOM in list(dict_dict_df_actions_originaux):
                echelle_df = dict_dict_df_actions_originaux[nom_CUSTOM].echelle_df
                if nom_CUSTOM not in dict_dict_info_REF['df_info_'+echelle_df]['NOM_' + echelle_df].to_list():
                    print("J'ai trouvé le tableau " + nom_CUSTOM + ", mais il n'est pas dans ma liste de MO")
                    del dict_dict_df_actions_originaux[nom_CUSTOM]
                if nom_CUSTOM in dict_dict_info_REF['df_info_'+echelle_df]['NOM_' + echelle_df].to_list():
                    print("J'ai trouvé le tableau " + nom_CUSTOM + ", et il est dans ma liste de MO !")
            for nom_CUSTOM,contenu_CUSTOM in dict_dict_df_actions_originaux.items():
                echelle_df = dict_dict_df_actions_originaux[nom_CUSTOM].echelle_df
                contenu_CUSTOM.CODE_CUSTOM = dict_dict_info_REF['df_info_'+echelle_df].dict_NOM_CODE[nom_CUSTOM]

        if len(dict_dict_df_actions_originaux)>0:
            print("J'ai trouvé cette liste de tableau actions :")
            print (list(dict_dict_df_actions_originaux))

        if len(dict_dict_df_actions_originaux)==0:
            print("Pas de tableau action trouvé")

        #dict_dict_df_actions_originaux = DictDFTableauxActionsMIA.attr_dict_nom_type_nb_df(dict_dict_df_actions_originaux)
        return dict_dict_df_actions_originaux

    def ajout_BDD_Osmose(dict_dict_df_actions_originaux):
        BDD_Osmose = dataframe.recuperer_BDD_Osmose()
        if len(BDD_Osmose)>0:
            dict_dict_df_actions_originaux['BDD_Osmose'] = DictDFTableauxActionsMIA({})
            dict_dict_df_actions_originaux['BDD_Osmose'].df = BDD_Osmose
            dict_dict_df_actions_originaux['BDD_Osmose'].echelle_df = "global_Osmose" 
            dict_dict_df_actions_originaux['BDD_Osmose'].CODE_CUSTOM = "BDD_OSMOSE"
            dict_dict_df_actions_originaux['BDD_Osmose'].annee_remplissage = str(datetime.today().year)


        for type_df,dict_df in dict_dict_df_actions_originaux.items():
            if dict_df.echelle_df =="global_Osmose":
                dict_df.dict_nom_col = dict_dict_config_df_actions_MIA['dict_conv_nom_col_Osmose']
                dict_df.dict_type_col = dict_dict_config_df_actions_MIA['dict_conv_type_col_Osmose']
                dict_df.dict_nb_chiffres = dict_dict_config_df_actions_MIA['dict_nb_chiffres_col_Osmose']                   
        return dict_dict_df_actions_originaux


    def col_ID_DORA(self):
        for nom_tableau,dict_df_actions_originaux in self.items():
            dict_df_actions_originaux.df['ID_DORA'] = dict_df_actions_originaux.CODE_CUSTOM + "#" + str(dict_df_actions_originaux.annee_remplissage) + "#" + dict_df_actions_originaux.df.index.astype(str)
        return self

    def recuperer_attribut_echelle_base_REF(self,dict_dict_info_REF,dict_CUSTOM_maitre):
        for nom_tableau,dict_df_actions_originaux in self.items():
            if dict_df_actions_originaux.echelle_df =="DEP":
                dict_df_actions_originaux.echelle_base_REF = "ME"
            if dict_df_actions_originaux.echelle_df =="MO":
                if dict_CUSTOM_maitre.echelle_base_REF!=None:
                    dict_df_actions_originaux.echelle_base_REF = dict_CUSTOM_maitre.echelle_base_REF
                if dict_CUSTOM_maitre.echelle_base_REF==None:
                    df_info_SME = dict_dict_info_REF['df_info_SME']
                    liste_CODE_CUSTOM_avec_echelle_SME = list(set(df_info_SME['MO_maitre'].to_list()))
                    if dict_df_actions_originaux.CODE_CUSTOM in liste_CODE_CUSTOM_avec_echelle_SME:
                        dict_df_actions_originaux.echelle_base_REF = "SME"
                    if dict_df_actions_originaux.CODE_CUSTOM not in liste_CODE_CUSTOM_avec_echelle_SME:
                        dict_df_actions_originaux.echelle_base_REF = "ME"   
            if dict_df_actions_originaux.echelle_df =="global_Osmose":
                dict_df_actions_originaux.echelle_base_REF = "ME"
        return self

    def verif_forme_fichiers_actions_MIA(dict_dict_df_actions_originaux,dict_dict_info_REF,dict_log):
        dict_log = DfTableauxActionsMIA.verif_nom_fichier(dict_dict_df_actions_originaux,dict_dict_info_REF,dict_log)
        dict_log = DfTableauxActionsMIA.verif_nom_col_fichier(dict_dict_df_actions_originaux,dict_log)
        return dict_log

    def mise_en_forme_tableau_actions_MIA_sans_modif_contenu(dict_dict_df_actions_originaux,dict_dict_info_REF):
        for nom_tableau,dict_df_actions_originaux in dict_dict_df_actions_originaux.items():
            ###Les verifs
            dict_dict_df_actions_originaux[nom_tableau].df = DfTableauxActionsMIA.ajout_colonne_manquante_df_type_DORA(dict_dict_df_actions_originaux[nom_tableau].df)
            dict_dict_df_actions_originaux[nom_tableau].df = DfTableauxActionsMIA.traitement_BDD_OSMOSE_format_DORA(dict_dict_df_actions_originaux[nom_tableau].df,dict_dict_info_REF,dict_dict_df_actions_originaux[nom_tableau].echelle_df)
        return dict_dict_df_actions_originaux
        
    def mise_en_forme_format_DORA_tableau_actions_MIA_avec_modif_contenu(dict_dict_df_actions_originaux,dict_relation_shp_liste,dict_dict_info_REF=None,dict_CUSTOM_maitre=None):
        for nom_tableau,dict_df_actions_originaux in dict_dict_df_actions_originaux.items():
            echelle_df = dict_dict_df_actions_originaux[nom_tableau].echelle_df
            echelle_base_REF = dict_dict_df_actions_originaux[nom_tableau].echelle_base_REF
            dict_nom_col = dict_dict_df_actions_originaux[nom_tableau].dict_nom_col
            dict_type_col = dict_dict_df_actions_originaux[nom_tableau].dict_type_col
            dict_nb_chiffres = dict_dict_df_actions_originaux[nom_tableau].dict_nb_chiffres

            ###Suppression lignes vides
            #dict_dict_df_actions_originaux[nom_tableau].df = DfTableauxActionsMIA.changement_type_col_tableaux_actions_MIA(dict_dict_df_actions_originaux[nom_tableau].df,echelle_df,dict_dict_info_REF,dict_type_col,dict_nb_chiffres)
            dict_dict_df_actions_originaux[nom_tableau].df = DfTableauxActionsMIA.mise_en_forme_col_action_phare(dict_dict_df_actions_originaux[nom_tableau].df)
            dict_dict_df_actions_originaux[nom_tableau].df = DfTableauxActionsMIA.definition_col_localisation_principale(dict_dict_df_actions_originaux[nom_tableau].df,echelle_df,echelle_base_REF,dict_type_col)
            dict_dict_df_actions_originaux[nom_tableau].df = DfTableauxActionsMIA.gestion_type_col_localisation_principale(dict_dict_df_actions_originaux[nom_tableau].df,dict_dict_info_REF,dict_dict_df_actions_originaux[nom_tableau].CODE_CUSTOM)
            dict_dict_df_actions_originaux[nom_tableau].df = DfTableauxActionsMIA.ajout_colonne_CODE_REF_permanent(dict_dict_df_actions_originaux[nom_tableau].df,dict_relation_shp_liste,dict_dict_info_REF,dict_dict_df_actions_originaux[nom_tableau].echelle_df,dict_dict_df_actions_originaux[nom_tableau].CODE_CUSTOM)
            #dict_dict_df_actions_originaux[nom_tableau].df = DfTableauxActionsMIA.chgt_CODE_hors_REF_base(dict_dict_df_actions_originaux[nom_tableau].df,echelle_df,dict_type_col,dict_relation_shp_liste)
        return dict_dict_df_actions_originaux
    
    def rassemblement_df_toutes_sources_pour_BDD_DORA(self):
        dict_config_col_BDD_DORA_vierge = config_DORA.import_dict_config_col_BDD_DORA_vierge()
        for nom_tableau,dict_df_actions_originaux in self.items():
            dict_df_actions_originaux.df['origine_df'] = dict_df_actions_originaux.CODE_CUSTOM + "#" + dict_df_actions_originaux.annee_remplissage
        df_BDD_DORA = pd.concat([v.df for k,v in self.items() if len(v.df)>0])
        list_col_BDD_DORA = list(dict_config_col_BDD_DORA_vierge['type_col'])
        list_col_BDD_DORA.append('origine_df')
        df_BDD_DORA = df_BDD_DORA[list_col_BDD_DORA]
        df_BDD_DORA = df_BDD_DORA.reset_index(drop=True)
        #df_BDD_DORA = DfTableauxActionsMIA.mise_en_forme_BDD_DORA(df_BDD_DORA,dict_config_col_BDD_DORA_vierge)
        return df_BDD_DORA

    def mise_en_forme_specifique_BDD_DORA_pour_carte(BDD_DORA,dict_dict_info_REF=None,dict_relation_shp_liste=None,dict_CUSTOM_maitre=None):
        if dict_CUSTOM_maitre.type_rendu=='carte' and dict_CUSTOM_maitre.type_donnees == 'action' and dict_CUSTOM_maitre.thematique == 'MIA':
            list_col_CODE_echelle_projet = ["CODE_" + x for x in dict_CUSTOM_maitre.liste_echelle_REF_projet]
            list_col_NOM_echelle_projet = ["NOM_" + x for x in dict_CUSTOM_maitre.liste_echelle_REF_projet]
            list_col_a_garder = ["echelle_princ_action","NOM_TYPE_ACTION_DORA","CODE_TYPE_ACTION_OSMOSE","CODE_ROE","Avancement_spe_conti","trav_prevu","Avancement","annee_action_ini",
                                    "annee_action_eng","annee_action_term","action_aba","Action_phare","description_action_phare"]
            list_col_a_garder = list_col_a_garder + list_col_CODE_echelle_projet + list_col_NOM_echelle_projet
            list_col_a_garder = [x for x in list(list_col_a_garder) if x in list(BDD_DORA.df)]
            BDD_DORA.df = BDD_DORA.df[list_col_a_garder]
            BDD_DORA.df = DfTableauxActionsMIA.ajout_CODE_REF(BDD_DORA.df,dict_CUSTOM_maitre)
        return BDD_DORA

    def mise_en_forme_df_pressions_carte_pressions(df_pressions):
        list_col_Pression = [x for x in list(df_pressions) if x.startswith("P_")]
        df_pressions[list_col_Pression] = df_pressions[list_col_Pression].replace([2,1],0)
        df_pressions[list_col_Pression] = df_pressions[list_col_Pression].replace(3,1)
        df_pressions.loc[(df_pressions['P_MORPHO']==1) | (df_pressions['P_HYDRO']==1) | (df_pressions['P_CONTI']==1),['P_HYDRO_TO']] = 1
        df_pressions.loc[(df_pressions['P_IRRIG']==1) | (df_pressions['P_PREAEP']==1) | (df_pressions['P_PREIND']==1),['P_IRRIG_TO']] = 1
        df_pressions.loc[(df_pressions['P_INDMAC']==1) | (df_pressions['P_IND_SUB']==1) | (df_pressions['P_IND_SITE']==1),['P_IND_TOT']] = 1
        #A ENELEVER
        df_pressions[[x for x in list(df_pressions) if x!="CODE_ME"]] = df_pressions[[x for x in list(df_pressions) if x!="CODE_ME"]].astype('object')
        df_pressions['CODE_ME'] = df_pressions['CODE_ME'].astype(str)
        df_pressions.fillna('0',inplace=True)
        df_pressions = df_pressions.astype({"P_DOMPON":"int","P_IND_TOT":"int","P_AZOTE":"int","P_PHYTO":"int","P_IRRIG_TO":"int","P_HYDRO_TO":"int"})
        #df_pressions['NB_type_icone'] = df_pressions[['P_DOMPON','P_IND_TOT','P_AZOTE','P_PHYTO','P_IRRIG_TO','P_HYDRO_TO']].sum(axis=1)
        df_pressions = df_pressions[['CODE_ME','P_DOMPON','P_IND_TOT','P_AZOTE','P_PHYTO','P_IRRIG_TO','P_HYDRO_TO','P_MORPHO','P_HYDRO','P_CONTI']]
        df_pressions = df_pressions.rename({'P_DOMPON':'P_DOM','P_IND_TOT':'P_IND','P_AZOTE':'P_N2','P_PHYTO':'P_PHY','P_IRRIG_TO':'P_IRR','P_HYDRO_TO':'P_HYDT'},axis=1)
        df_pressions = df_pressions.rename({'P_MORPHO':'P_MORPHO','P_HYDRO':'P_HYDRO','P_CONTI':'P_CONTI'},axis=1)
        return df_pressions
    
    def mise_en_forme_CODE_REF(df_pressions):
        df_pressions = df_pressions.rename({'CODE_ME':'CODE_REF'},axis=1)
        return df_pressions


    def selection_actions_BDD_DORA(self,dict_entite_CUSTOM,dict_dict_info_REF,dict_decoupREF,dict_relation_shp_liste):
        BDD_DORA = self['BDD_DORA']
        list_hierarchie_shp_hydro = ['SAGE','MO','PPG','BVG','ME','SME']
        CODE_CUSTOM = dict_entite_CUSTOM.CODE_CUSTOM
        
        liste_echelle_boite_normal = dict_entite_CUSTOM["dict_boite_maitre_normal"].liste_echelle_REF
        liste_echelle_boite_normal = dictGdfCompletREF.hierarchisation_liste_echelle(liste_echelle_boite_normal)[::-1]

        echelle_boite_normal_la_plus_basse = liste_echelle_boite_normal[-1]
        if "dict_boite_maitre_ortho" in dict_entite_CUSTOM:
            liste_echelle_boite_ortho = dict_entite_CUSTOM["dict_boite_maitre_ortho"].liste_echelle_REF
            echelle_boite_ortho_la_plus_haute = liste_echelle_boite_ortho[0]
        if "dict_boite_maitre_ortho" not in dict_entite_CUSTOM:
            liste_echelle_boite_ortho = []

        liste_echelle_REF_total = liste_echelle_boite_normal + liste_echelle_boite_ortho[::-1]

        liste_CODE_REF_boite_normal = dict_entite_CUSTOM["dict_boite_maitre_normal"]['df_bloc_texte'].df["CODE_REF"].to_list()

        df_BDD_DORA_tempo = copy.deepcopy(BDD_DORA.df)
        liste_echelle_autre = list_hierarchie_shp_hydro[:list_hierarchie_shp_hydro.index(echelle_boite_normal_la_plus_basse)]
        liste_echelle_autre = [x for x in liste_echelle_autre if x not in liste_echelle_REF_total]
        #Cas special avec les SME, on transforme les ME en SME en choissant la plus grosse
        if "SME" in liste_echelle_boite_normal:
            df_info_SME = dict_dict_info_REF['df_info_SME']
            df_info_SME_tempo = df_info_SME.rename({"ME_maitre":"CODE_ME"},axis=1)
            df_surface_SME = dict_decoupREF['gdf_decoupSME_CUSTOM'][["CODE_SME","surface_SME"]]
            df_info_SME_tempo = pd.merge(df_info_SME_tempo,df_surface_SME,on='CODE_SME')
            df_info_SME_tempo = df_info_SME_tempo.sort_values(by="surface_SME",ascending=False)
            df_info_SME_tempo = df_info_SME_tempo.groupby('CODE_ME').first()
            dict_CODE_ME_CODE_SME_max_surface = dict(zip(list(df_info_SME_tempo.index.values),df_info_SME_tempo['CODE_SME'].to_list()))
            if BDD_DORA.dict_type_col['CODE_ME']=='str':
                df_BDD_DORA_tempo.loc[df_BDD_DORA_tempo["echelle_princ_action"]=="ME","CODE_REF"] = df_BDD_DORA_tempo["CODE_REF"].apply(lambda x:dict_CODE_ME_CODE_SME_max_surface[x] if x in dict_CODE_ME_CODE_SME_max_surface else x)
            if BDD_DORA.dict_type_col['CODE_ME']=='list':
                df_BDD_DORA_tempo.loc[df_BDD_DORA_tempo["echelle_princ_action"]=="ME","CODE_REF"] = df_BDD_DORA_tempo["CODE_REF"].apply(lambda x:[dict_CODE_ME_CODE_SME_max_surface[CODE_ME] if CODE_ME in dict_CODE_ME_CODE_SME_max_surface else CODE_ME for CODE_ME in x])
            df_BDD_DORA_tempo.loc[df_BDD_DORA_tempo["echelle_princ_action"]=="ME","echelle_princ_action"] = "SME"
            liste_echelle_autre = [x for x in list(liste_echelle_autre) if x!="ME"]

        if "ME" in liste_echelle_boite_normal:
            df_info_SME = dict_dict_info_REF['df_info_SME']
            df_info_SME_tempo = df_info_SME.rename({"ME_maitre":"CODE_ME"},axis=1)
            dict_CODE_SME_CODE_ME = dict(zip(df_info_SME_tempo['CODE_SME'].to_list(),df_info_SME_tempo['CODE_ME'].to_list()))
            if BDD_DORA.dict_type_col["CODE_ME"] =="str":
                df_BDD_DORA_tempo.loc[df_BDD_DORA_tempo["echelle_princ_action"]=="SME","CODE_REF"] = df_BDD_DORA_tempo["CODE_REF"].apply(lambda x:dict_CODE_SME_CODE_ME[x] if x in dict_CODE_SME_CODE_ME else x)
            if BDD_DORA.dict_type_col["CODE_ME"] =="list":
                df_BDD_DORA_tempo.loc[df_BDD_DORA_tempo["echelle_princ_action"]=="SME","CODE_REF"] = df_BDD_DORA_tempo["CODE_REF"].apply(lambda x:[dict_CODE_SME_CODE_ME[CODE_ME] if CODE_ME in dict_CODE_SME_CODE_ME else CODE_ME for CODE_ME in x])
            df_BDD_DORA_tempo.loc[df_BDD_DORA_tempo["echelle_princ_action"]=="SME","echelle_princ_action"] = "ME"
            liste_echelle_autre = [x for x in list(liste_echelle_autre) if x!="SME"]
        #Maintenant on check si les actions qui se situent dans liste_echelle_ortho sont présentes dans plus de 3 entites de bases
        
        #df_BDD_DORA_tempo.loc[df_BDD_DORA_tempo["echelle_princ_action"]==entite_a_tester,"CODE_REF"] = df_BDD_DORA_tempo.loc[df_BDD_DORA_tempo["echelle_princ_action"]==entite_a_tester]['CODE_REF'].apply(lambda x:separation_entite_normal_ou_ortho(x,liste_echelle_autre))
        if dict_entite_CUSTOM.echelle_REF == "MO":
            df_BDD_DORA_tempo= DfTableauxActionsMIA.filtre_par_CODE_MO_si_df_DORA_MO_specifique_au_CUSTOM(df_BDD_DORA_tempo,self,dict_entite_CUSTOM.CODE_CUSTOM)
            df_BDD_DORA_tempo= DfTableauxActionsMIA.separation_entite_normal_ou_ortho_echelle_CUSTOM(df_BDD_DORA_tempo,self,liste_echelle_boite_normal,dict_relation_shp_liste,dict_entite_CUSTOM.echelle_REF,dict_entite_CUSTOM.CODE_CUSTOM)
        if dict_entite_CUSTOM.echelle_REF == "DEP":
            df_BDD_DORA_tempo = DfTableauxActionsMIA.separation_entite_normal_ou_ortho_echelle_CUSTOM(df_BDD_DORA_tempo,self,liste_echelle_boite_normal,dict_relation_shp_liste,dict_entite_CUSTOM.echelle_REF,dict_entite_CUSTOM.CODE_CUSTOM)
        df_BDD_DORA_tempo = df_BDD_DORA_tempo.explode('CODE_REF')
        df_BDD_DORA_tempo = df_BDD_DORA_tempo.loc[~df_BDD_DORA_tempo['CODE_REF'].isnull()]
        return df_BDD_DORA_tempo


    def creation_df_pour_onglets(self):
        for nom_tableau,dict_df_actions_originaux in self.items():
            len_df = len(dict_df_actions_originaux.df)
            df_vierge = pd.DataFrame([x for x in range(0,len_df)],columns =['CODE_IMPORT'])
            dict_df_actions_originaux.df_Points_de_blocage = DfTableauxActionsMIA(df_vierge)
            dict_df_actions_originaux.df_Etapes = DfTableauxActionsMIA(df_vierge)
            dict_df_actions_originaux.df_Financeurs = DfTableauxActionsMIA(df_vierge)
            dict_df_actions_originaux.df_attributs = DfTableauxActionsMIA(df_vierge)
        return self
    
    def creation_CODE_IMPORT(self):
        for nom_CUSTOM,dict_CUSTOM in self.items():
            self[nom_CUSTOM].df['CODE_IMPORT'] = self[nom_CUSTOM].df.index
        return self

##########################################################################################
#Osmose inverse
##########################################################################################  
    def ajout_info_supplementaire_CODE_Osmose(self):
        df_info_complementaire_info_CODE_Osmose = config_DORA.recuperation_df_info_complementaire_info_CODE_Osmose()
        dict_config_actions_MIA_DORA = config_DORA.import_dict_config_actions_MIA_DORA()
        for nom_CUSTOM,dict_CUSTOM in self.items():
            self[nom_CUSTOM].df = self[nom_CUSTOM].df[[x for x in list(self[nom_CUSTOM].df) if x!="CODE_TYPE_ACTION_OSMOSE"]]
            self[nom_CUSTOM].df = pd.merge(self[nom_CUSTOM].df,dict_config_actions_MIA_DORA['df_DORA_actions_MIA_conv_Osmose'][["NOM_TYPE_ACTION_DORA","CODE_TYPE_ACTION_OSMOSE"]],on="NOM_TYPE_ACTION_DORA")
            self[nom_CUSTOM].df = pd.merge(self[nom_CUSTOM].df,df_info_complementaire_info_CODE_Osmose,on="CODE_TYPE_ACTION_OSMOSE",suffixes=[None,'_a_supprimer'])
            self[nom_CUSTOM].df = self[nom_CUSTOM].df.loc[:,~self[nom_CUSTOM].df.columns.str.endswith('_a_supprimer')]
        return self
    
    def conv_colonne_avancement(self):
        dict_conversion_avancement = {"Prévisionnelle":"2-Prévisionnelle","Initiée":"3-Initiée","Engagée":"4-Engagée","Terminée":"5-Terminée","Abandonnée":"6-Abandonnée"}
        for nom_CUSTOM,dict_CUSTOM in self.items():
            self[nom_CUSTOM].df["Niveau d'avancement*"]=self[nom_CUSTOM].df['Avancement'].map(dict_conversion_avancement)
            dict_conversion_avancement_chiffre = {"2-Prévisionnelle":2,"3-Initiée":3,"4-Engagée":4,"5-Terminée":5,"6-Abandonnée":6}
            self[nom_CUSTOM].df.loc[self[nom_CUSTOM].df["Niveau d'avancement*"].isnull(),"Niveau d'avancement*"] = "2-Prévisionnelle"
            self[nom_CUSTOM].df["avancement_chiffre"]=self[nom_CUSTOM].df["Niveau d'avancement*"].map(dict_conversion_avancement_chiffre)
            self[nom_CUSTOM].df["avancement_chiffre"] = self[nom_CUSTOM].df["avancement_chiffre"].astype(int)
            self[nom_CUSTOM].df .loc[self[nom_CUSTOM].df["avancement_chiffre"]==6,"Motif abandon (si niveau avancement = abandonné)"] = "4-Motif économique"
        return self
    
    def Cycle_dimport(self):
        for nom_CUSTOM,dict_CUSTOM in self.items():
            self[nom_CUSTOM].df["cycle dimport"] = "Cycle 3 : 2022-2027"
        return self
    
    def titre_action_DORA(self):
        def nommage_nom_action_MO(x):
            if x['NOM_PERSO_ACTION']!="nan":
                if x['Action_phare']=="x":
                    x["Titre de l'action*"] = "[DORA MO - " + x["NOM_PPG"] + "] " + x['NOM_PERSO_ACTION'] + " [phare] " + x['description_action_phare']
                if x['Action_phare']=="nan":
                    x["Titre de l'action*"] = "[DORA MO - " + x["NOM_PPG"] + "] " + x['NOM_PERSO_ACTION']
            if x['NOM_PERSO_ACTION']=="nan":
                if x['Action_phare']=="x":
                    x["Titre de l'action*"] = "[DORA MO - " + x["NOM_PPG"] + "] " + x['NOM_TYPE_ACTION_DORA'] + " [phare] " + x['description_action_phare']
                if x['Action_phare']=="nan":
                   x["Titre de l'action*"] = "[DORA MO - " + x["NOM_PPG"] + "] " + x['NOM_TYPE_ACTION_DORA'] 
            return x
        def nommage_nom_action_DEP(x,num_dep):
            if x['NOM_PERSO_ACTION']!="nan":
                if x['Action_phare']=="x":
                    x["Titre de l'action*"] = "[DORA DEP - " + num_dep + "] " + x['NOM_PERSO_ACTION'] + " [phare] " + x['description_action_phare']
                if x['Action_phare']=="nan":
                    x["Titre de l'action*"] = "[DORA DEP - " + num_dep + "] " + x['NOM_PERSO_ACTION']
            if x['NOM_PERSO_ACTION']=="nan":
                if x['Action_phare']=="x":
                    x["Titre de l'action*"] = "[DORA DEP - " + num_dep + "] " + x['NOM_TYPE_ACTION_DORA'] + " [phare] " + x['description_action_phare']
                if x['Action_phare']=="nan":
                   x["Titre de l'action*"] = "[DORA DEP - " + num_dep + "] " + x['NOM_TYPE_ACTION_DORA'] 
            return x       

        for nom_CUSTOM,dict_CUSTOM in self.items():
            if dict_CUSTOM.echelle_df=="MO":
            #Remplissage des colonnes type actions
                self[nom_CUSTOM].df = self[nom_CUSTOM].df.apply(lambda x:nommage_nom_action_MO(x),axis=1)
            if dict_CUSTOM.echelle_df=="DEP":
            #Remplissage des colonnes type actions
                self[nom_CUSTOM].df = self[nom_CUSTOM].df.apply(lambda x:nommage_nom_action_DEP(x,self[nom_CUSTOM].CODE_CUSTOM),axis=1)                
        return self

    def col_debut_avancement_et_engagement(self):
        #Remplissage des colonnes type actions
        for nom_CUSTOM,dict_CUSTOM in self.items():
            annee_remplissage_tableau = dict_CUSTOM.annee_remplissage
            self[nom_CUSTOM].df.loc[~self[nom_CUSTOM].df['annee_action_ini'].isnull(),"Date de début du niveau d'avancement*"] = "01/01/" + self[nom_CUSTOM].df['annee_action_ini'].astype(str)
            self[nom_CUSTOM].df.loc[self[nom_CUSTOM].df['annee_action_ini'].isnull(),"Date de début du niveau d'avancement*"] = "01/01/" + str(annee_remplissage_tableau)
            self[nom_CUSTOM].df.loc[~self[nom_CUSTOM].df['annee_action_eng'].isnull(),"Année prévisionnelle d'engagement"] = self[nom_CUSTOM].df['annee_action_eng'].astype(str)
            self[nom_CUSTOM].df.loc[self[nom_CUSTOM].df['annee_action_eng'].isnull(),"Année prévisionnelle d'engagement"] = str(annee_remplissage_tableau)
        return self 

    def conv_CODE_ROE_str(self):
        for nom_CUSTOM,dict_CUSTOM in self.items():
            self[nom_CUSTOM].df = self[nom_CUSTOM].df.explode('CODE_ROE')
            self[nom_CUSTOM].dict_type_col["CODE_ROE"] = "str"
        return self

    def col_SUIVI_PARCE(self):
        df_Ouvrage_PARCE = dataframe.import_df_liste_ouvrage_prioritaire_PARCE()
        liste_ouvrage_ROE_PARCE = df_Ouvrage_PARCE['CODE_ROE'].to_list()
        for nom_CUSTOM,dict_CUSTOM in self.items():
            self[nom_CUSTOM].df.loc[self[nom_CUSTOM].df['CODE_ROE'].isin(liste_ouvrage_ROE_PARCE),"Suivi « PARCE »"] = "O"
            self[nom_CUSTOM].df.loc[~self[nom_CUSTOM].df['CODE_ROE'].isin(liste_ouvrage_ROE_PARCE),"Suivi « PARCE »"] = "N"
        return self

    def etape_continuite(self):
        df_etapes_conti=config_DORA.import_tableau_etapes_Osmose_continuite()
        for nom_CUSTOM,dict_CUSTOM in self.items():
            df_tempo_conti = self[nom_CUSTOM].df.loc[self[nom_CUSTOM].df["CODE_TYPE_ACTION_OSMOSE"].str.startswith('MIA03')]
            df_tempo_conti = df_tempo_conti[["CODE_IMPORT","CODE_TYPE_ACTION_OSMOSE",'Avancement_spe_conti',"Date de début du niveau d'avancement*"]]
            df_tempo_conti["Date de début de l'étape*"] = df_tempo_conti["Date de début du niveau d'avancement*"]
            self[nom_CUSTOM].df_etapes = pd.merge(df_tempo_conti,df_etapes_conti,on=["CODE_TYPE_ACTION_OSMOSE","Avancement_spe_conti"])
            '''self[nom_CUSTOM].df = pd.merge(self[nom_CUSTOM].df,self[nom_CUSTOM].df_etapes[["CODE_IMPORT","Avancement"]],how="left", on=['CODE_IMPORT'],suffixes=[None,'_y'])
            self[nom_CUSTOM].df.loc[~self[nom_CUSTOM].df["Avancement_y"].isnull(),"Avancement"] = self[nom_CUSTOM].df["Avancement_y"]
            self[nom_CUSTOM].df = self[nom_CUSTOM].df[[x for x in list(self[nom_CUSTOM].df) if x!="Avancement_y"]]'''
        return self

    def remplissage_CODE_ME_si_vide(self,dict_relation_shp_liste):
        for nom_CUSTOM,dict_CUSTOM in self.items():
            self.df['CODE_ME'] = self.df['CODE_ME'].apply(lambda x:";".join(x[2:]))
        return self

    def actualisation_CODE_BVG(self,dict_relation_shp_liste):
        def mapping_sur_liste(x,dict_BVG_par_REF):
            list_list_BVG = [dict_BVG_par_REF[CODE_REF] for CODE_REF in x]
            list_list_BVG = [x for xs in list_list_BVG for x in xs]
            return list_list_BVG        
        for nom_CUSTOM,dict_CUSTOM in self.items():
            dict_liste_REF_par_BVG = dict_relation_shp_liste["dict_liste_ME_par_BVG"]
            dict_BVG_par_REF = {}
            for k,v in dict_liste_REF_par_BVG.items():
                for x in v:
                    dict_BVG_par_REF.setdefault(x,[]).append(k)
            self[nom_CUSTOM].df['CODE_BVG']=self[nom_CUSTOM].df['CODE_ME'].apply(lambda x:mapping_sur_liste(x,dict_BVG_par_REF))
            self[nom_CUSTOM].df['CODE_BVG']=self[nom_CUSTOM].df['CODE_BVG'].apply(lambda x:list(set(x)))
            self[nom_CUSTOM].dict_type_col['CODE_BVG'] = "list"
        return self

    def col_departement_PAOT(self,df_info_ME):
        dict_relation_CODE_ME_liste_DEP_pilote = dict(zip(df_info_ME['CODE_ME'].to_list(),df_info_ME['liste_CODE_DEP'].to_list()))
        def mapping_sur_liste(x,dict_relation_CODE_ME_liste_DEP_pilote):
            list_list_BVG = [dict_relation_CODE_ME_liste_DEP_pilote[CODE_REF] for CODE_REF in x]
            list_list_BVG = [x for xs in list_list_BVG for x in xs]
            list_list_BVG = list(set(list_list_BVG))
            return list_list_BVG
        for nom_CUSTOM,dict_CUSTOM in self.items():
            self[nom_CUSTOM].df['Département PAOT']= self[nom_CUSTOM].df['CODE_ME'].apply(lambda x:mapping_sur_liste(x,dict_relation_CODE_ME_liste_DEP_pilote))
            self[nom_CUSTOM].df['Code des départements*'] = self[nom_CUSTOM].df['Département PAOT']
            self[nom_CUSTOM].dict_type_col['Code des départements*'] = 'list'
            self[nom_CUSTOM].dict_type_col['Département PAOT'] = 'list'
        return self   

    def col_departement_pilote(self,dict_dict_info_REF):
        for nom_CUSTOM,dict_CUSTOM in self.items():
            if dict_CUSTOM.echelle_df == "DEP":
                dict_numero_dep_nom_dep = config_DORA.import_dict_num_dep_nom_dep()
                df_info_DEP = dict_dict_info_REF['df_info_DEP']
                self[nom_CUSTOM].df['Département Pilote PAOT'] = int(df_info_DEP.dict_NOM_CODE[nom_CUSTOM])
                self[nom_CUSTOM].df['Intitulé du service pilote*'] = self[nom_CUSTOM].df['Département Pilote PAOT'].map(dict_numero_dep_nom_dep)
            if dict_CUSTOM.echelle_df == "MO":
                dict_numero_dep_nom_dep = config_DORA.import_dict_num_dep_nom_dep()
                self[nom_CUSTOM].df['Département Pilote PAOT'] = int(dict_CUSTOM.numero_dep)
                self[nom_CUSTOM].df['Intitulé du service pilote*'] = self[nom_CUSTOM].df['Département Pilote PAOT'].map(dict_numero_dep_nom_dep)
        return self
    
    def col_bassin_DCE(self,dict_dict_info_REF):
        for nom_CUSTOM,dict_CUSTOM in self.items():
            self[nom_CUSTOM].df['Code bassin dce*'] = "F-ADOUR GARONNE"
        return self

    ##########################################################################################
    #PAOT
    ##########################################################################################   
    def colonne_PAOT(self):
        #Remplissage des colonnes type actions
        for nom_CUSTOM,dict_CUSTOM in self.items():
            self[nom_CUSTOM].df.loc[self[nom_CUSTOM].df['proposition_DORA']=="oui","PAOT 1"] = "O"
            self[nom_CUSTOM].df.loc[self[nom_CUSTOM].df['proposition_DORA']=="non","PAOT 1"] = "N"
            self[nom_CUSTOM].df["PAOT 2"] = self[nom_CUSTOM].df["PAOT 1"]
        return self 


    ##########################################################################################
    #PDM
    ##########################################################################################     
    def lien_CODE_PDM(self,echelle_df,dict_type_col):
        for nom_CUSTOM,dict_CUSTOM in self.items():
            if echelle_df =='MO':
                dict_chgmt_type_col_df_tableau_action_MIA_REF = dict_type_col
            if echelle_df =='DEP':
                dict_chgmt_type_col_df_tableau_action_MIA_REF = dict_type_col
            df_PDM_AG = dataframe.import_df_PDM_AG() 
            df_PDM_AG['liste_CODE_TEMPO'] = df_PDM_AG.apply(lambda df_PDM_AG:[x + "$" + df_PDM_AG['CODE_TYPE_ACTION_OSMOSE'] for x in df_PDM_AG['CODE_BVG']],axis=1)
            df_PDM_AG = df_PDM_AG.explode('liste_CODE_TEMPO')
            df_PDM_AG = df_PDM_AG[['CODE_PDM','liste_CODE_TEMPO']]
            self[nom_CUSTOM].df['liste_CODE_TEMPO'] = self[nom_CUSTOM].df.apply(lambda x:[x + "$" + self[nom_CUSTOM].df['CODE_TYPE_ACTION_OSMOSE'] for x in self[nom_CUSTOM].df['CODE_BVG']],axis=1)
            self[nom_CUSTOM].df = self[nom_CUSTOM].df.explode('liste_CODE_TEMPO')
            self[nom_CUSTOM].df = pd.merge(self[nom_CUSTOM].df,df_PDM_AG,on="liste_CODE_TEMPO",how='left')
            self[nom_CUSTOM].df = self[nom_CUSTOM].df[[x for x in list(self[nom_CUSTOM].df) if x!="liste_CODE_TEMPO"]]
            #ATTENTION, on supprime les éventuels doublons (très rare) si plusieurs CODE_PDM qui match ! Pour LB, c'est chiant
            
            df_tempo = self[nom_CUSTOM].df[['CODE_IMPORT','CODE_PDM']].groupby('CODE_IMPORT').agg({'CODE_PDM':lambda x: list(x)})
            df_tempo["CODE_PDM"] = df_tempo["CODE_PDM"].apply(lambda x:list(set(x)))
            self[nom_CUSTOM].df = self[nom_CUSTOM].df[[x for x in list(self[nom_CUSTOM].df) if x!="CODE_PDM"]]
            self[nom_CUSTOM].df = pd.merge(self[nom_CUSTOM].df,df_tempo,on="CODE_IMPORT",how='left')
            self[nom_CUSTOM].df = self[nom_CUSTOM].df.drop_duplicates(subset = ['CODE_IMPORT'], keep = 'first')
            #On ne garde que les CODE PDM avec un seul resultat !
            self[nom_CUSTOM].df.loc[self[nom_CUSTOM].df['CODE_PDM'].apply(lambda x:len(x)>1),"CODE_PDM"] = "Pas de PDM"
            self[nom_CUSTOM].df.loc[self[nom_CUSTOM].df['CODE_PDM'].apply(lambda x:x[0]!=x[0]),"CODE_PDM"] = "Pas de PDM"
            self[nom_CUSTOM].df['CODE_PDM'] = [[] if x=="Pas de PDM" else x for x in self[nom_CUSTOM].df['CODE_PDM'].to_list()]
            dict_chgmt_type_col_df_tableau_action_MIA_REF['CODE_PDM'] = 'list'
        return self
    
    def info_MO(self,df_info_MO):
        #On fait en fonction de la colonne de l'echelle principale d'action
        df_info_MO_temp = df_info_MO[['CODE_MO','CODE_SIRET_SANDRE','CODE_SIRET',"TYPE_MO","nom_usuel"]]
        for nom_CUSTOM,dict_CUSTOM in self.items():
            self[nom_CUSTOM].df = pd.merge(self[nom_CUSTOM].df,df_info_MO_temp,on="CODE_MO",how='left')
            self[nom_CUSTOM].df['NOM_MO'] = self[nom_CUSTOM].df['nom_usuel'] + " [" + self[nom_CUSTOM].df['CODE_MO'] + "]"
            self[nom_CUSTOM].df['NOM_SANDRE'] = ""
            self[nom_CUSTOM].df["CODE_SIRET"] = self[nom_CUSTOM].df['CODE_SIRET_SANDRE']
            self[nom_CUSTOM].df["CODE_SIRET"] = self[nom_CUSTOM].df["CODE_SIRET"].astype(str)
            self[nom_CUSTOM].df.loc[self[nom_CUSTOM].df["CODE_SIRET"]=="Ab SANDRE","CODE_SIRET"] = ""
            dict_TYPE_MO_type_MO_Osmose = config_DORA.import_dict_TYPE_MO_type_MO_Osmose()
            self[nom_CUSTOM].df["Type du maitre d'ouvrage*"] = self[nom_CUSTOM].df['TYPE_MO'].map(dict_TYPE_MO_type_MO_Osmose)
            self[nom_CUSTOM].df.loc[self[nom_CUSTOM].df["Type du maitre d'ouvrage*"].isnull()] = ""
        return self

    def attributs(self):
        df_attributs_hors_conti=config_DORA.import_tableau_attributs_Osmose()
        for nom_CUSTOM,dict_CUSTOM in self.items():
            df_tempo_hors_conti = self[nom_CUSTOM].df.loc[~self[nom_CUSTOM].df["CODE_TYPE_ACTION_OSMOSE"].str.startswith('MIA03')]
            df_tempo_hors_conti = df_tempo_hors_conti[["CODE_IMPORT","NOM_TYPE_ACTION_DORA","CODE_TYPE_ACTION_OSMOSE","avancement_chiffre"]]
            df_tempo_hors_conti = pd.merge(df_tempo_hors_conti,df_attributs_hors_conti,on="CODE_TYPE_ACTION_OSMOSE")
            df_tempo_hors_conti.loc[df_tempo_hors_conti['Attribut']=="Action financée par l'Agence de l'eau",'Valeur (s)'] = "Non"
            #df_tempo_hors_conti.loc[df_tempo_hors_conti['Attribut']=="Code(s) Aides attribué(s) par l'Agence de l'eau",'Valeur (s)'] = "a faire"
            df_tempo_hors_conti = df_tempo_hors_conti.loc[df_tempo_hors_conti['Attribut']!="Code(s) Aides attribué(s) par l'Agence de l'eau"]
            df_tempo_hors_conti = df_tempo_hors_conti.loc[df_tempo_hors_conti["avancement_chiffre"]>=df_tempo_hors_conti["obligatoire_chiffre_avancement"]]
            #Liste de toutes les conneries réclamées par Osmose alors qu'ils savent pas gérer un apostrophe
            df_tempo_hors_conti.loc[df_tempo_hors_conti["Attribut"]=="Action issue de l'Observatoire du développement rural",'Valeur (s)'] = "Non"
            df_tempo_hors_conti.loc[df_tempo_hors_conti["Attribut"]=="Code(s) Action attribué(s) par l'Observatoire du développement rural",'Valeur (s)'] = "Serral goat"
            df_tempo_hors_conti.loc[df_tempo_hors_conti["Attribut"]=="Thématique(s) visée(s)",'Valeur (s)'] = "Milieux aquatiques"
            df_tempo_hors_conti.loc[df_tempo_hors_conti["Attribut"]=="ZCSE",'Valeur (s)'] = "Non"
            
            df_tempo_conti = self[nom_CUSTOM].df.loc[self[nom_CUSTOM].df["CODE_TYPE_ACTION_OSMOSE"].str.startswith('MIA03')]
            df_tempo_conti = df_tempo_conti[["CODE_IMPORT","NOM_TYPE_ACTION_DORA","CODE_TYPE_ACTION_OSMOSE",'CODE_ROE',"avancement_chiffre","trav_prevu"]]
            df_config_attributs_conti=config_DORA.import_tableau_attributs_Osmose_continuite()
            df_tempo_conti = pd.merge(df_tempo_conti,df_config_attributs_conti,on="CODE_TYPE_ACTION_OSMOSE")
            
            #Suppression des lignes "Code ROE" si l'action ne concerne pas un ROE
            df_tempo_conti.loc[(df_tempo_conti['CODE_ROE'].isnull())&(df_tempo_conti["Attribut"]=="Action concernant un ROE"),'Valeur (s)']= "Non"
            df_tempo_conti.loc[(~df_tempo_conti['CODE_ROE'].isnull())&(df_tempo_conti["Attribut"]=="Action concernant un ROE"),'Valeur (s)']= "Oui"
            list_CODE_IMPORT_sans_CODE_ROE = df_tempo_conti.loc[(df_tempo_conti["Attribut"]=="Action concernant un ROE")&(df_tempo_conti["Valeur (s)"]=="Non")]["CODE_IMPORT"].to_list()
            list_index_tempo_a_enlever = list(df_tempo_conti.loc[df_tempo_conti["CODE_IMPORT"].isin(list_CODE_IMPORT_sans_CODE_ROE)&(df_tempo_conti["Attribut"]=="Code ROE")].index)
            df_tempo_conti = df_tempo_conti.drop(list_index_tempo_a_enlever)
            
            df_tempo_conti.loc[df_tempo_conti['Attribut']=="Code ROE",'Valeur (s)'] = df_tempo_conti['CODE_ROE']
            df_tempo_conti.loc[df_tempo_conti['Attribut']=="Action financée par l'Agence de l'eau",'Valeur (s)'] = "Non"
            
            dict_tempo_CODE_IMPORT_trav_prevu = dict(zip(self[nom_CUSTOM].df["CODE_IMPORT"].to_list(),self[nom_CUSTOM].df["trav_prevu"].to_list()))
            df_tempo_conti.loc[df_tempo_conti['Attribut']=="Solution technique retenue",'Valeur (s)'] = df_tempo_conti["CODE_IMPORT"].map(dict_tempo_CODE_IMPORT_trav_prevu)
            df_tempo_conti = df_tempo_conti.loc[(df_tempo_conti["Valeur (s)"]!='nan')]
            
            df_tempo_conti.loc[(df_tempo_conti['Attribut']=="Solution technique retenue")&(df_tempo_conti["NOM_TYPE_ACTION_DORA"]=="Restauration de la continuité écologique par équipement des ouvrages")&(df_tempo_conti['Valeur (s)']=='nan'),'Valeur (s)'] = "Ouvrage aménagé"
            df_tempo_conti.loc[(df_tempo_conti['Attribut']=="Solution technique retenue")&(df_tempo_conti["NOM_TYPE_ACTION_DORA"]=="Effacement d'ouvrage transversal")&(df_tempo_conti['Valeur (s)']=='nan'),'Valeur (s)'] = "Suppression"
            df_tempo_conti.loc[(df_tempo_conti['Attribut']=="Solution technique retenue")&(df_tempo_conti["NOM_TYPE_ACTION_DORA"]=="Gestion des ouvrages")&(df_tempo_conti['Valeur (s)']=='nan'),'Valeur (s)'] = "Ouvrage avec règle de gestion"
            
            df_tempo_conti = df_tempo_conti.loc[df_tempo_conti['Attribut']!="Code(s) Aides attribué(s) par l'Agence de l'eau"]
            self[nom_CUSTOM].df_attributs = pd.concat([df_tempo_hors_conti,df_tempo_conti])
        return self    

    
    def suppression_lignes_df_actions_MIA(self,dict_dict_info_REF=None,dict_CUSTOM_maitre=None):
        for nom_tableau,dict_df_actions_originaux in self.items(): 
            echelle_df = self[nom_tableau].echelle_df
            echelle_base_REF = self[nom_tableau].echelle_base_REF
            dict_nom_col = self[nom_tableau].dict_nom_col
            dict_type_col = self[nom_tableau].dict_type_col
            dict_nb_chiffres = self[nom_tableau].dict_nb_chiffres            
            ###Suppression des lignes
            self[nom_tableau].df = DfTableauxActionsMIA.suppression_CODE_REF_principale_hors_BDD(self[nom_tableau].df,echelle_df,self[nom_tableau].echelle_base_REF,dict_dict_info_REF,dict_type_col)
            self[nom_tableau].df = DfTableauxActionsMIA.suppression_actions_sans_avancement(self[nom_tableau].df)
            self[nom_tableau].df = DfTableauxActionsMIA.suppression_actions_sans_echelle_REF(self[nom_tableau].df)
            self[nom_tableau].df = DfTableauxActionsMIA.suppression_actions_conti_sans_ROE(self[nom_tableau].df,dict_CUSTOM_maitre)    
        return self
    
    def mise_en_forme_tableau_actions_MIA_renommage_colonne(self,dict_CUSTOM_maitre):
        for nom_tableau,dict_df_actions_originaux in self.items(): 
            if dict_CUSTOM_maitre.type_rendu=='carte' or dict_CUSTOM_maitre.type_rendu=='tableau_DORA_vers_BDD':     
                ###Comme le ç de Hawai, tu sers à rien
                #self[nom_tableau].df = DfTableauxActionsMIA.col_invariantes(self[nom_tableau].df)
                self[nom_tableau].df = DfTableauxActionsMIA.conversion_CODE_ME_forme_FR(self[nom_tableau].df)
                self[nom_tableau].df = DfTableauxActionsMIA.conversion_contenu_col_list(self[nom_tableau].df,self[nom_tableau].echelle_df,self[nom_tableau].dict_type_col)
                self[nom_tableau].df = DfTableauxActionsMIA.renommage_col(self[nom_tableau].df)
                self[nom_tableau].df = DfTableauxActionsMIA.ajout_col_invariantes(self[nom_tableau].df)
                self[nom_tableau].df = DfTableauxActionsMIA.ordre_colonne_Osmose_LOL(self[nom_tableau].df)
                self[nom_tableau].df = DfTableauxActionsMIA.remplacer_mention_VIDE(self[nom_tableau].df)
        return self
    
    def renommage_onglets_annexes(dict_dict_df_actions_Osmose):
        for nom_tableau,dict_df_actions_originaux in dict_dict_df_actions_Osmose.items():
            dict_dict_df_actions_Osmose[nom_tableau].df_Points_de_blocage = DfTableauxActionsMIA.renommage_onglet_blocage(dict_dict_df_actions_Osmose[nom_tableau].df_Points_de_blocage)
            dict_dict_df_actions_Osmose[nom_tableau].df_Etapes = DfTableauxActionsMIA.renommage_onglet_etapes(dict_dict_df_actions_Osmose[nom_tableau].df_Etapes)
            dict_dict_df_actions_Osmose[nom_tableau].df_Financeurs = DfTableauxActionsMIA.renommage_onglet_financeurs(dict_dict_df_actions_Osmose[nom_tableau].df_Financeurs)
            dict_dict_df_actions_Osmose[nom_tableau].df_attributs = DfTableauxActionsMIA.renommage_onglet_attributs(dict_dict_df_actions_Osmose[nom_tableau].df_attributs)
        return dict_dict_df_actions_Osmose

    def rassemblement_tableaux_MIA(dict_dict_df_actions_originaux):
        liste_df_actions_tableaux = []
        for nom_tableau,dict_df_actions_originaux in dict_dict_df_actions_originaux.items():
            liste_df_actions_tableaux.append(dict_dict_df_actions_originaux[nom_tableau].df)
        df_gros_tableaux = pd.concat(liste_df_actions_tableaux, ignore_index=True)        
        return df_gros_tableaux

    ######################
    #Maj Osmose
    ######################
    def mise_en_forme_tableau_actions_maj_osmose(dict_dict_df_actions_originaux,dict_dict_info_REF=None,dict_relation_shp_liste=None,dict_CUSTOM_maitre=None):
        for nom_tableau,dict_df_actions_originaux in dict_dict_df_actions_originaux.items():
            ###Les verifs
            dict_dict_df_actions_originaux[nom_tableau].df = DfTableauxActionsMIA.changement_nom_col_tableaux_MAJ_osmose(dict_dict_df_actions_originaux[nom_tableau].df)
            dict_dict_df_actions_originaux[nom_tableau].df = dataframe.strip_et_tease_contenu_col(dict_dict_df_actions_originaux[nom_tableau].df)
            dict_dict_df_actions_originaux[nom_tableau].df = DfTableauxActionsMIA.modification_col_CODE_ME(dict_dict_df_actions_originaux[nom_tableau].df)
            dict_dict_df_actions_originaux[nom_tableau].df = DfTableauxActionsMIA.modification_col_CODE_TYPE_ACTION_OSMOSE(dict_dict_df_actions_originaux[nom_tableau].df)
            dict_dict_df_actions_originaux[nom_tableau] = dataframe.creation_doublon_df(dict_dict_df_actions_originaux[nom_tableau])
            dict_dict_df_actions_originaux[nom_tableau].df = DfTableauxActionsMIA.ajout_CODE_BVG(dict_dict_df_actions_originaux[nom_tableau].df,dict_relation_shp_liste)
            dict_dict_df_actions_originaux[nom_tableau] = DfTableauxActionsMIA.separation_actions_hors_PDM(dict_dict_df_actions_originaux[nom_tableau])
            dict_dict_df_actions_originaux[nom_tableau] = DfTableauxActionsMIA.recuperation_info_ME_BVG_CODE_PDM_pour_actions_abandonnees(dict_dict_df_actions_originaux[nom_tableau],dict_relation_shp_liste,dict_dict_info_REF)
            dict_dict_df_actions_originaux[nom_tableau] = DfTableauxActionsMIA.gestion_df_attributs(dict_dict_df_actions_originaux[nom_tableau])
            dict_dict_df_actions_originaux[nom_tableau] = DfTableauxActionsMIA.transfo_col_en_str(dict_dict_df_actions_originaux[nom_tableau])
            dict_dict_df_actions_originaux[nom_tableau] = DfTableauxActionsMIA.changement_nom_col_tableaux_MAJ_osmose_vers_osmose(dict_dict_df_actions_originaux[nom_tableau])
            dict_dict_df_actions_originaux[nom_tableau] = DfTableauxActionsMIA.mise_ordre_col_maj_osmose(dict_dict_df_actions_originaux[nom_tableau])
            
        return dict_dict_df_actions_originaux
    
    ######################
    #Verif tableau DORA
    ######################    
    def verif_df_actions_MIA(dict_df_actions_nettoye,dict_dict_df_actions_originaux,dict_dict_info_REF,dict_log_df_erreur):
        for nom_tableau,dict_df_actions_originaux in dict_dict_df_actions_originaux.items():
            echelle_df = dict_df_actions_nettoye[nom_tableau].echelle_df
            numero_dep = dict_df_actions_nettoye[nom_tableau].numero_dep
            dict_type_col = dict_df_actions_nettoye[nom_tableau].dict_type_col
            dict_nom_col = dict_df_actions_nettoye[nom_tableau].dict_nom_col
            dict_log_df_erreur[nom_tableau] = DfTableauxActionsMIA.verification_colonne_NOM_REF_existence_BDD(dict_log_df_erreur[nom_tableau],dict_df_actions_nettoye[nom_tableau].df,dict_dict_info_REF,echelle_df,numero_dep,dict_type_col)
            dict_log_df_erreur[nom_tableau] = DfTableauxActionsMIA.verification_colonne_CODE_REF(dict_log_df_erreur[nom_tableau],dict_df_actions_originaux,dict_dict_info_REF)
            dict_log_df_erreur[nom_tableau] = DfTableauxActionsMIA.verification_colonne_echelle_princ_action(dict_log_df_erreur[nom_tableau],dict_df_actions_nettoye[nom_tableau].df)
            dict_log_df_erreur[nom_tableau] = DfTableauxActionsMIA.verification_colonne_avancement(dict_log_df_erreur[nom_tableau],dict_df_actions_nettoye[nom_tableau].df)
            dict_log_df_erreur[nom_tableau] = DfTableauxActionsMIA.verification_colonnes_annee_avancement(dict_log_df_erreur[nom_tableau],dict_df_actions_originaux)
        return dict_log_df_erreur
    
    def jointure_tableau_origine_avec_erreur(dict_dict_df_actions_originaux,dict_log_df_erreur):
        for nom_tableau,dict_df_actions_originaux in dict_dict_df_actions_originaux.items():
            dict_renomage_inverse = {v:k for k,v in dict_df_actions_originaux.dict_nom_col.items()}
            dict_renomage_erreur_inverse = {"Erreur "+v:"Erreur "+k for k,v in dict_df_actions_originaux.dict_nom_col.items()}
            dict_df_actions_originaux.df = dict_df_actions_originaux.df.loc[list(dict_log_df_erreur[nom_tableau].index)]
            dict_log_df_erreur[nom_tableau] = dict_log_df_erreur[nom_tableau][[x for x in list(dict_log_df_erreur[nom_tableau]) if x!=0]]
            dict_log_df_erreur[nom_tableau] = dict_log_df_erreur[nom_tableau][[x for x in list(dict_log_df_erreur[nom_tableau]) if not all(x=='nan' for x in dict_log_df_erreur[nom_tableau][x].to_list())]]
            list_col_erreur = list(dict_log_df_erreur[nom_tableau])
            list_col_df_original = list(dict_dict_df_actions_originaux[nom_tableau].df_brut)
            for nom_col_erreur in list_col_erreur:
                nom_col_type_DORA = nom_col_erreur.split("Erreur ")[-1]
                if nom_col_type_DORA in list_col_df_original:
                    num_col = dict_log_df_erreur[nom_tableau].columns.get_loc(nom_col_erreur)
                    dict_log_df_erreur[nom_tableau].insert(num_col, nom_col_erreur.split("Erreur ")[-1], dict_dict_df_actions_originaux[nom_tableau].df_brut[nom_col_type_DORA].to_list())                    
            dict_log_df_erreur[nom_tableau] = dict_log_df_erreur[nom_tableau].rename(dict_renomage_inverse,axis=1)
            dict_log_df_erreur[nom_tableau] = dict_log_df_erreur[nom_tableau].rename(dict_renomage_erreur_inverse,axis=1)
        return dict_log_df_erreur
    
    def retour_nom_col_tableau_original(dict_log_df_erreur,dict_dict_df_actions_originaux):
        for nom_tableau,dict_df_erreur in dict_log_df_erreur.items():
            dict_inverse = {v: k for k, v in dict_dict_df_actions_originaux[nom_tableau].dict_nom_col.items()}
            dict_log_df_erreur[nom_tableau] = dict_log_df_erreur[nom_tableau].rename(dict_inverse,axis=1)
        return dict_log_df_erreur
    
    def suppression_si_col_vide(dict_log_df_erreur,dict_dict_df_actions_originaux_avec_mise_en_forme_sans_modif):
        for nom_tableau,df_erreur in dict_log_df_erreur.items():
            for nom_col_erreur in list(dict_log_df_erreur[nom_tableau]):
                if isinstance(nom_col_erreur, str):
                    if nom_col_erreur.startswith("Erreur ") and nom_col_erreur.split("Erreur ")[-1] in list(dict_dict_df_actions_originaux_avec_mise_en_forme_sans_modif[nom_tableau].df):
                        if all([True if x=='nan' else False for x in dict_log_df_erreur[nom_tableau][nom_col_erreur].to_list()]):
                            index_num_col_prec_a_supprimer_aussi = dict_log_df_erreur[nom_tableau].columns.get_loc(nom_col_erreur.split("Erreur ")[-1])
                            dict_log_df_erreur[nom_tableau] = dict_log_df_erreur[nom_tableau].drop(dict_log_df_erreur[nom_tableau].columns[[index_num_col_prec_a_supprimer_aussi, index_num_col_prec_a_supprimer_aussi+1]],axis = 1)
        return dict_log_df_erreur
    
    def mise_en_forme_avant_export(dict_log_df_erreur):
        for nom_tableau,df_erreur in dict_log_df_erreur.items():
            dict_log_df_erreur[nom_tableau] = dict_log_df_erreur[nom_tableau].replace({np.nan: None})
            dict_log_df_erreur[nom_tableau] = dict_log_df_erreur[nom_tableau].replace({'nan': None})
        return dict_log_df_erreur    
    
    def export_log_erreur(dict_log_df_erreur):
        for nom_tableau,df_erreur in dict_log_df_erreur.items():
            dict_log_df_erreur[nom_tableau] = dict_log_df_erreur[nom_tableau].replace({np.nan: None})
            dict_log_df_erreur[nom_tableau] = dict_log_df_erreur[nom_tableau].replace({'nan': None})
        return dict_log_df_erreur  
    
    ######################
    #Remplissage BDD DORA
    ###################### 
    def transfert_BDD_DORA(contenu_donnees):
        BDD_DORA = dataframe.import_BDD_DORA()
        for nom_MO,dict_MO in contenu_donnees.items():
            BDD_DORA = BDD_DORA.loc[BDD_DORA['origine_df']!=dict_MO.CODE_CUSTOM +"#"+dict_MO.annee_remplissage]
            df_tempo_import_BDD = dict_MO.df[list(BDD_DORA)]
            BDD_DORA = pd.concat([BDD_DORA,df_tempo_import_BDD], ignore_index=True)
            BDD_DORA.to_csv("/mnt/g/travail/carto/projets basiques/PAOT global 5.0/BDD_DORA.csv",index=False)
        return contenu_donnees


    ######################
    #Filtres tableau DORA
    ###################### 
    def actualisation_CODE_ME(contenu_donnees,dict_relation_shp_liste):
        for nom_tableau,dict_df_actions_traites in contenu_donnees.items():
                list_echelle_REF = list(set(contenu_donnees[nom_tableau].df['echelle_princ_action'].to_list()))
                for REF in list_echelle_REF:
                    CODE_REF = "CODE_"+ REF
                    if CODE_REF!="CODE_ME":
                        if contenu_donnees[nom_tableau].dict_type_col[CODE_REF]=='str':
                            contenu_donnees[nom_tableau].df.loc[contenu_donnees[nom_tableau].df['echelle_princ_action']==REF,"CODE_ME"] = contenu_donnees[nom_tableau].df[CODE_REF].map(dict_relation_shp_liste['dict_liste_ME_par_'+REF])
                        if contenu_donnees[nom_tableau].dict_type_col[CODE_REF]=='list':
                            contenu_donnees[nom_tableau].df.loc[contenu_donnees[nom_tableau].df['echelle_princ_action']==REF,"CODE_ME"] = contenu_donnees[nom_tableau].df[CODE_REF].apply(lambda x:[dict_relation_shp_liste['dict_liste_ME_par_'+REF][CODE_SME][0] if CODE_SME in dict_relation_shp_liste['dict_liste_ME_par_'+REF] else CODE_SME for CODE_SME in x])
        return contenu_donnees

    def filtre_pression_tableau_DORA(dict_filtre_tableau_DORA,contenu_donnees):
        df_pression_AG = dataframe.import_pression()
        df_pression_AG = df_pression_AG.rename({"P_MORPHO":"P_MOR","P_HYDRO":"P_HYD","P_CONTI":"P_CON"},axis=1)
        pression_hydromorpho =  ["MOR","CON","HYD"]
        dict_pression_AG = df_pression_AG.set_index('CODE_ME').to_dict('index')
        
        def chercher_au_moins_une_pression_signi(x,pression,dict_pression_AG):
            x['PRESSION_'+pression] = [dict_pression_AG[x]["P_"+pression] if x in dict_pression_AG else 0 for x in x['CODE_ME']]
            return x
        
        def calcul_pourcentage_pression_traite(x,pression):
            if len(x['PRESSION_'+pression])>0:
                x['PRESSION_'+pression] = len([ME_mauvais for ME_mauvais in x['PRESSION_'+pression] if ME_mauvais in [3]])/len(x['PRESSION_'+pression]) 
            else:
                x['PRESSION_'+pression]=0
            return x
        
        def pression_traitee(x):
            for pression in pression_hydromorpho:
                x["pression_traitee"] = "aucune"
                if x["filtre_pression_"+pression]==1:
                    x["pression_traitee"] = pression
                
            return x

        dict_config_actions_MIA_DORA = config_DORA.import_dict_config_actions_MIA_DORA()
        df_conv_action_DORA = dict_config_actions_MIA_DORA['df_actions_MIA_elu'].rename({"Hydrologie":"P_HYD_traite","Morpho":"P_MOR_traite","Continuité":"P_CON_traite"},axis=1)
        for nom_tableau,dict_df_actions_traites in contenu_donnees.items():
            contenu_donnees[nom_tableau].df = pd.merge(contenu_donnees[nom_tableau].df,df_conv_action_DORA,how="left",on="NOM_TYPE_ACTION_DORA")
            for pression in pression_hydromorpho:
                contenu_donnees[nom_tableau].df = contenu_donnees[nom_tableau].df.apply(lambda x:chercher_au_moins_une_pression_signi(x,pression,dict_pression_AG),axis=1)
                contenu_donnees[nom_tableau].df = contenu_donnees[nom_tableau].df.apply(lambda x:calcul_pourcentage_pression_traite(x,pression),axis=1)
                
                contenu_donnees[nom_tableau].df.loc[(contenu_donnees[nom_tableau].df['PRESSION_'+pression]>0)&(contenu_donnees[nom_tableau].df["P_" + pression + "_traite"]==1),"filtre_pression_"+pression] = 1
                contenu_donnees[nom_tableau].df["filtre_pression_"+pression] = contenu_donnees[nom_tableau].df["filtre_pression_"+pression].fillna(0)
                #contenu_donnees[nom_tableau].df["filtre_pression_"+pression] = (contenu_donnees[nom_tableau].df["filtre_pression_"+pression]*100).astype(str) + " %"
            contenu_donnees[nom_tableau].df = contenu_donnees[nom_tableau].df.apply(lambda x:pression_traitee(x),axis=1)
            contenu_donnees[nom_tableau].df["pression_traitee"] = contenu_donnees[nom_tableau].df["pression_traitee"].fillna("aucune")
            contenu_donnees[nom_tableau].df["filtre_pression"] = contenu_donnees[nom_tableau].df[["filtre_pression_"+pression for pression in pression_hydromorpho]].sum(axis=1)
            contenu_donnees[nom_tableau].df["filtre_pression"] = contenu_donnees[nom_tableau].df["filtre_pression"].astype(int)
            contenu_donnees[nom_tableau].df.loc[contenu_donnees[nom_tableau].df["filtre_pression"]==1,"filtre_pression_final"] = "oui"
            contenu_donnees[nom_tableau].df.loc[contenu_donnees[nom_tableau].df["filtre_pression"]==0,"filtre_pression_final"] = "non"
            for pression in pression_hydromorpho:
                contenu_donnees[nom_tableau].df.loc[contenu_donnees[nom_tableau].df["pression_traitee"]==pression,"pourcentage_ME"] = contenu_donnees[nom_tableau].df["PRESSION_"+pression]
            contenu_donnees[nom_tableau].df.loc[contenu_donnees[nom_tableau].df["pourcentage_ME"].isnull(),"pourcentage_ME"] = 0
            contenu_donnees[nom_tableau].df["pourcentage_ME"] = contenu_donnees[nom_tableau].df["pourcentage_ME"].apply(lambda x:str((round(x,2)*100)) + " %")
            dict_filtre_tableau_DORA[nom_tableau] = pd.merge(dict_filtre_tableau_DORA[nom_tableau],contenu_donnees[nom_tableau].df[["pourcentage_ME","filtre_pression_final","pression_traitee"]],left_index=True,right_index=True)
        return dict_filtre_tableau_DORA

    def filtre_PDM(dict_filtre_tableau_DORA,contenu_donnees,dict_relation_shp_liste):
        df_PDM_AG = BDD.import_df_PDM_AG()
        df_PDM_AG = df_PDM_AG.explode("CODE_TYPE_ACTION_OSMOSE")
        df_PDM_AG['CODE_ME'] = df_PDM_AG["CODE_BVG"].apply(lambda x:[dict_relation_shp_liste["dict_liste_ME_par_BVG"][CODE_BVG] if CODE_BVG in dict_relation_shp_liste["dict_liste_ME_par_BVG"] else ['nan'] for CODE_BVG in x])
        df_PDM_AG["CODE_ME"] = df_PDM_AG["CODE_ME"].apply(lambda x:[element for innerList in x for element in innerList])
        df_PDM_AG["CODE_ME"] = df_PDM_AG["CODE_ME"].apply(lambda x:[CODE_ME for CODE_ME in x if CODE_ME!='nan'])
        df_PDM_AG = df_PDM_AG[df_PDM_AG["CODE_ME"].map(len)>1]
        df_PDM_AG = df_PDM_AG.explode("CODE_ME")
        #Il faut remonter au BVG, et pas garder simplement la ME !
        '''dict_relation_shp_liste = dictGdfCompletREF.ajout_inversion_dict_relation_1_pour_1(dict_relation_shp_liste,"dict_liste_ME_par_BVG")
        df_PDM_AG["CODE_BVG"] = df_PDM_AG["CODE_ME"].map(dict_relation_shp_liste['dict_BVG_par_ME'])'''
        df_PDM_AG['CODE_OSMOSE_et_CODE_ME'] = df_PDM_AG['CODE_TYPE_ACTION_OSMOSE'] + "_" + df_PDM_AG['CODE_ME']
        dict_PDM = dict(zip(df_PDM_AG['CODE_OSMOSE_et_CODE_ME'],df_PDM_AG['CODE_PDM']))
        def recherche_liste_CODE_PDM(x,dict_PDM,dict_relation_shp_liste):
            x['CODE_OSMOSE_et_CODE_ME'] = x['CODE_TYPE_ACTION_OSMOSE'] + "_" + x['CODE_ME']
            if x["CODE_OSMOSE_et_CODE_ME"] in dict_PDM:
                x['CODE_PDM'] = dict_PDM[x["CODE_OSMOSE_et_CODE_ME"]]
            else:
                x['CODE_PDM'] = 'nan'
            return x
        dict_tempo_PDM = {k:pd.DataFrame([]) for k,v in contenu_donnees.items()}
        for nom_tableau,dict_df_actions_traites in contenu_donnees.items():
            dict_tempo_PDM[nom_tableau] = contenu_donnees[nom_tableau].df.explode("CODE_ME")
            dict_tempo_PDM[nom_tableau] = dict_tempo_PDM[nom_tableau].loc[~dict_tempo_PDM[nom_tableau]["CODE_ME"].isnull()]
            #dict_tempo_PDM[nom_tableau]["CODE_BVG"] = dict_tempo_PDM[nom_tableau]['CODE_ME'].map(dict_relation_shp_liste['dict_BVG_par_ME'])
            dict_tempo_PDM[nom_tableau] = dict_tempo_PDM[nom_tableau].apply(lambda x:recherche_liste_CODE_PDM(x,dict_PDM,dict_relation_shp_liste),axis=1)
            dict_tempo_PDM[nom_tableau].reset_index(inplace=True)
            dict_tempo_PDM[nom_tableau] = dict_tempo_PDM[nom_tableau].groupby("index").agg({"CODE_PDM":lambda x:list(x)})
            #Ici, on ne veut garder le CODE PDM que si la liste est composée de la même valeur
            dict_tempo_PDM[nom_tableau]['CODE_PDM'] = dict_tempo_PDM[nom_tableau]['CODE_PDM'].apply(lambda x:list(set(x)))
            dict_tempo_PDM[nom_tableau]['CODE_PDM'] = dict_tempo_PDM[nom_tableau]['CODE_PDM'].apply(lambda x:x if len(x)==1 else ['nan'])
            dict_tempo_PDM[nom_tableau]['CODE_PDM'] = dict_tempo_PDM[nom_tableau]['CODE_PDM'].apply(lambda x:x[0])
            dict_tempo_PDM[nom_tableau].loc[dict_tempo_PDM[nom_tableau]["CODE_PDM"]!='nan',"filtre_PDM"] = "oui"
            dict_tempo_PDM[nom_tableau].loc[dict_tempo_PDM[nom_tableau]["CODE_PDM"]=='nan',"filtre_PDM"] = "non"
            dict_filtre_tableau_DORA[nom_tableau] = pd.merge(dict_filtre_tableau_DORA[nom_tableau],dict_tempo_PDM[nom_tableau],left_index=True,right_index=True)
        return dict_filtre_tableau_DORA
    
    def filtre_bon_etat(dict_filtre_tableau_DORA,contenu_donnees):
        df_pression_AG = dataframe.import_pression()
        dict_BE = dict(zip(df_pression_AG['CODE_ME'],df_pression_AG['ETAT_ECO']))
        def calcul_pourcentage_ME_traite(x):
            if len(x)>0:
                x = len([ME_mauvais for ME_mauvais in x if ME_mauvais in [3,4,5]])/len(x)
            else:
                x=0
            return x

        for nom_tableau,dict_df_actions_traites in contenu_donnees.items():
            contenu_donnees[nom_tableau].df['etat_eco'] = contenu_donnees[nom_tableau].df['CODE_ME'].apply(lambda x:[dict_BE[CODE_ME] if CODE_ME in dict_BE else 0 for CODE_ME in x])
            contenu_donnees[nom_tableau].df['etat_eco'] = contenu_donnees[nom_tableau].df['etat_eco'].apply(lambda x:calcul_pourcentage_ME_traite(x))
            contenu_donnees[nom_tableau].df.loc[contenu_donnees[nom_tableau].df['etat_eco']>0,"filtre_BE"] = 'oui'
            contenu_donnees[nom_tableau].df['etat_eco'] = (contenu_donnees[nom_tableau].df['etat_eco']*100).astype(str) + " %"
            dict_filtre_tableau_DORA[nom_tableau] = pd.merge(dict_filtre_tableau_DORA[nom_tableau],contenu_donnees[nom_tableau].df[['etat_eco',"filtre_BE"]],left_index=True,right_index=True)
        return dict_filtre_tableau_DORA

    def filtre_finale(dict_filtre_tableau_DORA):
        def resultat_filtre(x,list_col_filtre):
            if any([x[col_filtre]=="oui" for col_filtre in list_col_filtre]):
                x['proposition_DORA'] = "oui"
            else:
                x['proposition_DORA'] = "non"
            return x['proposition_DORA']
        for nom_tableau,dict_df_actions_traites in dict_filtre_tableau_DORA.items():
            list_col_filtre = [x for x in list(dict_df_actions_traites) if x.startswith('filtre')]
            dict_df_actions_traites['proposition_DORA'] = dict_df_actions_traites.apply(lambda x:resultat_filtre(x,list_col_filtre),axis=1)
        return dict_filtre_tableau_DORA
    
    def ajout_CODE_DORA_unique(dict_filtre_tableau_DORA,contenu_donnees):
        for nom_tableau,dict_df_actions_traites in dict_filtre_tableau_DORA.items():
            contenu_donnees[nom_tableau].df['ID_DORA'] = contenu_donnees[nom_tableau].df['etat_eco']
            dict_filtre_tableau_DORA[nom_tableau] = pd.merge(dict_filtre_tableau_DORA[nom_tableau],contenu_donnees[nom_tableau].df[['etat_eco',"filtre_BE"]],left_index=True,right_index=True)
        return dict_filtre_tableau_DORA