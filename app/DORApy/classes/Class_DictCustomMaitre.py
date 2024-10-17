# -*- coding: utf-8 -*-
import pandas as pd
import geopandas as gpd
import glob

from app.DORApy.classes.modules import CUSTOM,dataframe,tableau_excel,connect_path
from app.DORApy.classes import Class_NGdfREF

from app.DORApy.classes.Class_DictBoiteComplete import DictBoiteComplete
from app.DORApy.classes.Class_Bloc import BlocTexteSimple,BlocIcone,BlocLignesMultiples
from app.DORApy.classes.Class_dictGdfCompletREF import dictGdfCompletREF
from app.DORApy.classes.Class_GdfCompletREF import ListGdfCUSTOM,GdfFondCarte,GdfCompletREF
from app.DORApy.classes.Class_DictDfInfoShp import DictDfInfoShp
from app.DORApy.classes.Class_DictDFTableauxActionsMIA import DictDFTableauxActionsMIA
from app.DORApy.classes.modules import config_DORA
dict_config_espace,dict_config_police = config_DORA.creation_dicts_config()

import os.path
import itertools
import numpy as np
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
from openpyxl.worksheet.datavalidation import DataValidationList
from openpyxl.styles import NamedStyle
from datetime import date
import os 
ENVIRONMENT = os.getenv('ENVIRONMENT')

import time
start_time = time.time()
import datetime
annee_actuelle = datetime.date.today().year

chemin_dossier = "/mnt/g/travail/carto/projets basiques/PAOT global 5.0/"

##########################################################################################
#Config_projet
##########################################################################################
class DictCustomMaitre(dict):
    @property
    def _constructor(self):
        return DictCustomMaitre
    
    ##########################################################################################
    #Projet : Partie initiation
    ##########################################################################################
    def set_config_type_projet(self,type_rendu,type_donnees=None,thematique=None,public_cible=None,
                               liste_echelle_shp_CUSTOM_a_check=None,liste_grand_bassin=['AG'],info_fond_carte=None,echelle_REF=None,echelle_base_REF=None):
        self.type_rendu = type_rendu
        self.type_donnees = type_donnees
        self.thematique = thematique
        self.public_cible = public_cible
        self.liste_echelle_shp_CUSTOM_a_check = liste_echelle_shp_CUSTOM_a_check
        self.liste_grand_bassin = liste_grand_bassin
        self.info_fond_carte = info_fond_carte
        self.echelle_REF = echelle_REF
        self.echelle_base_REF = echelle_base_REF
        if self.liste_echelle_shp_CUSTOM_a_check==['MO']:
            self.taille_carto = 'petite'
        if self.liste_echelle_shp_CUSTOM_a_check==['DEP']:
            self.taille_carto = 'grande'
        if type_donnees!=None and thematique!=None and public_cible!=None:
            self.nom_dossier_maitre = type_donnees + '_' + thematique + '_' + public_cible
        return self

    def creation_dict_boite_complete_normal(self,nom_boite_maitre,avancement_max=5,taille_globale_carto='petite'):
        dict_boite_complete = DictBoiteComplete(taille_globale_carto)
        dict_boite_complete.nom_boite_maitre = nom_boite_maitre
        self['dict_boite_maitre_normal'] = dict_boite_complete
        self['dict_boite_maitre_normal'].orientation = 'normal'
        return self

    def creation_dict_boite_complete_orthogonal(self,nom_boite_maitre,avancement_max=5,taille_globale_carto='petite'):
        dict_boite_complete = DictBoiteComplete(taille_globale_carto)
        dict_boite_complete.nom_boite_maitre = nom_boite_maitre
        self['dict_boite_maitre_ortho'] = dict_boite_complete
        self['dict_boite_maitre_ortho'].orientation = 'orthogonal'
        return self    

    def remplissage_bloc_REF_dict_dict_boite_maitre(self,dict_relation_shp_liste):
        for nom_CUSTOM,entite_CUSTOM in self.items():
            CODE_CUSTOM = entite_CUSTOM.CODE_CUSTOM
            for nom_boite,dict_boite in entite_CUSTOM.items():
                liste_echelle_base_REF = dict_boite.liste_echelle_REF
                list_tempo_df = [] 
                for echelle_REF in liste_echelle_base_REF:
                    if echelle_REF!='CUSTOM':
                        df_tempo = pd.DataFrame(dict_relation_shp_liste['dict_liste_'+ echelle_REF +'_par_CUSTOM'][CODE_CUSTOM], columns=['CODE_REF'])
                    if echelle_REF=='CUSTOM':
                        df_tempo = pd.DataFrame([CODE_CUSTOM], columns=['CODE_REF'])
                    df_tempo['echelle_REF'] = echelle_REF
                    list_tempo_df.append(df_tempo)
                df_normal_CODE_REF_du_CUSTOM = pd.concat(list_tempo_df)
                dict_boite.df_CODE_REF = df_normal_CODE_REF_du_CUSTOM
                for nom_bloc,bloc in dict_boite.items():
                    bloc.df = df_normal_CODE_REF_du_CUSTOM
        return self

    def remplissage_boite_REF_dict_dict_boite_maitre(self,dict_relation_shp_liste):
        for nom_entite_CUSTOM,contenu_CUSTOM in self.items():
            for nom_CUSTOM,dict_boite in contenu_CUSTOM.items():
                liste_df_CODE_REF_tempo = []
                for nom_bloc,dict_bloc in dict_boite.items():
                    liste_df_CODE_REF_tempo.append(dict_bloc.df[['CODE_REF','echelle_REF']])
                df_CODE_REF_tempo = pd.concat(liste_df_CODE_REF_tempo)
                df_CODE_REF_tempo = df_CODE_REF_tempo.drop_duplicates(subset="CODE_REF", keep='first')
                dict_boite.df = df_CODE_REF_tempo
        return self

    def attributs_liste_echelle_REF_projet(self):
        def ajouter_echelle_REF(self,liste_echelle_REF_projet):
            if hasattr(self,"liste_echelle_CUSTOM"):
                liste_echelle_REF_projet.extend(self.liste_echelle_CUSTOM) 
            for nom_CUSTOM,entite_CUSTOM in self.items():
                if hasattr(entite_CUSTOM,"echelle_REF"):
                    liste_echelle_REF_projet.append(entite_CUSTOM.echelle_REF)
                if hasattr(entite_CUSTOM,"echelle_base_REF"):
                    liste_echelle_REF_projet.append(entite_CUSTOM.echelle_base_REF)  
                for type_boite,dict_boite in entite_CUSTOM.items():
                    if hasattr(dict_boite,"liste_echelle_REF"):
                        liste_echelle_REF_projet.extend(dict_boite.liste_echelle_REF)
            return liste_echelle_REF_projet
        
        liste_echelle_REF_projet = []
        if (self.type_rendu=='carte' and self.type_donnees=='action') or self.type_rendu=='verif_tableau_DORA':
            liste_echelle_REF_projet = ['SAGE','MO','PPG','BVG','ME','SME']
            liste_echelle_REF_projet.extend(self.liste_echelle_shp_CUSTOM_a_check)
            liste_echelle_REF_projet = list(set(liste_echelle_REF_projet))

        if (self.type_rendu=='carte' and self.type_donnees=='toutes_pressions'):
            liste_echelle_REF_projet = ['ME']

        if self.type_rendu=='tableau_vierge' and self.type_donnees=='action':
            liste_echelle_REF_projet = ['MO','SAGE','PPG','BVG','ME','ROE']
            liste_echelle_REF_projet = ajouter_echelle_REF(self,liste_echelle_REF_projet)
            for nom_CUSTOM,entite_CUSTOM in self.items():
                if hasattr(entite_CUSTOM,"echelle_REF"):
                    liste_echelle_REF_projet.append(entite_CUSTOM.echelle_base_REF)
        if self.type_rendu=='tableau_DORA_vers_BDD' and self.type_donnees=='action':
            liste_echelle_REF_projet = ['SAGE','MO','PPG','BVG','ME','SME']
            for nom_CUSTOM,entite_CUSTOM in self.items():
                if hasattr(entite_CUSTOM,"echelle_REF"):
                    liste_echelle_REF_projet.append(entite_CUSTOM.echelle_base_REF)             
        if self.type_rendu=='tableau_MAJ_Osmose':
            liste_echelle_REF_projet = ['MO','SAGE','PPG','BVG','ME','ROE']            
            
        self.liste_echelle_REF_projet = list(set(liste_echelle_REF_projet))
        return self

    def attributs_liste_echelle_base_REF(self):
        if self.type_rendu=='carte':
            liste_echelle_base_REF = list({k:v.echelle_base_REF for k,v in self.items()}.values())
            liste_echelle_base_REF = list(set(liste_echelle_base_REF))
            self.liste_echelle_base_REF = liste_echelle_base_REF
        return self

    def definition_attributs_liste_echelle_boite_par_CUSTOM(self,dict_dict_info_CUSTOM):
        if self.type_rendu=='carte' and self.type_donnees=='action' and self.thematique=='MIA':
            if self.liste_echelle_CUSTOM == ['MO']:
                for nom_CUSTOM,entite_CUSTOM in self.items():
                    CODE_CUSTOM = entite_CUSTOM.CODE_CUSTOM
                    for type_boite,dict_boite in entite_CUSTOM.items():
                        if dict_dict_info_CUSTOM[CODE_CUSTOM]['PPG_inclus_dans_integral_CUSTOM'] == True:
                            if dict_boite.orientation=="normal":
                                dict_boite.liste_echelle_boite = [entite_CUSTOM.echelle_base_REF]
                            if dict_boite.orientation=="orthogonal":
                                dict_boite.liste_echelle_boite = ['PPG','CUSTOM']
                        if dict_dict_info_CUSTOM[CODE_CUSTOM]['PPG_inclus_dans_integral_CUSTOM'] == False:
                            if dict_boite.orientation=="normal":
                                dict_boite.liste_echelle_boite = [entite_CUSTOM.echelle_base_REF,"PPG"]
                            if dict_boite.orientation=="orthogonal":
                                dict_boite.liste_echelle_boite = ['CUSTOM']
            if self.liste_echelle_CUSTOM == ['DEP']:
                for nom_CUSTOM,entite_CUSTOM in self.items():
                    CODE_CUSTOM = entite_CUSTOM.CODE_CUSTOM
                    for type_boite,dict_boite in entite_CUSTOM.items():
                        dict_boite.liste_echelle_boite = ["MO"]
            
        if self.type_donnees!='action':
            for nom_CUSTOM,entite_CUSTOM in self.items():
                CODE_CUSTOM = entite_CUSTOM.CODE_CUSTOM
                for nom_boite_maitre,dict_boite_complete in self.items():
                    self[CODE_CUSTOM]['liste_echelle_boite_normal'] = [self.echelle_base_REF]
        return self
    ##########################################################################################
    #Projet : Création des boites
    ##########################################################################################
    def creation_boite_projet_carto(self):
        if self.type_rendu=='carte':
            if self.liste_echelle_shp_CUSTOM_a_check==['MO']:
                taille_globale_carto = "petite"
            if self.liste_echelle_shp_CUSTOM_a_check==['DEP']:
                taille_globale_carto = "grande"

            if self.type_donnees=='toutes_pressions' and self.liste_echelle_base_REF==['ME']:
                for nom_CUSTOM,entite_CUSTOM in self.items():
                    entite_CUSTOM = DictCustomMaitre.creation_dict_boite_complete_normal(entite_CUSTOM,nom_boite_maitre='dict_boite_complete_pressions_globales')
                for nom_CUSTOM,entite_CUSTOM in self.items():
                    entite_CUSTOM = NomCUSTOMMaitre.creation_bloc_texte_simple(entite_CUSTOM,orientation="normal",type_icone='texte_simple',sous_type=['NOM_REF'],colonne_texte='nom_simple_REF',nom_CUSTOM=nom_CUSTOM)
                    entite_CUSTOM = NomCUSTOMMaitre.creation_bloc_icone(entite_CUSTOM,orientation="normal",type_icone='icone_pression',sous_type=["pressions"],colonne_nb_icone='NB_type_icone',nom_CUSTOM=nom_CUSTOM)

            if self.type_donnees=='action' and self.liste_echelle_shp_CUSTOM_a_check==['MO'] and any(x in self.liste_echelle_base_REF for x in ['ME','SME']):
                for nom_CUSTOM,entite_CUSTOM in self.items():
                    entite_CUSTOM = DictCustomMaitre.creation_dict_boite_complete_normal(entite_CUSTOM,nom_boite_maitre='dict_boite_complete_action_MIA')
                    entite_CUSTOM = DictCustomMaitre.creation_dict_boite_complete_orthogonal(entite_CUSTOM,nom_boite_maitre='dict_boite_complete_action_ortho',taille_globale_carto=taille_globale_carto)
                for nom_CUSTOM,entite_CUSTOM in self.items():
                    entite_CUSTOM = NomCUSTOMMaitre.creation_bloc_texte_simple(entite_CUSTOM,orientation="normal",type_icone='texte_simple',sous_type=['NOM_REF'],colonne_texte='nom_simple_REF',nom_CUSTOM=nom_CUSTOM)
                    entite_CUSTOM = NomCUSTOMMaitre.creation_bloc_texte_simple(entite_CUSTOM,orientation="orthogonal",type_icone='texte_simple',sous_type=['NOM_REF'],colonne_texte='nom_simple_REF',nom_CUSTOM=nom_CUSTOM)
                    entite_CUSTOM = NomCUSTOMMaitre.creation_bloc_icone(entite_CUSTOM,orientation="normal",type_icone='icone_action_MIA',sous_type=['nombre_actions','avancement'],colonne_nb_icone='NB_type_icone',nom_CUSTOM=nom_CUSTOM)
                    entite_CUSTOM = NomCUSTOMMaitre.creation_bloc_icone(entite_CUSTOM,orientation="orthogonal",type_icone='icone_action_MIA',sous_type=['nombre_actions','avancement'],colonne_nb_icone='NB_type_icone',nom_CUSTOM=nom_CUSTOM) 
                    entite_CUSTOM = NomCUSTOMMaitre.creation_bloc_lignes_multiples(entite_CUSTOM,orientation="normal",type_icone='ap_MIA',sous_type=['actions_phares'],colonne_texte='description_action_phare',nom_CUSTOM=nom_CUSTOM)
                    entite_CUSTOM = NomCUSTOMMaitre.creation_bloc_lignes_multiples(entite_CUSTOM,orientation="orthogonal",type_icone='ap_MIA',sous_type=['actions_phares'],colonne_texte='description_action_phare',nom_CUSTOM=nom_CUSTOM)
            if self.type_donnees=='action' and self.liste_echelle_shp_CUSTOM_a_check==['DEP'] and self.liste_echelle_base_REF==['MO']:
                for nom_CUSTOM,entite_CUSTOM in self.items():
                    entite_CUSTOM = DictCustomMaitre.creation_dict_boite_complete_normal(entite_CUSTOM,nom_boite_maitre='dict_boite_complete_action_MIA',taille_globale_carto=taille_globale_carto)
                    entite_CUSTOM = DictCustomMaitre.creation_dict_boite_complete_orthogonal(entite_CUSTOM,nom_boite_maitre='dict_boite_complete_action_ortho',taille_globale_carto=taille_globale_carto)
                for nom_CUSTOM,entite_CUSTOM in self.items():
                    entite_CUSTOM = NomCUSTOMMaitre.creation_bloc_texte_simple(entite_CUSTOM,orientation="normal",type_icone='texte_simple',sous_type=['NOM_REF'],colonne_texte='nom_simple_REF',nom_CUSTOM=nom_CUSTOM,taille_globale_carto=taille_globale_carto)
                    entite_CUSTOM = NomCUSTOMMaitre.creation_bloc_icone(entite_CUSTOM,orientation="normal",type_icone='icone_action_MIA',sous_type=['nombre_actions','pressions_MIA'],colonne_nb_icone='NB_type_icone',nom_CUSTOM=nom_CUSTOM,taille_globale_carto=taille_globale_carto)
                    entite_CUSTOM = NomCUSTOMMaitre.creation_bloc_texte_simple(entite_CUSTOM,orientation="orthogonal",type_icone='texte_simple',sous_type=['NOM_REF'],colonne_texte='nom_simple_REF',nom_CUSTOM=nom_CUSTOM,taille_globale_carto=taille_globale_carto)
                    entite_CUSTOM = NomCUSTOMMaitre.creation_bloc_icone(entite_CUSTOM,orientation="orthogonal",type_icone='icone_action_MIA',sous_type=['nombre_actions','pressions_MIA'],colonne_nb_icone='NB_type_icone',nom_CUSTOM=nom_CUSTOM,taille_globale_carto=taille_globale_carto) 
        return self
    
    def definition_liste_echelle_boite_projet_carto(self,dict_relation_shp_liste):
        if self.type_rendu=='carte':
            for nom_CUSTOM,entite_CUSTOM in self.items():
                if self.type_donnees=='action' and self.liste_echelle_shp_CUSTOM_a_check==['MO'] and any(x in self.liste_echelle_base_REF for x in ['ME','SME']):
                    for type_boite,dict_boite in entite_CUSTOM.items():
                        if len(dict_relation_shp_liste['dict_liste_PPG_par_CUSTOM'][entite_CUSTOM.CODE_CUSTOM])>1:
                            if dict_boite.orientation == "normal":
                                self[nom_CUSTOM][type_boite].liste_echelle_REF = ["PPG",entite_CUSTOM.echelle_base_REF]
                            if dict_boite.orientation == "orthogonal":
                                self[nom_CUSTOM][type_boite].liste_echelle_REF = ["MO"]   
                        if len(dict_relation_shp_liste['dict_liste_PPG_par_CUSTOM'][entite_CUSTOM.CODE_CUSTOM])<2:
                            if dict_boite.orientation == "normal":
                                self[nom_CUSTOM][type_boite].liste_echelle_REF = [entite_CUSTOM.echelle_base_REF]
                            if dict_boite.orientation == "orthogonal":
                                self[nom_CUSTOM][type_boite].liste_echelle_REF = ["PPG","MO"]                                                                                         
                if self.type_donnees=='action' and self.liste_echelle_shp_CUSTOM_a_check==['DEP'] and self.liste_echelle_base_REF==['MO']:
                    for type_boite,dict_boite in entite_CUSTOM.items():
                        if dict_boite.orientation == "normal":
                            dict_boite.liste_echelle_REF = ['MO']
                        if dict_boite.orientation == "orthogonal":
                            dict_boite.liste_echelle_REF = ['DEP']       
                if self.type_donnees=='toutes_pressions':
                    for type_boite,dict_boite in entite_CUSTOM.items():
                        if dict_boite.orientation == "normal":
                            dict_boite.liste_echelle_REF = ['ME']
        return self    

    def initialisation_bloc_avec_liste_entite_base_REF(dict_CUSTOM_maitre,dict_relation_shp_liste):
        for nom_CUSTOM,entite_CUSTOM in dict_CUSTOM_maitre.items():
            CODE_CUSTOM = entite_CUSTOM.CODE_CUSTOM
            if 'dict_boite_maitre_normal' in entite_CUSTOM:
                liste_echelle_boite_normal = entite_CUSTOM['dict_boite_maitre_normal'].liste_echelle_REF
                liste_df_boite_normal = []                
                for echelle_boite_normal in liste_echelle_boite_normal:
                    df_normal_CODE_REF_du_CUSTOM = pd.DataFrame(dict_relation_shp_liste['dict_liste_'+ echelle_boite_normal +'_par_CUSTOM'][CODE_CUSTOM], columns=['CODE_'+echelle_boite_normal])
                    df_normal_CODE_REF_du_CUSTOM = df_normal_CODE_REF_du_CUSTOM.rename({'CODE_'+echelle_boite_normal:'CODE_REF'},axis=1)
                    df_normal_CODE_REF_du_CUSTOM['echelle_REF']=echelle_boite_normal
                    liste_df_boite_normal.append(df_normal_CODE_REF_du_CUSTOM)
                df_normal_CODE_REF_du_CUSTOM = pd.concat(liste_df_boite_normal)
                for nom_boite_maitre,bloc in entite_CUSTOM['dict_boite_maitre_normal'].items():
                    bloc.df = df_normal_CODE_REF_du_CUSTOM

            
            if 'dict_boite_maitre_ortho' in entite_CUSTOM:
                liste_echelle_boite_ortho = entite_CUSTOM['dict_boite_maitre_ortho'].liste_echelle_REF
                liste_df_boite_ortho = []
                for echelle_boite_ortho in liste_echelle_boite_ortho:
                    if echelle_boite_ortho==entite_CUSTOM.echelle_REF:
                        df_ortho_CODE_REF_du_CUSTOM = pd.DataFrame([entite_CUSTOM.CODE_CUSTOM],columns=['CODE_REF'])
                    if echelle_boite_ortho!=entite_CUSTOM.echelle_REF:
                        df_ortho_CODE_REF_du_CUSTOM = pd.DataFrame(dict_relation_shp_liste['dict_liste_'+ echelle_boite_ortho +'_par_CUSTOM'][CODE_CUSTOM], columns=['CODE_'+echelle_boite_ortho])
                        df_ortho_CODE_REF_du_CUSTOM = df_ortho_CODE_REF_du_CUSTOM.rename({'CODE_'+echelle_boite_ortho:'CODE_REF'},axis=1)
                    df_ortho_CODE_REF_du_CUSTOM['echelle_REF']=echelle_boite_ortho
                    liste_df_boite_ortho.append(df_ortho_CODE_REF_du_CUSTOM)
                df_ortho_CODE_REF_du_CUSTOM = pd.concat(liste_df_boite_ortho)
                for nom_boite_maitre,bloc in entite_CUSTOM['dict_boite_maitre_ortho'].items():
                    bloc.df = df_ortho_CODE_REF_du_CUSTOM
        return dict_CUSTOM_maitre

    def repartition_df_donnees_dans_bloc(self,dict_df_donnees,dict_dict_info_REF,dict_decoupREF,dict_relation_shp_liste):
        for nom_CUSTOM,entite_CUSTOM in self.items():
            for nom_boite,dict_boite in entite_CUSTOM.items():
                for nom_bloc,dict_bloc in dict_boite.items():
                    if dict_bloc.type=='bloc_texte_simple':
                        dict_bloc.df = pd.merge(dict_bloc.df,dict_df_donnees['df_nom_REF_simple'],on='CODE_REF')
                    if dict_bloc.type=='bloc_icone' or dict_bloc.type=='bloc_lignes_multiples':
                        if self.type_donnees=="actions":
                            df_BDD_DORA_CUSTOM = DictDFTableauxActionsMIA.selection_actions_BDD_DORA(dict_df_donnees,entite_CUSTOM,dict_dict_info_REF,dict_decoupREF,dict_relation_shp_liste)
                            dict_bloc.df = pd.merge(dict_bloc.df,df_BDD_DORA_CUSTOM,on='CODE_REF')
                        if self.type_donnees=="toutes_pressions":
                            dict_bloc.df = pd.merge(dict_bloc.df,dict_df_donnees['df_pression'],on='CODE_REF')
                        if len(dict_bloc.df)==0:
                            print("attention, le "+ nom_bloc + " est vide pour le " + nom_boite)
        return self

    def traitement_special_bloc_ortho_si_plusieurs_echelle(self,dict_relation_shp_liste):
        for nom_CUSTOM,entite_CUSTOM in self.items():
            for nom_boite,dict_boite in entite_CUSTOM.items():
                if dict_boite.orientation=="orthogonal" and len(dict_boite.liste_echelle_REF)>1:
                    liste_echelle_REF = dictGdfCompletREF.hierarchisation_liste_echelle(dict_boite.liste_echelle_REF)
                    for nom_bloc,dict_bloc in dict_boite.items():
                        if dict_bloc.type=='bloc_texte_simple':
                            dict_bloc.df = dict_bloc.df.loc[dict_bloc.df["echelle_REF"]==liste_echelle_REF[-1]]
                        if dict_bloc.type=='bloc_icone' or dict_bloc.type=='bloc_lignes_multiples':
                            dict_bloc.df["echelle_REF"]==liste_echelle_REF[-1]
                            dict_bloc.df["CODE_REF"] = dict_relation_shp_liste['dict_liste_' + liste_echelle_REF[-1] + '_par_CUSTOM'][entite_CUSTOM.CODE_CUSTOM][0]
        return self    
    

    def traitement_bloc_avant_calcul_taille(self,dict_df_donnees,dict_relation_shp_liste):
        ##########################################################################################
        #Dict info bloc icone : fonction pour tableau pression
        ##########################################################################################
        def creation_liste_type_icone_toutes_pressions_toutes_pressions(self,public_cible):
            liste_categorie = ['dom','ind','azo','phy','irr','hyd']
            return(liste_categorie)

        def ajout_num_par_categorie_pression_toutes_pressions(self,liste_categorie):
            self[['icone_' + categorie for categorie in liste_categorie]] = self[['icone_' + categorie for categorie in liste_categorie]].replace(0, np.nan)
            self[['num_' + categorie for categorie in liste_categorie]] = self[['icone_' + categorie for categorie in liste_categorie]].cumsum(axis = 1)
            self[['icone_' + categorie for categorie in liste_categorie]] = self[['icone_' + categorie for categorie in liste_categorie]].replace(np.nan, 0)
            self[['num_' + categorie for categorie in liste_categorie]] = self[['num_' + categorie for categorie in liste_categorie]].replace(np.nan, 0)
            self[['num_' + categorie for categorie in liste_categorie]] = self[['num_' + categorie for categorie in liste_categorie]].astype(int)
            return self

        ##########################################################################################
        #Dict info bloc icone : fonction pour pression global
        ##########################################################################################
        def actualisation_liste_nom_colonne_a_garder_bloc_icone_pression_global(self,liste_categorie):
            self.liste_nom_colonne_a_garder.extend(['P_MORPHO','P_HYDRO','P_CONTI'])
            for categorie in liste_categorie:
                self.liste_nom_colonne_a_garder.extend(['icone_' + categorie])
            self.liste_nom_colonne_a_garder = list(set(self.liste_nom_colonne_a_garder))
            return self

        ##########################################################################################
        #Dict info bloc texte simple : Fonctions tableau texte simple
        ##########################################################################################
        for nom_entite_CUSTOM,contenu_CUSTOM in self.items():
            for nom_CUSTOM,dict_boite in contenu_CUSTOM.items():
                for nom_bloc,dict_bloc in dict_boite.items():
                    if dict_bloc.type == "bloc_texte_simple":
                        dict_bloc = BlocTexteSimple.decoupage_bloc_texte_ligne_simple(dict_bloc)

                ##########################################################################################
                #Dict info bloc icone : Fonctions bloc icone
                ##########################################################################################
                    if dict_bloc.type == "bloc_icone":
                        if dict_bloc.type_icone=="icone_action_MIA":
                            liste_categorie = BlocIcone.creation_liste_type_icone_nombres_actions_MIA(self)
                            dict_bloc = BlocIcone.conversion_typologie_actions_MIA(dict_bloc,self.public_cible,liste_categorie)    

                        if dict_bloc.type_icone=="icone_pression":
                            liste_categorie = BlocIcone.creation_liste_type_icone_toutes_pressions(self)

                        if "avancement" in dict_bloc.sous_type and self.thematique == "MIA":
                            dict_bloc = BlocIcone.conversion_niveau_avancement_tableau_actions(dict_bloc,annee_actuelle)
                        
                        liste_df_info_sup_icone = []
                        if "nombre_actions" in dict_bloc.sous_type and self.thematique == "MIA":  
                            df_bloc_avancement = BlocIcone.creation_dict_compte_chaque_categorie_et_avancement_nombres_actions(dict_bloc.df,liste_categorie,dict_bloc.avancement_max)
                            liste_df_info_sup_icone.append(df_bloc_avancement)

                        if "pressions_MIA" in dict_bloc.sous_type and self.thematique == "MIA":  
                            df_bloc_pressions_MIA = BlocIcone.indications_nombre_echelle_REF_avec_pression(dict_bloc,liste_categorie,dict_df_donnees,dict_boite.df_CODE_REF,dict_relation_shp_liste)                        
                            liste_df_info_sup_icone.append(df_bloc_pressions_MIA)

                        if dict_bloc.type_icone=="icone_action_MIA":
                            dict_bloc = BlocIcone.rassemblement_et_ajout_df_sup(dict_bloc,liste_categorie,liste_df_info_sup_icone,self.liste_contenu_bloc)
                        dict_bloc.df = BlocIcone.ajout_si_icone_par_categorie(dict_bloc.df,liste_categorie,dict_bloc.type_icone)
                        dict_bloc.df = BlocIcone.ajout_numero_par_binome_categorie(dict_bloc.df,liste_categorie,self.liste_contenu_bloc)
                        if dict_bloc.type_icone=="icone_pression":
                            dict_bloc.df = BlocIcone.garder_type_specifique_action(dict_bloc.df,liste_categorie)

                        if "nombre_actions" in dict_bloc.sous_type:
                            dict_bloc = BlocIcone.actualisation_liste_nom_colonne_a_garder_bloc_icone_nombres_actions(dict_bloc,liste_categorie)
                        if "avancement" in dict_bloc.sous_type:
                            dict_bloc = BlocIcone.actualisation_liste_nom_colonne_a_garder_bloc_icone_avancement(dict_bloc,liste_categorie) 
                        if "pressions" in dict_bloc.sous_type:
                            dict_bloc = BlocIcone.actualisation_liste_nom_colonne_a_garder_bloc_icone_pressions(dict_bloc,liste_categorie)                                                              

                    ##########################################################################################
                    #Dict info bloc phrases multiples : Fonctions bloc multiples lignes
                    ##########################################################################################
                    if dict_bloc.type == "bloc_lignes_multiples":
                        if "actions_phares" in dict_bloc.sous_type:
                            dict_bloc.df = BlocLignesMultiples.garder_actions_phares(dict_bloc.df,dict_bloc.colonne_texte)
                        if "actions_phares" in dict_bloc.sous_type and self.thematique == "MIA":
                            dict_bloc = BlocLignesMultiples.conversion_niveau_avancement_tableau_actions(dict_bloc,annee_actuelle)
                            dict_bloc = BlocLignesMultiples.conversion_typologie_actions_MIA(dict_bloc,self.public_cible,liste_categorie)
                            dict_bloc = BlocLignesMultiples.casse_lignes_multiples(dict_bloc)
                            dict_bloc = BlocLignesMultiples.decoupage_ligne_texte_indiv(dict_bloc)
                            dict_bloc = BlocLignesMultiples.actualisation_nom_colonne_a_garder_bloc_lm_ap(dict_bloc)
        return self

##########################################################################################
#Focntions travail geometries
##########################################################################################
    def suppression_blocs_vides(self):
        for CODE_CUSTOM in list(self):
            for nom_CUSTOM in list(self[CODE_CUSTOM]):
                for nom_bloc in list(self[CODE_CUSTOM][nom_CUSTOM]):
                    if len(self[CODE_CUSTOM][nom_CUSTOM][nom_bloc].df)==0:
                        del self[CODE_CUSTOM][nom_CUSTOM][nom_bloc]
        return self

    def ajout_infos_geometriques_decoupREF_boite_normal(self,dict_decoupREF,dict_df_buffer_CUSTOM):
        if self.type_rendu=='carte':
            for CODE_CUSTOM,contenu_CUSTOM in self.items():
                for nom_boite,dict_boite in contenu_CUSTOM.items():
                    if dict_boite.orientation == 'normal':
                        liste_tempo_REF = []
                        for echelle_REF in dict_boite.liste_echelle_REF:
                            df_info_decoupREF = dict_decoupREF['gdf_decoup' + echelle_REF + '_CUSTOM']
                            df_info_decoupREF = df_info_decoupREF.gdf.rename({x:x.replace(echelle_REF, 'REF') for x in list(df_info_decoupREF.gdf) if echelle_REF in x},axis=1)
                            liste_tempo_REF.append(df_info_decoupREF)
                        df_info_decoupREF_total = pd.concat(liste_tempo_REF)
                        
                        dict_boite.df = pd.merge(dict_boite.df,df_info_decoupREF_total.loc[df_info_decoupREF_total["CODE_CUSTOM"]==CODE_CUSTOM],on='CODE_REF',how='left',suffixes=[None,'_a_supprimer'])
                        dict_boite.df = dict_boite.df.loc[:,~dict_boite.df.columns.str.endswith('_a_supprimer')]
                        dict_boite.df = dict_boite.df.set_geometry('geometry')
                        dict_boite.df['centre_decoupREF'] = dict_boite.df.representative_point()
                        dict_boite.df['X_centre_decoupREF']=dict_boite.df['centre_decoupREF'].x
                        dict_boite.df['Y_centre_decoupREF']=dict_boite.df['centre_decoupREF'].y
                        dict_boite.df = pd.merge(dict_boite.df,self.df_info_CUSTOM,on='CODE_CUSTOM',suffixes=[None,'_a_supprimer'])
                        dict_boite.df = dict_boite.df.loc[:,~dict_boite.df.columns.str.endswith('_a_supprimer')]
                        dict_boite = dictGdfCompletREF.definition_gauche_droite_haut_bas_decoupREF_par_rapport_au_CUSTOM(dict_boite,dict_df_buffer_CUSTOM,contenu_CUSTOM.CODE_CUSTOM)
                        liste_colonne_info_geom = ['NOM_CUSTOM','echelle_REF','orient_CUSTOM','X_centre_decoupREF','Y_centre_decoupREF','orient_GD','orient_BH','X_centre_CUSTOM','Y_centre_CUSTOM']
                        dict_boite.liste_colonne_info_geom = liste_colonne_info_geom
                        dict_boite.df = dict_boite.df[liste_colonne_info_geom + ['CODE_REF']]
            return self
        else:
            return None
        
    def ajout_infos_boite_ortho(self,dict_decoupREF,dict_dict_info_CUSTOM,df_info_CUSTOM,projet):
        if projet.type_rendu=='carte':
            for nom_entite_CUSTOM,contenu_CUSTOM in self.items():
                for nom_boite,dict_boite in contenu_CUSTOM.items():
                    if dict_boite.orientation == 'orthogonal':
                        liste_tempo_REF = []
                        liste_echelle_ortho_sans_CUSTOM = dict_boite.liste_echelle_boite
                        liste_echelle_ortho_sans_CUSTOM = [x for x in liste_echelle_ortho_sans_CUSTOM if x !='CUSTOM']
                        if len(liste_echelle_ortho_sans_CUSTOM)>0:
                            for echelle_REF in liste_echelle_ortho_sans_CUSTOM:
                                df_info_decoupREF = dict_decoupREF['gdf_decoup' + echelle_REF + '_CUSTOM']
                                df_info_decoupREF = df_info_decoupREF.rename({x:x.replace(echelle_REF, 'REF') for x in list(df_info_decoupREF) if echelle_REF in x},axis=1)
                                liste_tempo_REF.append(df_info_decoupREF)
                            df_info_decoupREF_total = pd.concat(liste_tempo_REF)
                            dict_boite.df = pd.merge(dict_boite.df,df_info_decoupREF_total,on='CODE_REF',how='left')
                        df_info_CUSTOM_avec_CODE_REF = df_info_CUSTOM.rename({'CODE_CUSTOM':'CODE_REF'},axis=1)
                        dict_boite.df = pd.merge(dict_boite.df,df_info_CUSTOM_avec_CODE_REF,on='CODE_REF',suffixes=[None,'_a_supprimer'],how='left')
                        dict_boite.df = dict_boite.df.loc[:,~dict_boite.df.columns.str.endswith('_a_supprimer')]
                        liste_colonne_info_geom = ['NOM_CUSTOM','echelle_REF','orient_CUSTOM','X_centre_CUSTOM','Y_centre_CUSTOM']
                        dict_boite.liste_colonne_info_geom = liste_colonne_info_geom
                        dict_boite.df = dict_boite.df[liste_colonne_info_geom + ['CODE_REF']]
            return self
        
    def calcul_taille_bloc(self):
        for nom_entite_CUSTOM,contenu_CUSTOM in self.items():
            for nom_boite_maitre,dict_boite_maitre in contenu_CUSTOM.items():
                for nom_bloc,dict_bloc in dict_boite_maitre.items():
                    if dict_bloc.type == 'bloc_texte_simple':
                        if "NOM_REF" in dict_bloc.sous_type:
                            dict_bloc = BlocTexteSimple.calcul_taille_bloc_texte_simple(dict_bloc)
                    if dict_bloc.type == 'bloc_icone':
                        dict_bloc = BlocIcone.calcul_taille_bloc_icone(dict_bloc)
                    if dict_bloc.type == 'bloc_lignes_multiples':
                        dict_bloc = BlocLignesMultiples.calcul_taille_lignes_textes_multiples_indiv(dict_bloc)
                        #Dedoublement du df pour avoir ap indiv ET ap en groupe par REF de base
                        dict_bloc.df_indiv = dict_bloc.df
                        dict_bloc = BlocLignesMultiples.calcul_taille_bloc_lignes_multiples(dict_bloc)
        return self

    def calcul_taille_boite_complete(self):
        for nom_entite_CUSTOM,contenu_CUSTOM in self.items():
            contenu_CUSTOM = DictBoiteComplete.rassemblement_par_boite_complete(contenu_CUSTOM)
            contenu_CUSTOM = DictBoiteComplete.suppression_boite_complete_vide(contenu_CUSTOM)
        return self
    
    def ajout_contour_geometry_boite_complete(self,type_boite_placement:list=['normal']):
        for nom_entite_CUSTOM,contenu_CUSTOM in self.items():
            contenu_CUSTOM = DictBoiteComplete.creation_df_contour(contenu_CUSTOM,type_boite_placement)
            contenu_CUSTOM = DictBoiteComplete.import_hauteur_deviation(contenu_CUSTOM,type_boite_placement)
            contenu_CUSTOM= DictBoiteComplete.creation_contour_boite_complete(contenu_CUSTOM,type_boite_placement)
        return self   
    
    def replacement_eventuel_boite(self,type_boite_placement:list=['normal']):
        for nom_entite_CUSTOM,contenu_CUSTOM in self.items():
            contenu_CUSTOM = DictBoiteComplete.decalage_contour_si_contact(contenu_CUSTOM,type_boite_placement)
            contenu_CUSTOM = DictBoiteComplete.transfert_info_geom_dans_df(contenu_CUSTOM,type_boite_placement)
            '''for nom_boite_maitre,dict_boite_maitre in contenu_CUSTOM.items():
                if dict_boite_maitre.orientation=="normal":
                    dict_boite_maitre = DictBoiteComplete.actualisation_bas_haut_gauche_droite(dict_boite_maitre)
                    dict_boite_maitre = DictBoiteComplete.recuperation_bas_haut_gauche_droite_dans_df_contour(dict_boite_maitre)'''
        return self

    def actualisation_info_geom_dans_bloc(self,type_boite_placement:list=['normal']):
        #Faut actualiser gauche,droite et point interception par boite
        for nom_entite_CUSTOM,contenu_CUSTOM in self.items():
            for nom_boite_maitre,dict_boite_maitre in contenu_CUSTOM.items():
                if dict_boite_maitre.orientation in type_boite_placement:
                    df_boite_normal = dict_boite_maitre.df
                    for nom_bloc,dict_bloc in dict_boite_maitre.items():
                        dict_bloc.df = pd.merge(dict_bloc.df,df_boite_normal[["CODE_REF",'geometry_point_interception','haut_boite_complete','bas_boite_complete','gauche_boite_complete','droite_boite_complete']],on="CODE_REF",suffixes=['_a_supprimer',None])
                        dict_bloc.df = dict_bloc.df.loc[:,~dict_bloc.df.columns.str.endswith('_a_supprimer')]
        return self
    
    def suppression_CUSTOM_a_reduire_boite_complete(self,dict_dict_info_CUSTOM):
        dict_CODE_CUSTOM_a_enlever = {}
        for nom_boite_maitre,dict_boite_maitre in self.items():
            dict_CODE_CUSTOM_a_enlever[nom_boite_maitre] = []
            for CODE_CUSTOM,df_CUSTOM in dict_boite_maitre.boite_complete.items():
                if dict_dict_info_CUSTOM[CODE_CUSTOM]['CUSTOM_a_reduire'] == True:
                    dict_CODE_CUSTOM_a_enlever[nom_boite_maitre].append(CODE_CUSTOM)

        for nom_boite_maitre,dict_boite_maitre in dict_CODE_CUSTOM_a_enlever.items():
            for CODE_CUSTOM in dict_boite_maitre:
                del self[nom_boite_maitre].boite_complete[CODE_CUSTOM]

        dict_CODE_CUSTOM_a_enlever = {}
        for nom_boite_maitre,dict_boite_maitre in self.items():
            dict_CODE_CUSTOM_a_enlever[nom_boite_maitre] = {}
            for nom_bloc,dict_bloc in dict_boite_maitre.items():
                dict_CODE_CUSTOM_a_enlever[nom_boite_maitre][nom_bloc] = []
                for CODE_CUSTOM,df_CUSTOM in dict_bloc.items():
                    if dict_dict_info_CUSTOM[CODE_CUSTOM]['CUSTOM_a_reduire'] == True:
                        dict_CODE_CUSTOM_a_enlever[nom_boite_maitre][nom_bloc].append(CODE_CUSTOM)

        for nom_boite_maitre,dict_boite_maitre in dict_CODE_CUSTOM_a_enlever.items():
            for nom_bloc,dict_bloc in dict_boite_maitre.items():
                for CODE_CUSTOM in dict_bloc:
                    del self[nom_boite_maitre][nom_bloc][CODE_CUSTOM]

        return self        
    
#################################################################################################################
#Placement : Le boss niveau
#################################################################################################################
    def definition_point_ancrage_complet_REF_entre_eux(self,dict_df_buffer_CUSTOM):
        for CODE_CUSTOM,contenu_CUSTOM in self.items():
            dict_info_CUSTOM = self.df_info_CUSTOM.loc[self.df_info_CUSTOM['CODE_CUSTOM']==CODE_CUSTOM].to_dict(orient='records')[0]
            gdf_buffer = dict_df_buffer_CUSTOM[contenu_CUSTOM.CODE_CUSTOM]
            for nom_boite_maitre,dict_boite_maitre in contenu_CUSTOM.items():
                if dict_boite_maitre.orientation=="normal":
                    #On trace des lignes verticales si portrait, et horizontales si paysage
                    dict_boite_maitre = DictBoiteComplete.ajout_colonne_placement_boite_final(dict_boite_maitre)
                    dict_boite_maitre = DictBoiteComplete.placement_boite_complet_ME_entre_eux(dict_boite_maitre,dict_info_CUSTOM,"placement_boite_classique")
                    dict_boite_maitre = DictBoiteComplete.tracer_ligne_pour_intersection_buffer(dict_boite_maitre,dict_info_CUSTOM,"placement_boite_classique")
                    #On a placé des lignes (méridiens ou paralléles suivants paysage ou portrait)
                    dict_boite_maitre = DictBoiteComplete.intersection_ligne_buffer(dict_boite_maitre,gdf_buffer,dict_info_CUSTOM,"placement_boite_classique")
                    dict_boite_maitre = DictBoiteComplete.gestion_erreurs_interceptions_ligne_buffer(dict_boite_maitre,dict_info_CUSTOM,"placement_boite_classique")
                    dict_boite_maitre = DictBoiteComplete.extraction_liste_coord_apres_interception(dict_boite_maitre,dict_info_CUSTOM,"placement_boite_classique")
                    dict_boite_maitre = DictBoiteComplete.actualisation_orient_GD_et_BH(dict_boite_maitre,dict_info_CUSTOM)
                    dict_boite_maitre = DictBoiteComplete.calcul_nombre_boite_qui_depassent_a_deplacer(dict_boite_maitre,dict_info_CUSTOM)
                    dict_boite_maitre = DictBoiteComplete.calcul_valeur_limite_si_boite_a_replacer(dict_boite_maitre,dict_info_CUSTOM)
                    dict_boite_maitre = DictBoiteComplete.actualisation_hauteur_et_largeur_boites_normales_qui_depassent(dict_boite_maitre,dict_info_CUSTOM)
                    dict_boite_maitre = DictBoiteComplete.placement_boite_complet_ME_entre_eux(dict_boite_maitre,dict_info_CUSTOM,"placement_boite_extremite_qui_depassent")
                    dict_boite_maitre = DictBoiteComplete.tracer_ligne_pour_intersection_buffer(dict_boite_maitre,dict_info_CUSTOM,"placement_boite_extremite_qui_depassent")
                    dict_boite_maitre = DictBoiteComplete.intersection_ligne_buffer(dict_boite_maitre,gdf_buffer,dict_info_CUSTOM,"placement_boite_extremite_qui_depassent")
                    dict_boite_maitre = DictBoiteComplete.gestion_erreurs_interceptions_ligne_buffer(dict_boite_maitre,dict_info_CUSTOM,"placement_boite_extremite_qui_depassent")
                    dict_boite_maitre = DictBoiteComplete.extraction_liste_coord_apres_interception(dict_boite_maitre,dict_info_CUSTOM,"placement_boite_extremite_qui_depassent")
                    dict_boite_maitre = DictBoiteComplete.calcul_limite_boites_extremitees(dict_boite_maitre,dict_info_CUSTOM)
                    dict_boite_maitre = DictBoiteComplete.replacement_boites_extremitees(dict_boite_maitre,dict_info_CUSTOM)
                    dict_boite_maitre = DictBoiteComplete.creation_point_unique_par_REF(dict_boite_maitre,dict_info_CUSTOM)
                    '''dict_boite_maitre = DictBoiteComplete.actualisation_type_placement_boite_final(dict_boite_maitre,dict_info_CUSTOM)
                    dict_boite_maitre = DictBoiteComplete.actualisation_bas_haut_gauche_droite(dict_boite_maitre)
                    dict_boite_maitre = DictBoiteComplete.decalage_final(dict_boite_maitre)'''

                    dict_boite_maitre = DictBoiteComplete.actualisation_type_placement_boite_final(dict_boite_maitre,dict_info_CUSTOM)
                    dict_boite_maitre = DictBoiteComplete.actualisation_bas_haut_gauche_droite(dict_boite_maitre)
        return self

    ##########################################################################################
    #Placement intérieur boite
    ##########################################################################################
    def recuperation_infos_geom_dans_bloc(self,type_boite_placement:list=['normal']):
        for nom_entite_CUSTOM,contenu_CUSTOM in self.items():
            for nom_boite_maitre,dict_boite_maitre in contenu_CUSTOM.items():
                dict_boite_maitre = DictBoiteComplete.ajout_info_point_interception_et_cote(dict_boite_maitre,type_boite_placement)
                dict_boite_maitre = DictBoiteComplete.creation_dict_hauteur_largeur_blocs_dans_meme_boite(dict_boite_maitre,type_boite_placement)
                dict_boite_maitre = DictBoiteComplete.ajout_ecart_origine_et_numero_bloc(dict_boite_maitre,type_boite_placement)
        return self

    def placement_bloc_interieur_boite_complete(self,type_boite_placement):
        for nom_entite_CUSTOM,contenu_CUSTOM in self.items():
            for nom_boite_maitre,dict_boite_maitre in contenu_CUSTOM.items():
                if dict_boite_maitre.orientation in type_boite_placement:
                    dict_boite_maitre = DictBoiteComplete.placement_des_blocs_interieur_boite_complete(dict_boite_maitre,type_boite_placement)
                    dict_boite_maitre = DictBoiteComplete.placement_objets_interieur_blocs(dict_boite_maitre,type_boite_placement)
                    dict_boite_maitre = DictBoiteComplete.ajout_colonne_pour_orientation_et_alignement_objet(dict_boite_maitre)
                    dict_boite_maitre = DictBoiteComplete.conversion_colonne_hauteur_largeur_boite_complete_vers_bloc(dict_boite_maitre,type_boite_placement)
        return self


    ##########################################################################################
    #PARTIE ortho
    ##########################################################################################
    def placement_eventuel_boite_ortho(self,type_boite_placement:list=['orthogonal']):
        for CODE_CUSTOM,contenu_CUSTOM in self.items():
            dict_info_CUSTOM = self.df_info_CUSTOM.loc[self.df_info_CUSTOM['CODE_CUSTOM']==CODE_CUSTOM].to_dict(orient='records')[0]
            for nom_boite_maitre,dict_boite_maitre in contenu_CUSTOM.items():
                if dict_boite_maitre.orientation in type_boite_placement:
                    dict_boite_maitre = DictBoiteComplete.placement_boite_ortho(dict_boite_maitre,dict_info_CUSTOM,contenu_CUSTOM)
                    dict_boite_maitre = DictBoiteComplete.actualisation_info_boite_ortho(dict_boite_maitre)
                    dict_boite_maitre = DictBoiteComplete.actualisation_bas_haut_gauche_droite(dict_boite_maitre)
                    dict_boite_maitre = DictBoiteComplete.recuperation_bas_haut_gauche_droite_dans_df_contour(dict_boite_maitre)
                    dict_boite_maitre = DictBoiteComplete.augmentation_esthetique_df_contour(dict_boite_maitre)
        return self


    ##########################################################################################
    #Projet : Création des dossiers
    ##########################################################################################
    def creation_dossiers_projet(self,liste_echelle_REF_projet):
        for nom_couche_REF in liste_echelle_REF_projet:
            if not os.path.exists("/mnt/g/travail/carto/projets basiques/PAOT global 5.0/" + self.nom_dossier_maitre + '/couche_CUSTOM_tempo'):
                os.makedirs("/mnt/g/travail/carto/projets basiques/PAOT global 5.0/" + self.nom_dossier_maitre + '/couche_CUSTOM_tempo')
            if not os.path.exists("/mnt/g/travail/carto/projets basiques/PAOT global 5.0/" + self.nom_dossier_maitre + '/couche_geom'):
                os.makedirs("/mnt/g/travail/carto/projets basiques/PAOT global 5.0/" + self.nom_dossier_maitre + '/couche_geom')
            if not os.path.exists("/mnt/g/travail/carto/projets basiques/PAOT global 5.0/" + self.nom_dossier_maitre + '/couche_bloc'):
                os.makedirs("/mnt/g/travail/carto/projets basiques/PAOT global 5.0/" + self.nom_dossier_maitre + '/couche_bloc')
            if not os.path.exists("/mnt/g/travail/carto/projets basiques/PAOT global 5.0/" + self.nom_dossier_maitre + '/couche_boite'):
                os.makedirs("/mnt/g/travail/carto/projets basiques/PAOT global 5.0/" + self.nom_dossier_maitre + '/couche_boite')
            if not os.path.exists("/mnt/g/travail/carto/projets basiques/PAOT global 5.0/" + self.nom_dossier_maitre + '/couche_fleche'):
                os.makedirs("/mnt/g/travail/carto/projets basiques/PAOT global 5.0/" + self.nom_dossier_maitre + '/couche_fleche')

    ##########################################################################################
    #CUSTOM
    ##########################################################################################
    def creation_entite_CUSTOM(self,dict_dict_info_REF=None,dict_geom_REF=None,liste_CODE_CUSTOM=None,TYPE_REF=None):
        for CODE_CUSTOM in liste_CODE_CUSTOM:
            self[CODE_CUSTOM] = NomCUSTOMMaitre({})
            self[CODE_CUSTOM] = NomCUSTOMMaitre.attributs_CUSTOM(self[CODE_CUSTOM],self.liste_echelle_shp_CUSTOM_a_check,CODE_CUSTOM,dict_dict_info_REF,TYPE_REF)
            self[CODE_CUSTOM] = NomCUSTOMMaitre.chercher_gdf_CUSTOM(self[CODE_CUSTOM],dict_geom_REF)

        return self    

    def attributs_echelle_base_REF(self,dict_dict_info_REF):
        if self.echelle_base_REF!=None:
                for nom_CUSTOM,entite_CUSTOM in self.items():
                    entite_CUSTOM.echelle_base_REF = self.echelle_base_REF

        if self.echelle_base_REF==None:
            if (self.type_rendu=='carte' and self.type_donnees=='toutes_pressions'):
                for nom_CUSTOM,entite_CUSTOM in self.items():
                    entite_CUSTOM.echelle_base_REF = 'ME'
            if self.type_rendu=='tableau_MAJ_Osmose':
                for nom_CUSTOM,entite_CUSTOM in self.items():
                    entite_CUSTOM.echelle_base_REF = 'ME'
            if ((self.type_rendu=='carte' or self.type_rendu=='tableau_DORA_vers_BDD' or self.type_rendu=='verif_tableau_DORA' or  self.type_rendu=='tableau_vierge')
                and self.type_donnees=='action'):
                df_info_SME = dict_dict_info_REF['df_info_SME']
                liste_CODE_CUSTOM_avec_echelle_SME = list(set(df_info_SME['MO_maitre'].to_list()))
                for nom_CUSTOM,entite_CUSTOM in self.items():
                    list_echelle_base_REF = []
                    if entite_CUSTOM.CODE_CUSTOM in liste_CODE_CUSTOM_avec_echelle_SME:
                        print("Pour "+nom_CUSTOM+", j'ai trouvé des SOUS ME !")
                        entite_CUSTOM.echelle_base_REF = "SME"
                    
                    if entite_CUSTOM.CODE_CUSTOM not in liste_CODE_CUSTOM_avec_echelle_SME:
                        print("Pour "+nom_CUSTOM+", j'ai pas trouvé des SOUS ME !")
                        entite_CUSTOM.echelle_base_REF = "ME"

        return self
    
    def creation_gdf_CUSTOM(self):
        liste_gdf_CUSTOM = [entite_CUSTOM.gdf for nom_CUSTOM,entite_CUSTOM in self.items()]
        self.gdf_CUSTOM = pd.concat(liste_gdf_CUSTOM)
        return self

    def ajout_df_info_CUSTOM(self):
        if self.type_rendu=="carte":
            list_colonne_a_garder_issues_gdf_gros_CUSTOM = ["CODE_CUSTOM","NOM_CUSTOM","echelle",'X_centre_CUSTOM','Y_centre_CUSTOM','orient_CUSTOM','min_x_CUSTOM','min_y_CUSTOM','max_x_CUSTOM','max_y_CUSTOM']
            gdf_CUSTOM = self.gdf_CUSTOM
            df_info_CUSTOM_tempo = gdf_CUSTOM.join(gdf_CUSTOM.bounds)
            df_info_CUSTOM_tempo = df_info_CUSTOM_tempo.rename({"maxx":'max_x_CUSTOM',"maxy":'max_y_CUSTOM',"minx":'min_x_CUSTOM',"miny":'min_y_CUSTOM'},axis=1)
            df_info_CUSTOM_tempo['taille_GD']=df_info_CUSTOM_tempo['max_y_CUSTOM']-df_info_CUSTOM_tempo['min_y_CUSTOM']
            df_info_CUSTOM_tempo['taille_BH']=df_info_CUSTOM_tempo['max_x_CUSTOM']-df_info_CUSTOM_tempo['min_x_CUSTOM']
            #Orientation
            df_info_CUSTOM_tempo.loc[df_info_CUSTOM_tempo['taille_GD']>df_info_CUSTOM_tempo['taille_BH'],'orient_CUSTOM']='portrait'
            df_info_CUSTOM_tempo.loc[df_info_CUSTOM_tempo['taille_GD']<df_info_CUSTOM_tempo['taille_BH'],'orient_CUSTOM']='paysage'

            #Echelle
            dict_NOM_CUSTOM_echelle = {k:v.echelle_carto_globale for k,v in self.items()}
            df_info_CUSTOM_tempo['echelle'] = df_info_CUSTOM_tempo['CODE_CUSTOM'].map(dict_NOM_CUSTOM_echelle)

            #Centre
            df_info_CUSTOM_tempo['geometry_centre_CUSTOM'] = df_info_CUSTOM_tempo.representative_point()
            df_info_CUSTOM_tempo['X_centre_CUSTOM']=df_info_CUSTOM_tempo['geometry_centre_CUSTOM'].x
            df_info_CUSTOM_tempo['Y_centre_CUSTOM']=df_info_CUSTOM_tempo['geometry_centre_CUSTOM'].y
            df_info_CUSTOM = df_info_CUSTOM_tempo.reset_index()[list_colonne_a_garder_issues_gdf_gros_CUSTOM]
            df_info_CUSTOM["boite_a_replacer"] = False
            #ajout de l'ID
            df_info_CUSTOM['id_atlas'] = df_info_CUSTOM['CODE_CUSTOM'] + '%' + self.type_donnees + '%' + self.thematique + '%' + self.public_cible
            self.df_info_CUSTOM = df_info_CUSTOM
        return self

    def ajout_attributs_projet(self):
        def ajout_liste_categorie_icone(self):
            liste_categorie = []
            if self.type_rendu == "carte" and self.type_donnees == "action":
                if self.thematique =="MIA":
                    if self.public_cible == "elu" or self.public_cible == 'prefet':
                        liste_categorie = ['hyd','mor','con','ino','gou']
                    if self.public_cible == 'tech':
                        liste_categorie = ['rip','mor','con','mob', 'ZH','rui','ino','gou']
                if self.thematique =="ASS":
                    liste_categorie = ['etu','plu','tra','ANC','con','aut']
            if self.type_rendu == "carte" and self.type_donnees == "toutes_pressions":
                liste_categorie = ['dom','ind','azo','phy','irr','hyd']
            self.liste_categorie = liste_categorie
            return self

        #Attribution CODE MO pour chaque entite du CUSTOM
        self.nom_entite_CUSTOM = self.gdf_shp_CUSTOM.nom_entite_REF
        self.nom_geometry_CUSTOM = self.gdf_shp_CUSTOM.colonne_geometry
        self.dict_nom_CUSTOM_CODE_CUSTOM = dict(zip(self.gdf_shp_CUSTOM['NOM_CUSTOM'].to_list(),self.gdf_shp_CUSTOM['CODE_CUSTOM']))
        self.liste_nom_CUSTOM = self.gdf_shp_CUSTOM['NOM_CUSTOM'].to_list()
        self.liste_CODE_CUSTOM = self.gdf_shp_CUSTOM['CODE_CUSTOM'].tolist()
        self = ajout_liste_categorie_icone(self)
        return self

    def actualisation_liste_CUSTOM_projet(self,dict_dict_boite_maitre):
        if self.type_rendu=='carte' and self.type_donnees == 'action' and self.thematique == 'MIA':
            self.liste_CODE_CUSTOM_tableau_actions = [self.dict_nom_CUSTOM_CODE_CUSTOM[x] for x in self.liste_nom_CUSTOM_tableau_actions if x in self.dict_nom_CUSTOM_CODE_CUSTOM]
            self.liste_CODE_CUSTOM = [x for x in self.liste_CODE_CUSTOM if x in self.liste_CODE_CUSTOM_tableau_actions]
        return self

    def creation_dict_dict_info_CUSTOM(self,df_info_CUSTOM):
        dict_dict_info_CUSTOM = df_info_CUSTOM.set_index('CODE_CUSTOM').to_dict('index')
        return dict_dict_info_CUSTOM

    def attributs_dict_dict_info_CUSTOM(self,dict_dict_info_CUSTOM):
        for CUSTOM in dict_dict_info_CUSTOM:
            dict_dict_info_CUSTOM[CUSTOM]['CUSTOM_a_reduire']=False
            dict_dict_info_CUSTOM[CUSTOM]['boite_a_replacer']=False
            dict_dict_info_CUSTOM[CUSTOM]['PPG_inclus_dans_integral_CUSTOM']=False
            dict_dict_info_CUSTOM[CUSTOM]['cartouche_boite_ortho_separe']=False
        return dict_dict_info_CUSTOM

    def ajout_info_df_parties_CUSTOM(self,dict_decoupCUSTOM,dict_dict_info_CUSTOM,dict_dict_info_REF):
        df_nom_ME_simple = dataframe.recuperation_tableaux_nom_ME_simple()
        dict_mapping_geometry = {}
        dict_mapping_nom_partie = {}
        liste_echelle_shp_par_CUSTOM = self.gdf_shp_CUSTOM.liste_echelle_shp_par_CUSTOM
        for CODE_CUSTOM,dict_partie_CUSTOM in dict_decoupCUSTOM.items():
            dict_mapping_nom_partie[CODE_CUSTOM] = dict_decoupCUSTOM[CODE_CUSTOM].index.tolist()
            dict_mapping_geometry[CODE_CUSTOM]= dict_decoupCUSTOM[CODE_CUSTOM]['geometry'].to_list()
            dict_tempo_mapping_CODE_CUSTOM = {CODE_CUSTOM:dict_decoupCUSTOM[CODE_CUSTOM]['CODE_CUSTOM'].to_list()}
            self.gdf_shp_CUSTOM["CODE_CUSTOM_tempo"] = self.gdf_shp_CUSTOM["CODE_CUSTOM"].map(dict_tempo_mapping_CODE_CUSTOM).fillna(self.gdf_shp_CUSTOM["CODE_CUSTOM"])
            self.gdf_shp_CUSTOM = self.gdf_shp_CUSTOM.explode("CODE_CUSTOM_tempo")
            self.gdf_shp_CUSTOM = self.gdf_shp_CUSTOM.reset_index(drop=True)
            self.gdf_shp_CUSTOM.loc[self.gdf_shp_CUSTOM["CODE_CUSTOM"]==CODE_CUSTOM,"geometry_CUSTOM"] = dict_decoupCUSTOM[CODE_CUSTOM]['geometry'].to_list()
            self.gdf_shp_CUSTOM.loc[self.gdf_shp_CUSTOM["CODE_CUSTOM"]==CODE_CUSTOM,"ALIAS"] = self.gdf_shp_CUSTOM["CODE_CUSTOM_tempo"]
            self.gdf_shp_CUSTOM.loc[self.gdf_shp_CUSTOM["CODE_CUSTOM"]==CODE_CUSTOM,"NOM_MO"] = self.gdf_shp_CUSTOM["CODE_CUSTOM_tempo"]
            self.gdf_shp_CUSTOM.loc[self.gdf_shp_CUSTOM["CODE_CUSTOM"]==CODE_CUSTOM,"CODE_CUSTOM"] = self.gdf_shp_CUSTOM["CODE_CUSTOM"] + "é" + self.gdf_shp_CUSTOM["CODE_CUSTOM_tempo"]
        self.gdf_shp_CUSTOM = ListGdfCUSTOM(self.gdf_shp_CUSTOM)
        self.gdf_shp_CUSTOM.attribution_GdfCompletREF('CUSTOM')
        self.gdf_shp_CUSTOM.liste_echelle_shp_par_CUSTOM = list(set(liste_echelle_shp_par_CUSTOM))
        self.gdf_shp_CUSTOM['surface_MO'] = self.gdf_shp_CUSTOM.area
        
        '''self.gdf_shp_CUSTOM['nom_partie'] = self.gdf_shp_CUSTOM.apply(lambda x: dict_map_CODE_ME_nom_simple[x['NOM_MO'].replace("partie_","").split('$')[0]] if x['NOM_MO'].startswith('partie_') else x['ALIAS'], axis=1)
        self.gdf_shp_CUSTOM.loc[self.gdf_shp_CUSTOM['CODE_CUSTOM'].isin(dict_mapping_geometry),'ALIAS'] = self.gdf_shp_CUSTOM['ALIAS'] + ' ' + self.gdf_shp_CUSTOM['nom_partie']
        self.gdf_shp_CUSTOM['NOM_MO'] = self.gdf_shp_CUSTOM.apply(lambda x: dict_map_CODE_CUSTOM_NOM_MO[x['CODE_CUSTOM']] + ' ' + dict_map_CODE_ME_nom_simple[x['NOM_MO'].replace("partie_","").split('$')[0]] if x['NOM_MO'].startswith('partie_') else x['NOM_MO'], axis=1)
        self.gdf_shp_CUSTOM.loc[self.gdf_shp_CUSTOM['CODE_CUSTOM'].isin(dict_mapping_geometry),'CODE_CUSTOM'] = self.gdf_shp_CUSTOM['CODE_CUSTOM'] + '_' + self.gdf_shp_CUSTOM['nom_partie']'''
        return self

    def actualisation_attributs_projet_apres_decoupage_CUSTOM(self):
        self.liste_nom_CUSTOM = self.liste_nom_CUSTOM_tableau_actions
        self.liste_CODE_CUSTOM = self.liste_CODE_CUSTOM_tableau_actions
        return self

    def ajout_attributs_CUSTOM_reduit(self,dict_decoupCUSTOM,dict_dict_info_CUSTOM,dict_dict_info_REF):
        for CODE_CUSTOM,dict_decoup_CUSTOM in dict_decoupCUSTOM.items():
            for nom_CUSTOM,dict_df_partie_CUSTOM in dict_decoup_CUSTOM.items():
                #C'est sale mais chaque gdf n'a qu'une ligne donc bon...
                self.dict_nom_CUSTOM_CODE_CUSTOM = dict(zip(self.gdf_shp_CUSTOM['NOM_MO'].to_list(),self.gdf_shp_CUSTOM['CODE_CUSTOM'].to_list()))
                self.liste_CODE_CUSTOM_tableau_actions = self.gdf_shp_CUSTOM['CODE_CUSTOM'].to_list()
                self.liste_nom_CUSTOM_tableau_actions = self.gdf_shp_CUSTOM['NOM_MO'].to_list()
        return self

########################################################################################################
#PARTIE atlas
########################################################################################################
    def definition_colonne_a_garder_pour_export_vers_QGIS_bloc(self):
        for nom_entite_CUSTOM,contenu_CUSTOM in self.items():
            for type_boite,contenu_boite in contenu_CUSTOM.items():
                for nom_bloc,dict_bloc in contenu_boite.items():
                    dict_bloc.liste_nom_colonne_a_garder.extend(['CODE_REF','echelle_REF','id_atlas'])
                    dict_bloc.liste_nom_colonne_a_garder.extend(['sens',"type_placement_boite_final"])
                    if dict_bloc.type=='bloc_texte_simple':
                        dict_bloc.liste_nom_colonne_a_garder.extend(['geometry_bloc_texte_simple','nom_police','taille_police','ls_decoup_texte_simple','gauche_bloc','droite_bloc','haut_bloc','bas_bloc','alignement','Quadrant'])
                    if dict_bloc.type=='bloc_icone':
                        dict_bloc.liste_nom_colonne_a_garder.extend(['geometry_bloc_icone','gauche_bloc_bloc_icone','droite_bloc_bloc_icone','haut_bloc_bloc_icone','bas_bloc_bloc_icone','largeur_icone','hauteur_icone'])
                    if dict_bloc.type=='bloc_lignes_multiples':
                        dict_bloc.liste_nom_colonne_a_garder.extend(['geometry_point_bloc_lignes_multiples','nom_police','taille_police','lm_decoup_bloc_ligne_multiples','alignement','Quadrant','al_lm_ind','angle'])
        return self

   
    def definition_colonne_a_garder_pour_export_vers_QGIS_boite(self):
        for nom_entite_CUSTOM,contenu_CUSTOM in self.items():
            for type_boite,contenu_boite in contenu_CUSTOM.items():
                contenu_boite.liste_nom_colonne_a_garder = []
                contenu_boite.liste_nom_colonne_a_garder.extend(['CODE_REF','echelle_REF','id_atlas','geom_boite','type_placement_boite_final'])
        return self    
    
    def reduction_nom_colonne_pour_export_QGIS(self):
        for nom_entite_CUSTOM,contenu_CUSTOM in self.items():
            for type_boite,contenu_boite in contenu_CUSTOM.items():
                if contenu_boite.orientation=="normal":
                    contenu_boite.df_contour = dataframe.reduction_nom_colonne_via_fichier_csv(contenu_boite.df_contour,"df_contour")
        for nom_entite_CUSTOM,contenu_CUSTOM in self.items():
            for type_boite,contenu_boite in contenu_CUSTOM.items():
                    for type_boite,contenu_boite in contenu_CUSTOM.items():
                        for type_bloc,contenu_bloc in contenu_boite.items():
                            if hasattr(contenu_bloc,"df_indiv"):
                                contenu_bloc.df_indiv = dataframe.reduction_nom_colonne_via_fichier_csv(contenu_bloc.df_indiv,contenu_bloc.type)  
                            if not hasattr(contenu_bloc,"df_indiv"):    
                                contenu_bloc.df = dataframe.reduction_nom_colonne_via_fichier_csv(contenu_bloc.df,contenu_bloc.type)                        
        return self

    def generation_col_atlas(self):
        self.gdf_CUSTOM['id_atlas'] = self.gdf_CUSTOM['CODE_CUSTOM'] + '%' + self.type_donnees + '%' + self.thematique + '%' + self.public_cible + '%' + self.info_fond_carte
        self.gdf_fond_carte['id_atlas'] = self.gdf_fond_carte['CODE_CUSTOM'] + '%' + self.type_donnees + '%' + self.thematique + '%' + self.public_cible + '%' + self.info_fond_carte
        for nom_entite_CUSTOM,contenu_CUSTOM in self.items():
            for type_boite,contenu_boite in contenu_CUSTOM.items():
                contenu_boite.df_contour['id_atlas'] = contenu_CUSTOM.CODE_CUSTOM + '%' + self.type_donnees + '%' + self.thematique + '%' + self.public_cible + '%' + self.info_fond_carte
        for nom_entite_CUSTOM,contenu_CUSTOM in self.items():
            for type_boite,contenu_boite in contenu_CUSTOM.items():
                for type_bloc,contenu_bloc in contenu_boite.items():                                  
                    contenu_bloc.df['id_atlas'] = contenu_CUSTOM.CODE_CUSTOM + '%' + self.type_donnees + '%' + self.thematique + '%' + self.public_cible + '%' + self.info_fond_carte
        return self

    def transfert_eventuel_info_bloc_df_vers_bloc_indiv(self):
        for nom_entite_CUSTOM,contenu_CUSTOM in self.items():
            for nom_boite_maitre,dict_boite_maitre in contenu_CUSTOM.items():
                dict_boite_maitre = DictBoiteComplete.transfert_info_df_vers_df_indiv(dict_boite_maitre)
        return self
    
    def garder_colonne_de_attributs_colonne_a_garder(self):
        for nom_entite_CUSTOM,contenu_CUSTOM in self.items():
            for type_boite,contenu_boite in contenu_CUSTOM.items():
                if contenu_boite.orientation=="normal":
                    contenu_boite.df_contour = contenu_boite.df_contour[contenu_boite.liste_nom_colonne_a_garder]
        for nom_entite_CUSTOM,contenu_CUSTOM in self.items():
            for type_boite,contenu_boite in contenu_CUSTOM.items():
                    for type_boite,contenu_boite in contenu_CUSTOM.items():
                        for type_bloc,contenu_bloc in contenu_boite.items():
                            if contenu_bloc.type=="bloc_lignes_multiples":
                                contenu_bloc.df_indiv = contenu_bloc.df_indiv[contenu_bloc.liste_nom_colonne_a_garder] 
                            else:
                                contenu_bloc.df = contenu_bloc.df[contenu_bloc.liste_nom_colonne_a_garder]
        return self                               

    def recuperer_donnes_df_info_CUSTOM_sur_atlas(self,df_info_CUSTOM):
        self.gdf_shp_CUSTOM = pd.merge(self.gdf_shp_CUSTOM,df_info_CUSTOM,on=['CODE_CUSTOM','NOM_CUSTOM'],suffixes=[None,'_a_supprimer'])
        self.gdf_shp_CUSTOM = self.gdf_shp_CUSTOM.loc[:,~self.gdf_shp_CUSTOM.columns.str.endswith('_a_supprimer')]
        return self

    def garder_colonne_de_attributs_colonne_a_garder_atlas(self):
        liste_colonne_a_garder = ['NOM_CUSTOM','ALIAS', 'geometry_CUSTOM', 'CODE_CUSTOM', 'echelle', 'X_centre_CUSTOM', 'Y_centre_CUSTOM', 'orient_CUSTOM', 'id_atlas', 'geom_princ', 'min_x_CUSTOM', 'max_x_CUSTOM', 'min_y_CUSTOM', 'max_y_CUSTOM']
        if len(self.liste_echelle_fond_carte)==2:
            liste_colonne_a_garder.append('geom_sec')
        self.gdf_shp_CUSTOM = self.gdf_shp_CUSTOM[liste_colonne_a_garder]
        return self
    ##########################################################################################
    #CUSTOM spécial : département
    ##########################################################################################
    def import_shp_CUSTOM_departement(self,dep,gdf_BVG,df_info_BVG):
        liste_gdf_tempo = []
        df_info_tempo = df_info_BVG.loc[df_info_BVG.apply(lambda x : dep in x["list_dep"],axis=1)]
        list_BVG_dep = df_info_tempo['CODE_BVG'].to_list()
        gdf_BVG_tempo = gdf_BVG.loc[gdf_BVG['CODE_BVG'].isin(list_BVG_dep)]
        gdf_BVG_tempo = gdf_BVG_tempo.dissolve()
        gdf_BVG_tempo['NOM_CUSTOM'] = str(dep)
        gdf_BVG_tempo['CODE_CUSTOM'] = "dep " + str(dep)
        gdf_BVG_tempo = gdf_BVG_tempo.rename({"geometry_BVG":'geometry_CUSTOM'},axis=1)
        gdf_BVG_tempo = gdf_BVG_tempo.set_geometry('geometry_CUSTOM')
        gdf_BVG_tempo = gdf_BVG_tempo[["CODE_CUSTOM",'NOM_CUSTOM','geometry_CUSTOM']]
        liste_gdf_tempo.append(gdf_BVG_tempo)
            
        gdf_gros_CUSTOM = pd.concat(liste_gdf_tempo)
        gdf_gros_CUSTOM = gdf_gros_CUSTOM.set_geometry('geometry_CUSTOM')
        gdf_gros_CUSTOM = gdf_gros_CUSTOM.reset_index()
        gdf_gros_CUSTOM = ListGdfCUSTOM(gdf_gros_CUSTOM)
        gdf_gros_CUSTOM.echelle_shp_CUSTOM = "CUSTOM"
        gdf_gros_CUSTOM.liste_echelle_shp_par_CUSTOM = ['DEP']
        return gdf_gros_CUSTOM

    ##########################################################################################
    #Reférentiels
    ##########################################################################################
    def creation_dict_couche_geom(self,liste_echelle_REF_projet):
        dict_geom_REF = {}
        self.liste_echelle_REF_projet = liste_echelle_REF_projet
        for echelle_carto_REF in self.liste_echelle_REF_projet:
            if echelle_carto_REF == 'ME':
                perimetre_ME= gpd.read_file("/mnt/g/travail/carto/projets basiques/PAOT global 5.0/couche geom/ME/BV ME AG 2021.shp")
                #On dégage le premier FR
                perimetre_ME[self.colonne_fond_carte] = perimetre_ME[self.colonne_fond_carte].str[2:]
                perimetre_ME = perimetre_ME.rename(columns={'geometry':'geometry_ME'})
                perimetre_ME = perimetre_ME.set_geometry('geometry_ME')
                perimetre_ME['surface_ME'] = perimetre_ME.area
                dict_geom_REF['gdf_ME'] = perimetre_ME
            if echelle_carto_REF == 'PPG':
                perimetre_PPG = PPG.import_shp_PPG(self)
                perimetre_PPG['surface_PPG'] = perimetre_PPG.area
                dict_geom_REF['gdf_PPG'] = perimetre_PPG
        return dict_geom_REF

    def ajout_code_geom_REF(self,dict_geom_REF,dict_code_geom_REF):
        for echelle_carto_REF in self.liste_echelle_REF_projet:
            if echelle_carto_REF == 'PPG':
                dict_geom_REF['gdf_PPG'] = pd.merge(dict_geom_REF['gdf_PPG'],dict_code_geom_REF['info_gdf_PPG'],on='CODE_PPG')
        return dict_geom_REF

    ##########################################################################################
    #Attribution géographique : entre CUSTOM et rérentiels
    ##########################################################################################
    def attribution_REF_CUSTOM(self,dict_geom_REF,gdf_gros_CUSTOM):
        dict_geom_decoupREF = {}
        def creation_decoup_REF_CUSTOM(dict_geom_REF,gdf_gros_CUSTOM):
            for echelle_carto_REF in self.liste_echelle_REF_projet:
                tempo_gdf_gros_CUSTOM = gdf_gros_CUSTOM
                tempo_gdf_gros_CUSTOM['NOM_MO'] = tempo_gdf_gros_CUSTOM.index
                gdf_decoup_REF_CUSTOM = gpd.overlay(dict_geom_REF['gdf_' + echelle_carto_REF],tempo_gdf_gros_CUSTOM, how='intersection')
                gdf_decoup_REF_CUSTOM['surface_decoup' + echelle_carto_REF +'_par_CUSTOM'] = gdf_decoup_REF_CUSTOM.area
                dict_geom_decoupREF['gdf_decoup' + echelle_carto_REF]=gdf_decoup_REF_CUSTOM
            return dict_geom_decoupREF

        def calcul_ratio_surf_REF_CUSTOM(dict_geom_decoupREF):
            for echelle_carto_REF in self.liste_echelle_REF_projet:
                dict_geom_decoupREF['gdf_decoup' + echelle_carto_REF]['ratio_surf'] = dict_geom_decoupREF['gdf_decoup' + echelle_carto_REF]['surface_decoup' + echelle_carto_REF +'_par_CUSTOM']/dict_geom_decoupREF['gdf_decoup' + echelle_carto_REF]['surface_'+echelle_carto_REF]
            return dict_geom_decoupREF

        def regles_tri_REF_CUSTOM(dict_geom_decoupREF):
            #Les régles de tri sont différentes pour le tableaux persos, on est plus laxiste pour le tableau perso par MO
            for echelle_carto_REF in self.liste_echelle_REF_projet:
                if self.type_donnees =='action' or self.type_donnees =='toutes_pressions':
                    if echelle_carto_REF=='ME':
                        dict_geom_decoupREF['gdf_decoup' + echelle_carto_REF] = dict_geom_decoupREF['gdf_decoup' + echelle_carto_REF].loc[(dict_geom_decoupREF['gdf_decoup' + echelle_carto_REF]['ratio_surf'] > 0.3) | (dict_geom_decoupREF['gdf_decoup' + echelle_carto_REF]['surface_decoup' + echelle_carto_REF +'_par_CUSTOM']>10000000)]
                    if echelle_carto_REF=='PPG':
                        dict_geom_decoupREF['gdf_decoup' + echelle_carto_REF] = dict_geom_decoupREF['gdf_decoup' + echelle_carto_REF].loc[(dict_geom_decoupREF['gdf_decoup' + echelle_carto_REF]['ratio_surf'] > 0.3)]
                if self.type_donnees =='tableaux_perso_par_MO_ou_dep':
                    if echelle_carto_REF=='ME':
                        dict_geom_decoupREF['gdf_decoup' + echelle_carto_REF] = dict_geom_decoupREF['gdf_decoup' + echelle_carto_REF].loc[(dict_geom_decoupREF['gdf_decoup' + echelle_carto_REF]['ratio_surf'] > 0.2) | (dict_geom_decoupREF['gdf_decoup' + echelle_carto_REF]['surface_decoup' + echelle_carto_REF +'_par_CUSTOM']>200000)]
                    if echelle_carto_REF=='PPG':
                        dict_geom_decoupREF['gdf_decoup' + echelle_carto_REF] = dict_geom_decoupREF['gdf_decoup' + echelle_carto_REF].loc[(dict_geom_decoupREF['gdf_decoup' + echelle_carto_REF]['ratio_surf'] > 0.2)]
            return dict_geom_decoupREF

        def creation_dict_dict_CUSTOM_listeREF(dict_geom_decoupREF):
            dict_df_listeREF_CUSTOM = {}
            dict_dict_listeREF_CUSTOM = {}
            for echelle_carto_REF in self.liste_echelle_REF_projet:
                dict_df_listeREF_CUSTOM[echelle_carto_REF] = dict_geom_decoupREF['gdf_decoup' + echelle_carto_REF].groupby('NOM_MO').agg({'CODE_'+echelle_carto_REF:lambda x: list(x)})
                dict_df_listeREF_CUSTOM[echelle_carto_REF].columns = ["liste_decoup" + echelle_carto_REF + "_CUSTOM"]
                dict_dict_listeREF_CUSTOM[echelle_carto_REF] = dict_df_listeREF_CUSTOM[echelle_carto_REF].to_dict()
            dict_dict_listeREF_CUSTOM = {'decoup'+echelle_carto_REF + '_CUSTOM': v for echelle_carto_REF,v in dict_dict_listeREF_CUSTOM.items()}
            return dict_df_listeREF_CUSTOM,dict_dict_listeREF_CUSTOM

        def creation_dict_dict_CUSTOM_decoupREF(dict_geom_decoupREF):
            dict_dict_decoupREF_CUSTOM = {}
            for echelle_carto_REF in self.liste_echelle_REF_projet:
                dict_dict_decoupREF_CUSTOM['gdf_decoup' + echelle_carto_REF] = {}
                for CUSTOM in self:
                    dict_dict_decoupREF_CUSTOM['gdf_decoup' + echelle_carto_REF][CUSTOM]=dict_geom_decoupREF['gdf_decoup' + echelle_carto_REF].loc[dict_geom_decoupREF['gdf_decoup' + echelle_carto_REF]['NOM_MO']==CUSTOM]
            return dict_dict_decoupREF_CUSTOM

            #découpage des REF par le CUSTOM
        dict_geom_decoupREF = creation_decoup_REF_CUSTOM(dict_geom_REF,gdf_gros_CUSTOM)
        dict_geom_decoupREF = calcul_ratio_surf_REF_CUSTOM(dict_geom_decoupREF)
        dict_geom_decoupREF = regles_tri_REF_CUSTOM(dict_geom_decoupREF)
        dict_df_listeREF_CUSTOM,dict_dict_listeREF_CUSTOM = creation_dict_dict_CUSTOM_listeREF(dict_geom_decoupREF)
        dict_dict_decoupREF_CUSTOM = creation_dict_dict_CUSTOM_decoupREF(dict_geom_decoupREF)
        return dict_df_listeREF_CUSTOM,dict_dict_listeREF_CUSTOM,dict_geom_decoupREF,dict_dict_decoupREF_CUSTOM
    #gdf_decoupREF_CUSTOM,gdf_decoupREF_tried_CUSTOM,df_CUSTOM_decoup_listeREF,dict_listeREF_CUSTOM

    def actualisation_keys_dict_si_CUSTOM_vide(self,dict_dict_listeREF_CUSTOM):
        for echelle_carto_REF in self.liste_echelle_REF_projet:
            liste_CUSTOM_non_vide = list(dict_dict_listeREF_CUSTOM["decoup" + echelle_carto_REF + "_CUSTOM"]["liste_decoup" + echelle_carto_REF + "_CUSTOM"].keys())
        liste_complete_CUSTOM = list(self.keys())
        liste_CUSTOM_vide = list(set(liste_complete_CUSTOM) - set(liste_CUSTOM_non_vide))
        for CUSTOM_vide in liste_CUSTOM_vide:
            self.pop(CUSTOM_vide)
        liste_nom_CUSTOM = liste_CUSTOM_non_vide
        return self

    def actualisation_liste_nom_CUSTOM(self):
        for echelle_carto_REF in self.liste_echelle_REF_projet:
            liste_CUSTOM_non_vide = list(self.keys())
        liste_nom_CUSTOM = liste_CUSTOM_non_vide
        return liste_nom_CUSTOM

    def creation_dict_dict_listeREF_REF(self,dict_geom_REF):
        dict_geomREF_decoupREF = {}
        def creation_decoup_REF_REF(dict_geom_REF,liste_echelle_REF_projet):
            liste_combinaison_REF = list(itertools.combinations(liste_echelle_REF_projet, 2))
            liste_combinaison_REF = [list(x) for x in liste_combinaison_REF]
            for [REF1,REF2] in liste_combinaison_REF:
                gdf_decoup_REF1_REF2 = gpd.overlay(dict_geom_REF['gdf_' + REF1][['CODE_'+REF1,'geometry_'+REF1]], dict_geom_REF['gdf_' + REF2][['CODE_'+REF2,'geometry_'+REF2,'surface_'+REF2]], how='intersection')
                gdf_decoup_REF2_REF1 = gpd.overlay(dict_geom_REF['gdf_' + REF2][['CODE_'+REF2,'geometry_'+REF2]], dict_geom_REF['gdf_' + REF1][['CODE_'+REF1,'geometry_'+REF1,'surface_'+REF1]], how='intersection')
                dict_geomREF_decoupREF['gdf_decoup' + REF1 +'_' + REF2]=gdf_decoup_REF1_REF2
                dict_geomREF_decoupREF['gdf_decoup' + REF2 +'_' + REF1]=gdf_decoup_REF2_REF1
                dict_geomREF_decoupREF['gdf_decoup' + REF1 +'_' + REF2]['surface_decoup' + REF1] = dict_geomREF_decoupREF['gdf_decoup' + REF1 +'_' + REF2]['geometry'].area
                dict_geomREF_decoupREF['gdf_decoup' + REF2 +'_' + REF1]['surface_decoup' + REF2] = dict_geomREF_decoupREF['gdf_decoup' + REF2 +'_' + REF1]['geometry'].area
                dict_geomREF_decoupREF['gdf_decoup' + REF1 +'_' + REF2]['ratio_surf'] = dict_geomREF_decoupREF['gdf_decoup' + REF1 +'_' + REF2]['surface_decoup' + REF1]/dict_geomREF_decoupREF['gdf_decoup' + REF1 +'_' + REF2]['surface_' + REF2]
                dict_geomREF_decoupREF['gdf_decoup' + REF2 +'_' + REF1]['ratio_surf'] = dict_geomREF_decoupREF['gdf_decoup' + REF2 +'_' + REF1]['surface_decoup' + REF2]/dict_geomREF_decoupREF['gdf_decoup' + REF2 +'_' + REF1]['surface_' + REF1]

            return dict_geomREF_decoupREF,liste_combinaison_REF

        def regle_tri_decoupREF_REF(dict_geomREF_decoupREF):
            #On dégage si un périmétre est clairement plus grand que l'autre (ex : PPG > ME)
            for nom_decoup,df_decoupREF_REF in dict_geomREF_decoupREF.items():
                df_decoupREF_REF = df_decoupREF_REF.loc[df_decoupREF_REF['ratio_surf'] > 0.1]
                dict_geomREF_decoupREF[nom_decoup] = df_decoupREF_REF
            return dict_geomREF_decoupREF

        def apply_dict_dict_REF_listeREF(dict_geomREF_decoupREF):
            dict_dict_listeREF_REF = {}
            for nom_decoup,ddf_decoupREF_REF in dict_geomREF_decoupREF.items():
                REF1 = list(ddf_decoupREF_REF)[0].split('_')[1]
                REF2 = list(ddf_decoupREF_REF)[1].split('_')[1]
                ddf_decoupREF_REF = ddf_decoupREF_REF.groupby('CODE_'+REF1).agg({'CODE_'+REF2:lambda x: list(x)})
                ddf_decoupREF_REF.columns = ['liste_decoup' + REF2 +'_' + REF1]
                dict_dict_listeREF_REF['decoup' + REF2 + '_' + REF1] = ddf_decoupREF_REF.to_dict()
            return dict_dict_listeREF_REF


            #découpage des REF par le CUSTOM
        dict_geomREF_decoupREF,liste_combinaison_REF = creation_decoup_REF_REF(dict_geom_REF,self.liste_echelle_REF_projet)
        dict_geomREF_decoupREF = regle_tri_decoupREF_REF(dict_geomREF_decoupREF)
        dict_dict_listeREF_REF = apply_dict_dict_REF_listeREF(dict_geomREF_decoupREF)
        return dict_dict_listeREF_REF

    ##########################################################################################
    #Couches de fond de carte (toujours é la ME pour l'eau)
    ##########################################################################################
    def creation_gdf_fond_carte_REF(self,dict_decoup_REF,dict_relation_shp_liste):
        list_gdf_fond_carte = []
        for nom_CUSTOM,entite_CUSTOM in self.items():
            CODE_CUSTOM = entite_CUSTOM.CODE_CUSTOM
            echelle_REF = entite_CUSTOM.echelle_base_REF
            gdf_fond_carte = GdfFondCarte(dict_decoup_REF['gdf_decoup'+echelle_REF+"_CUSTOM"].gdf)
            gdf_fond_carte = gdf_fond_carte.loc[gdf_fond_carte["CODE_"+echelle_REF].isin(dict_relation_shp_liste["dict_liste_"+ echelle_REF + "_par_CUSTOM"][CODE_CUSTOM])]
            gdf_fond_carte["echelle_REF"] = echelle_REF
            gdf_fond_carte["CODE_CUSTOM"] = CODE_CUSTOM
            gdf_fond_carte.columns = gdf_fond_carte.columns.str.replace(echelle_REF, "REF", regex=True)
            list_gdf_fond_carte.append(gdf_fond_carte)
        self.gdf_fond_carte = pd.concat(list_gdf_fond_carte)
        self.gdf_fond_carte = self.gdf_fond_carte.set_crs("epsg:2154")
        return self
    

    def ajout_info_fond_carte(self,dict_dict_info_REF):
        if self.info_fond_carte=="pression_MIA" or self.info_fond_carte=="etat_eco":
            df_pression_AG = dataframe.import_pression()
        if self.type_rendu =="carte":
            if self.info_fond_carte=="pression_MIA":
                if self.thematique=="MIA" and "MO" in self.liste_echelle_shp_CUSTOM_a_check:
                    df_pression_AG = df_pression_AG[[x for x in list(df_pression_AG) if x.startswith("P_")]+["CODE_ME"]]
                    df_pression_AG = df_pression_AG[['CODE_ME','P_hydromorpho',"P_MORPHO",'P_HYDRO','P_CONTI']]
                    df_pression_AG = df_pression_AG.rename({"CODE_ME":"CODE_REF"},axis=1)

                    liste_echelle_fond_carte = list(set(self.gdf_fond_carte['echelle_REF'].to_list()))
                    for REF in liste_echelle_fond_carte:
                        if "SME" in liste_echelle_fond_carte:
                            df_tempo_SME_ME = dict_dict_info_REF['df_info_SME'][['CODE_SME',"ME_maitre"]].rename({'ME_maitre':"CODE_REF"},axis=1)
                            df_tempo_SME_ME = df_tempo_SME_ME.groupby('CODE_REF').agg({'CODE_SME':lambda x: list(x)})
                            dict_CODE_ME_CODE_SME = dict(zip(df_tempo_SME_ME.index.to_list(),df_tempo_SME_ME["CODE_SME"].to_list()))
                            df_tempo_pression_AG_SME = copy.deepcopy(df_pression_AG)
                            df_tempo_pression_AG_SME['CODE_REF'] = df_tempo_pression_AG_SME['CODE_REF'].map(dict_CODE_ME_CODE_SME)
                            df_tempo_pression_AG_SME = df_tempo_pression_AG_SME.loc[~df_tempo_pression_AG_SME["CODE_REF"].isnull()]
                            df_tempo_pression_AG_SME = df_tempo_pression_AG_SME.explode("CODE_REF")
                            df_pression_AG = pd.concat([df_pression_AG,df_tempo_pression_AG_SME])
                            
                        self.gdf_fond_carte = pd.merge(self.gdf_fond_carte,df_pression_AG,on="CODE_REF")
                if self.thematique=="MIA" and "DEP" in self.liste_echelle_shp_CUSTOM_a_check:
                    pass
            if self.info_fond_carte=="etat_eco":
                df_pression_AG = df_pression_AG.rename({'CODE_ME':"CODE_REF"},axis=1)
                self.gdf_fond_carte = pd.merge(self.gdf_fond_carte,df_pression_AG,on="CODE_REF")

        return self


    def suppression_attributs_liste_CODE_CUSTOM(self,dict_dict_info_CUSTOM):
        for CODE_CUSTOM in dict_dict_info_CUSTOM:
            if dict_dict_info_CUSTOM[CODE_CUSTOM]['CUSTOM_a_reduire'] == True:
                self.liste_CODE_CUSTOM = [x for x in self.liste_CODE_CUSTOM if x!=CODE_CUSTOM]
        return self

    def suppression_CODE_CUSTOM_du_df_info_CUSTOM(self,df_info_CUSTOM,dict_dict_info_CUSTOM):
        for CODE_CUSTOM in dict_dict_info_CUSTOM:
            if dict_dict_info_CUSTOM[CODE_CUSTOM]['CUSTOM_a_reduire'] == True:
                df_info_CUSTOM = df_info_CUSTOM.loc[df_info_CUSTOM['CODE_MO']!=CODE_CUSTOM]
        return self

    def definition_si_CUSTOM_a_reduire(self,dict_special_CUSTOM_a_reduire,dict_boite_complete_pour_placement,gdf_CUSTOM,dict_dict_info_CUSTOM):
        for nom_boite_maitre,dict_boite_maitre in dict_boite_complete_pour_placement.items():
            for CODE_CUSTOM,df_CUSTOM in dict_boite_maitre.boite_complete.items():

                
                if dict_special_CUSTOM_a_reduire[CODE_CUSTOM]['df_taille_boite_complete']['surface_boite'].sum()/1000000>dict_config_espace['surface_limite'][self.taille_globale_carto]:
                    dict_special_CUSTOM_a_reduire[CODE_CUSTOM]['CUSTOM_a_reduire'] = True
                '''if gdf_CUSTOM.loc[gdf_CUSTOM['CODE_CUSTOM']==CODE_CUSTOM]["surface_CUSTOM"].iloc[0]:
                    dict_special_CUSTOM_a_reduire[CODE_CUSTOM]['CUSTOM_a_reduire'] = True'''
        return dict_special_CUSTOM_a_reduire

    def actualisation_dict_special_CUSTOM_a_reduire(self,dict_special_CUSTOM_a_reduire,df_info_CUSTOM,dict_dict_info_CUSTOM):
        for CODE_CUSTOM in dict_special_CUSTOM_a_reduire:
            dict_special_CUSTOM_a_reduire[CODE_CUSTOM]['CUSTOM_a_reduire'] = False
        return dict_special_CUSTOM_a_reduire

    def suppression_CODE_CUSTOM_du_dict_dict_info_CUSTOM(self,df_info_CUSTOM,dict_dict_info_CUSTOM):
        dict_dict_info_CUSTOM = {k:v for k,v in dict_dict_info_CUSTOM.items() if v['CUSTOM_a_reduire'] == False}
        return self

    ############################################################################################################################
    #Actualisation de la bb pour encadrer tout le CUSTOM ET les boites
    ############################################################################################################################
    def actualisation_dict_dict_info_CUSTOM_avec_bb(self):
        def modif_droite_boite_si_bloc_lm_present(self,taille_globale_carto):
            if "hauteur_ligne_indiv_droit" in list(self):
                facteur_angle = np.cos(dict_config_espace['angle_rotation_lm_paysage'][taille_globale_carto] * np.pi / 180. )
                self.loc[(self["type_placement_boite_final"].isin(["H","B"]))&(~self["hauteur_ligne_indiv_droit"].isnull()),"droite_boite_complete"] = self["droite_boite_complete"] + (self["taille_hauteur_boite_biais"]-self["ecart_hauteur_origine"])*facteur_angle
            return self

        df_info_CUSTOM = self.df_info_CUSTOM
        list_temp_contour = []
        for nom_CUSTOM,entite_CUSTOM in self.items():
            for type_boite,dict_boite_complete in entite_CUSTOM.items():
                df_contour_tempo = dict_boite_complete.df_contour
                df_contour_tempo['CODE_CUSTOM'] = entite_CUSTOM.CODE_CUSTOM
                if dict_boite_complete.orientation =="normal":
                    df_contour_tempo = modif_droite_boite_si_bloc_lm_present(df_contour_tempo,dict_boite_complete.taille_globale_carto)
                list_temp_contour.append(df_contour_tempo)
        df_contour_global = pd.concat(list_temp_contour)
        df_contour_global = df_contour_global.groupby("CODE_CUSTOM").agg({'gauche_boite_complete':'min','droite_boite_complete':'max','haut_boite_complete':'max','bas_boite_complete':'min'})
        self.gdf_CUSTOM = DictBoiteComplete.actualisation_cote_bb_CUSTOM(self.gdf_CUSTOM,df_info_CUSTOM,df_contour_global)
        self.df_info_CUSTOM = pd.merge(self.gdf_CUSTOM[["CODE_CUSTOM","min_x_CUSTOM","max_x_CUSTOM","min_y_CUSTOM","max_y_CUSTOM"]],self.df_info_CUSTOM,on='CODE_CUSTOM',suffixes=[None,'_a_supprimer'])
        self.df_info_CUSTOM =self.df_info_CUSTOM.loc[:,~self.df_info_CUSTOM.columns.str.endswith('_a_supprimer')]
        dict_info_CUSTOM = self.df_info_CUSTOM.to_dict(orient='records')[0]
        for nom_CUSTOM,entite_CUSTOM in self.items():
            dict_info_CUSTOM['ratio_hauteur_largeur'] = (dict_info_CUSTOM['max_y_CUSTOM']-dict_info_CUSTOM['min_y_CUSTOM'])/(dict_info_CUSTOM['max_x_CUSTOM']-dict_info_CUSTOM['min_x_CUSTOM'])
            if self.format_rendu == 'A3':
                ratio_hauteur_largeur = 310/296
                if dict_info_CUSTOM['ratio_hauteur_largeur']>ratio_hauteur_largeur:
                    dict_info_CUSTOM['max_x_CUSTOM'] = dict_info_CUSTOM['X_centre_CUSTOM'] + (dict_info_CUSTOM['max_y_CUSTOM']-dict_info_CUSTOM['min_y_CUSTOM'])/ratio_hauteur_largeur/2
                    dict_info_CUSTOM['min_x_CUSTOM'] = dict_info_CUSTOM['X_centre_CUSTOM'] - (dict_info_CUSTOM['max_y_CUSTOM']-dict_info_CUSTOM['min_y_CUSTOM'])/ratio_hauteur_largeur/2
            if (dict_info_CUSTOM['max_y_CUSTOM']-dict_info_CUSTOM['min_y_CUSTOM'])>(dict_info_CUSTOM['max_x_CUSTOM']-dict_info_CUSTOM['min_x_CUSTOM']):
                dict_info_CUSTOM['orient_CUSTOM'] = "portrait"
            if (dict_info_CUSTOM['max_y_CUSTOM']-dict_info_CUSTOM['min_y_CUSTOM'])<(dict_info_CUSTOM['max_x_CUSTOM']-dict_info_CUSTOM['min_x_CUSTOM']):
                dict_info_CUSTOM['orient_CUSTOM'] = "paysage" 
        return self
    
    ############################################################################################################################
    #Export boite, bloc, atlas
    ############################################################################################################################    
    def export_bloc(self):
        id_projet = self.type_donnees + '_' + self.thematique + '_' + self.public_cible
        for nom_entite_CUSTOM,contenu_CUSTOM in self.items():
            for type_boite,contenu_boite in contenu_CUSTOM.items():
                '''df_faits = pd.read_csv("/mnt/g/travail/carto/projets basiques/PAOT global 5.0/test/repartition_df_faits.csv")
                contenu_boite.df_contour = pd.merge(contenu_boite.df_contour,df_faits[['CODE_MO',"etat"]].rename({'CODE_MO':"CODE_REF"},axis=1),on="CODE_REF",how="left")'''
                contenu_boite.df_contour = contenu_boite.df_contour.to_file("/mnt/g/travail/carto/projets basiques/PAOT global 5.0/"+ id_projet +"/couche_boite/boite_complete_" + id_projet + "_" + contenu_boite.orientation + ".shp",engine="fiona", encoding='utf-8')
        for nom_entite_CUSTOM,contenu_CUSTOM in self.items():
            for type_boite,contenu_boite in contenu_CUSTOM.items():
                for type_bloc,contenu_bloc in contenu_boite.items():
                    if hasattr(contenu_bloc,"df_indiv"):
                        contenu_bloc.df_indiv = contenu_bloc.df_indiv.to_file("/mnt/g/travail/carto/projets basiques/PAOT global 5.0/"+ id_projet +"/couche_bloc/" + contenu_bloc.type + '_' + contenu_boite.orientation + ".shp",engine="fiona", encoding='utf-8') 
                    if not hasattr(contenu_bloc,"df_indiv"):
                        contenu_bloc.df = contenu_bloc.df.to_file("/mnt/g/travail/carto/projets basiques/PAOT global 5.0/"+ id_projet +"/couche_bloc/" + contenu_bloc.type + '_' + contenu_boite.orientation + ".shp",engine="fiona", encoding='utf-8') 
        return self    
    
    def export_atlas(self):
        id_projet = self.type_donnees + '_' + self.thematique + '_' + self.public_cible
        self.gdf_CUSTOM = self.gdf_CUSTOM.set_geometry('geometry_CUSTOM')
        self.gdf_CUSTOM['date'] = date.today().strftime("%d/%m/%Y")
        self.gdf_CUSTOM.to_file("/mnt/g/travail/carto/projets basiques/PAOT global 5.0/" + id_projet + "/atlas_" + id_projet + ".shp",engine="fiona")
        return self
    

    def export_fond_carte(self):
        id_projet = self.type_donnees + '_' + self.thematique + '_' + self.public_cible
        self.gdf_fond_carte.to_file("/mnt/g/travail/carto/projets basiques/PAOT global 5.0/" + id_projet + "/couche_fond_carte/fond_carte_" + id_projet + ".shp",engine="fiona")
        return self    
    ############################################################################################################################
    #Osmose
    ############################################################################################################################
    def repartition_donnees_dans_nom_CUSTOM_maitre(dict_CUSTOM_maitre,dict_df_donnees):
        for entite_CUSTOM,contenu_CUSTOM in dict_CUSTOM_maitre.items():
            contenu_CUSTOM.df = dict_df_donnees["dict_dict_df_actions_originaux"][entite_CUSTOM].df
            contenu_CUSTOM.df_Points_de_blocage = dict_df_donnees["dict_dict_df_actions_originaux"][entite_CUSTOM].df_Points_de_blocage
            contenu_CUSTOM.df_Etapes = dict_df_donnees["dict_dict_df_actions_originaux"][entite_CUSTOM].df_Etapes
            contenu_CUSTOM.df_Financeurs = dict_df_donnees["dict_dict_df_actions_originaux"][entite_CUSTOM].df_Financeurs
            contenu_CUSTOM.df_attributs = dict_df_donnees["dict_dict_df_actions_originaux"][entite_CUSTOM].df_attributs
        return dict_CUSTOM_maitre

    ############################################################################################################################
    #Tableaux excel
    ############################################################################################################################
    def export_fichier_excel_perso(self,dict_relation_shp_liste,dict_dict_info_REF,dict_decoupREF):
        #Remplissage des onglets
        dict_dict_info_REF['df_info_ME'] = DictDfInfoShp.ajout_surface_ME(dict_dict_info_REF['df_info_ME'],dict_decoupREF)
        dict_dict_info_REF = DictDfInfoShp.boost_df_info_ME(dict_dict_info_REF)
        if "df_info_SME" in list(dict_dict_info_REF):
            dict_dict_info_REF = DictDfInfoShp.boost_df_info_SME(dict_dict_info_REF)

        for entite_CUSTOM,contenu_CUSTOM in self.items():
            contenu_CUSTOM = NomCUSTOMMaitre.definition_liste_REF_hors_echelle_base_REF(contenu_CUSTOM)
            
        for entite_CUSTOM,contenu_CUSTOM in self.items():
            #Creation fichier tableaux excel avec onglets vierges
            if contenu_CUSTOM.echelle_REF == "MO":
                excel_modif = config_DORA.recuperation_excel_MIA_MO_vierge_DORA()
                worksheet= excel_modif['tableau a remplir']
                worksheet['A1']='Tableau DORA actions MIA MO ' + entite_CUSTOM
            if contenu_CUSTOM.echelle_REF == "DEP":
                excel_modif = config_DORA.recuperation_excel_MIA_DEP_vierge_DORA()
                worksheet= excel_modif['tableau a remplir']
                worksheet['A1']='Tableau DORA actions MIA DEP ' + entite_CUSTOM                
            #excel_modif = tableau_excel.ajout_onglet_AIDE_liste_ME(excel_modif,contenu_CUSTOM.CODE_CUSTOM,dict_relation_shp_liste,dict_dict_info_REF)
            excel_modif = tableau_excel.ajout_onglet_AIDE_liste_REF(self,excel_modif,contenu_CUSTOM.CODE_CUSTOM,dict_relation_shp_liste,dict_dict_info_REF,contenu_CUSTOM)
            excel_modif = tableau_excel.ajout_onglet_info_PPG(excel_modif,contenu_CUSTOM.CODE_CUSTOM,dict_relation_shp_liste,dict_dict_info_REF)
            
            excel_modif = tableau_excel.ajout_onglet_Lien_REF_ME(self,excel_modif,contenu_CUSTOM.CODE_CUSTOM,dict_relation_shp_liste,dict_dict_info_REF,contenu_CUSTOM)
            #excel_modif.save("/mnt/h//Tableau_suivi_MIA_" + contenu_CUSTOM.echelle_REF + "_vierge_" + entite_CUSTOM + ".xlsx")
            return excel_modif

    def export_fichier_excel_osmose_conversion_dora(self):
        workbook = config_DORA.recuperation_excel_vierge_Osmose()
        for entite_CUSTOM,contenu_CUSTOM in self.items():
            workbook.save("/mnt/g/travail/carto/projets basiques/PAOT global 5.0/Tableau actions/MIA/" + contenu_CUSTOM.echelle_REF + "/Tableau_suivi_osmose_" + entite_CUSTOM + ".xlsx")        
            excel_modif = load_workbook("/mnt/g/travail/carto/projets basiques/PAOT global 5.0/Tableau actions/MIA/" + contenu_CUSTOM.echelle_REF + "/Tableau_suivi_osmose_" + entite_CUSTOM + ".xlsx", read_only=False, keep_vba=False)
            excel_modif = tableau_excel.ajout_onglet_actions(excel_modif,contenu_CUSTOM.df)
            excel_modif = tableau_excel.ajout_onglet_attributs(excel_modif,contenu_CUSTOM.df_attributs)
            excel_modif.save("/mnt/g/travail/carto/projets basiques/PAOT global 5.0/Tableau actions/MIA/" + contenu_CUSTOM.echelle_REF + "/Tableau_suivi_osmose_" + entite_CUSTOM + ".xlsx")

    def export_fichier_excel_osmose_maj(self):
        workbook = load_workbook("/mnt/g/travail/carto/projets basiques/PAOT global 5.0/Tableau actions/TABLEAU_MAJ_OSMOSE_VIERGE.xlsx", read_only=False)
        for type_dict_donnees,dict_donnees in self.items():
            for entite_CUSTOM,contenu_CUSTOM in dict_donnees.items():
                workbook.save("/mnt/g/travail/carto/projets basiques/PAOT global 5.0/Tableau actions/MAJ Osmose/tableau_propre_MAJ/Tableau_suivi_MIA_MO_vierge_" + entite_CUSTOM + ".xlsx")        
                excel_modif = load_workbook("/mnt/g/travail/carto/projets basiques/PAOT global 5.0/Tableau actions/MAJ Osmose/tableau_propre_MAJ/Tableau_suivi_MIA_MO_vierge_" + entite_CUSTOM + ".xlsx", read_only=False, keep_vba=False)
                excel_modif = tableau_excel.ajout_onglet_actions(excel_modif,contenu_CUSTOM.df)
                excel_modif = tableau_excel.ajout_onglet_attributs(excel_modif,contenu_CUSTOM.df_attributs)
                excel_modif.save("/mnt/g/travail/carto/projets basiques/PAOT global 5.0/Tableau actions/MAJ Osmose/tableau_propre_MAJ/Tableau_suivi_MIA_MO_vierge_" + entite_CUSTOM + ".xlsx")
                if hasattr(contenu_CUSTOM,"df_actions_a_recreer"):
                    excel_modif = load_workbook("/mnt/g/travail/carto/projets basiques/PAOT global 5.0/Tableau actions/TABLEAU_CREATION_OSMOSE_VIERGE.xlsx", read_only=False, keep_vba=False)
                    excel_modif = tableau_excel.ajout_onglet_actions(excel_modif,contenu_CUSTOM.df_actions_a_recreer)
                    excel_modif = tableau_excel.ajout_onglet_attributs(excel_modif,contenu_CUSTOM.df_attributs_df_actions_a_creer)
                    excel_modif.save("/mnt/g/travail/carto/projets basiques/PAOT global 5.0/Tableau actions/MAJ Osmose/tableau_propre_MAJ/Tableau_suivi_MIA_MO_vierge_" + entite_CUSTOM + "_creer_actions" + ".xlsx")

    def attributs_df_filtre_dans_dict_CUSTOM_maitre(self,dict_df_donnees):
        for entite_CUSTOM,contenu_CUSTOM in self.items():
            contenu_CUSTOM.df_filtre_tableau = dict_df_donnees["dict_dict_df_actions_originaux"][contenu_CUSTOM.NOM_CUSTOM].df_filtre_tableau
        return self

    def attributs_df_log_erreur_dans_dict_CUSTOM_maitre(self,dict_df_donnees):
        for entite_CUSTOM,contenu_CUSTOM in self.items():
            contenu_CUSTOM.df_log_erreur = dict_df_donnees["dict_dict_df_actions_originaux"][contenu_CUSTOM.NOM_CUSTOM].df_log_erreur
        return self

    def attributs_echelle_REF_dans_dict_CUSTOM_maitre(self,dict_df_donnees):
        for entite_CUSTOM,contenu_CUSTOM in self.items():
            contenu_CUSTOM.echelle_df = dict_df_donnees["dict_dict_df_actions_originaux"][contenu_CUSTOM.NOM_CUSTOM].echelle_df
        return self

    def export_log_df_erreur(self):
        for entite_CUSTOM,contenu_CUSTOM in self.items():
            print("attention, il y a plusieurs logs")
            path = os.path.join(self.folder_path,"log_"+contenu_CUSTOM.NOM_CUSTOM+".csv")
            connect_path.upload_file_vers_s3("CUSTOM",contenu_CUSTOM.df_log_erreur,path)
            
        return contenu_CUSTOM.df_log_erreur

    def export_tableau_excel_complet(self, dict_df_donnees):
        for entite_CUSTOM, contenu_CUSTOM in self.items():
            fichier_excel = dict_df_donnees["dict_dict_df_actions_originaux"][contenu_CUSTOM.NOM_CUSTOM].fichier_brut
            wb = load_workbook(fichier_excel)
            
            #On doit d'abbord créer un fichier "tableau_final" à partir du fichier brut. Puis c'est ce fichier qu'on va modifier avec le fichier issu de Python
            path = os.path.join(self.folder_path,"tableau_proposition_"+contenu_CUSTOM.NOM_CUSTOM+".xlsx")
            connect_path.upload_file_vers_s3("CUSTOM",wb,path)

            # Charger le fichier avec load_workbook (fichier_excel est un BytesIO)
            

            # Accéder à la feuille "tableau a remplir"
            feuille_a_remplir = wb['tableau a remplir']

            # Créer un NamedStyle et copier les propriétés d'une cellule
            format_cellule_rouge = NamedStyle(name="format_cellule_rouge")
            format_cellule_rouge.font = copy(feuille_a_remplir.cell(row=4, column=1).font)
            format_cellule_rouge.alignment = copy(feuille_a_remplir.cell(row=4, column=1).alignment)
            format_cellule_rouge.fill = copy(feuille_a_remplir.cell(row=4, column=1).fill)
            format_cellule_rouge.border = copy(feuille_a_remplir.cell(row=4, column=1).border)

            # Ajouter le NamedStyle au Workbook
            wb.add_named_style(format_cellule_rouge)

            # Appliquer ce style à des cellules spécifiques (colonnes d'en-tête)
            for numero_col, nom_colonne in enumerate(list(contenu_CUSTOM.df_filtre_tableau.columns)):
                feuille_a_remplir.cell(row=4, column=50 + numero_col).value = nom_colonne
                feuille_a_remplir.cell(row=4, column=50 + numero_col).style = "format_cellule_rouge"

            # Remplir les cellules avec les données du DataFrame
            for index, row in contenu_CUSTOM.df_filtre_tableau.iterrows():
                for numero_col, contenu_ligne in enumerate(row):
                    feuille_a_remplir.cell(row=6 + index, column=50 + numero_col).value = contenu_ligne

            # Ajout d'une valeur à la cellule A1
            feuille_a_remplir.cell(row=1, column=1).value = (
                "Tableau DORA actions MIA " + contenu_CUSTOM.echelle_df + " final " + contenu_CUSTOM.NOM_CUSTOM
            )

            # Sauvegarder le fichier Excel modifié
            path = os.path.join(self.folder_path,"tableau_proposition_"+contenu_CUSTOM.NOM_CUSTOM+".xlsx")
            connect_path.upload_file_vers_s3("CUSTOM",wb,path)

        return "Exportation terminée"

class NomCUSTOMMaitre(dict):
    def __init__(self,TYPE_REF=None):
        self.TYPE_REF = TYPE_REF

    def attributs_CUSTOM(self,liste_echelle_shp_CUSTOM_a_check,CODE_CUSTOM,dict_dict_info_REF,TYPE_REF=None):
        self.CODE_CUSTOM = CODE_CUSTOM
        if TYPE_REF!=None:
            dict_conv_CODE_NOM = dict_dict_info_REF['df_info_'+TYPE_REF].dict_CODE_NOM
            NOM_CUSTOM = dict_conv_CODE_NOM[CODE_CUSTOM]
            self.NOM_CUSTOM = NOM_CUSTOM
            self.echelle_REF = TYPE_REF
        if TYPE_REF==None:
            for echelle_REF in liste_echelle_shp_CUSTOM_a_check:
                if CODE_CUSTOM in dict_dict_info_REF['df_info_'+echelle_REF]['CODE_'+echelle_REF].to_list():
                    dict_conv_CODE_NOM = dict_dict_info_REF['df_info_'+echelle_REF].dict_CODE_NOM
                    NOM_CUSTOM = dict_conv_CODE_NOM[CODE_CUSTOM]
                    self.NOM_CUSTOM = NOM_CUSTOM
                    self.echelle_REF = echelle_REF
                    if echelle_REF=="MO":
                        self.echelle_carto_globale = 'petite'                        
                    if echelle_REF=="DEP":
                        self.echelle_carto_globale = 'grande'
                    pass
            for echelle_REF in liste_echelle_shp_CUSTOM_a_check:
                if hasattr(self,"echelle_REF")==False:
                    self.echelle_REF = "CUSTOM"
                    self.echelle_carto_globale = 'petite'
        return self
    
    def chercher_gdf_CUSTOM(self,dict_geom_REF):
        TYPE_REF = self.echelle_REF
        if TYPE_REF!="CUSTOM":
            self.gdf = dict_geom_REF['gdf_'+TYPE_REF].gdf.loc[dict_geom_REF['gdf_'+TYPE_REF].gdf['CODE_'+TYPE_REF]==self.CODE_CUSTOM]
            self.gdf = self.gdf.rename({"surface_"+TYPE_REF:"surface_CUSTOM","CODE_"+TYPE_REF:"CODE_CUSTOM","NOM_"+TYPE_REF:"NOM_CUSTOM","geometry_"+TYPE_REF:"geometry_CUSTOM"},axis=1)
            self.gdf = self.gdf.set_geometry("geometry_CUSTOM")
        return self 

    def ajout_info_geom_CUSTOM(self):
        gdf_CUSTOM = self.gdf_CUSTOM
        """ajout des info échelle, centre, bb et orientation"""
        list_colonne_a_garder_issues_gdf_gros_CUSTOM = ["CODE_CUSTOM","NOM_CUSTOM","echelle",'X_centre_CUSTOM','Y_centre_CUSTOM','orient_CUSTOM','min_x_CUSTOM','min_y_CUSTOM','max_x_CUSTOM','max_y_CUSTOM']
        df_info_CUSTOM_tempo = gdf_CUSTOM.join(gdf_CUSTOM.bounds)
        df_info_CUSTOM_tempo = df_info_CUSTOM_tempo.rename({"maxx":'max_x_CUSTOM',"maxy":'max_y_CUSTOM',"minx":'min_x_CUSTOM',"miny":'min_y_CUSTOM'},axis=1)
        #BB
        df_info_CUSTOM_tempo['taille_GD']=df_info_CUSTOM_tempo['max_y_CUSTOM']-df_info_CUSTOM_tempo['min_y_CUSTOM']
        df_info_CUSTOM_tempo['taille_BH']=df_info_CUSTOM_tempo['max_x_CUSTOM']-df_info_CUSTOM_tempo['min_x_CUSTOM']
        #Orientation
        df_info_CUSTOM_tempo.loc[df_info_CUSTOM_tempo['taille_GD']>df_info_CUSTOM_tempo['taille_BH'],'orient_CUSTOM']='portrait'
        df_info_CUSTOM_tempo.loc[df_info_CUSTOM_tempo['taille_GD']<df_info_CUSTOM_tempo['taille_BH'],'orient_CUSTOM']='paysage'

        #Echelle
        dict_NOM_CUSTOM_echelle = {k:v.echelle_carto_globale for k,v in self.items()}
        df_info_CUSTOM_tempo['echelle'] = df_info_CUSTOM_tempo['NOM_CUSTOM'].map(dict_NOM_CUSTOM_echelle)

        #Centre
        df_info_CUSTOM_tempo['geometry_centre_CUSTOM'] = df_info_CUSTOM_tempo.representative_point()
        df_info_CUSTOM_tempo['X_centre_CUSTOM']=df_info_CUSTOM_tempo['geometry_centre_CUSTOM'].x
        df_info_CUSTOM_tempo['Y_centre_CUSTOM']=df_info_CUSTOM_tempo['geometry_centre_CUSTOM'].y
        df_info_CUSTOM = df_info_CUSTOM_tempo.reset_index()[list_colonne_a_garder_issues_gdf_gros_CUSTOM]
        #ajout de l'ID
        df_info_CUSTOM['id_atlas'] = df_info_CUSTOM['CODE_CUSTOM'] + '%' + self.type_donnees + '%' + self.thematique + '%' + self.public_cible

        #Gestion de l'ALIAS
        liste_nom_potentiellement_alias = ['ALIAS',"Alias","alias"]
        for x in liste_nom_potentiellement_alias:
            df_info_CUSTOM = df_info_CUSTOM.rename({x:"ALIAS"},axis=1)
        if "ALIAS" in list(df_info_CUSTOM):
            df_info_CUSTOM.loc[df_info_CUSTOM['ALIAS']!=df_info_CUSTOM['ALIAS'],"ALIAS"] = df_info_CUSTOM['NOM_CUSTOM']
        if "ALIAS" not in list(df_info_CUSTOM):
            df_info_CUSTOM['ALIAS'] = df_info_CUSTOM['NOM_CUSTOM']

        return df_info_CUSTOM            
        

    def definition_liste_REF_hors_echelle_base_REF(self):
        if self.echelle_REF=="MO":
            self.liste_REF = ["PPG","ROE"]
        if self.echelle_REF=="DEP":
            self.liste_REF = ["MO","SAGE","BVG","PPG","ROE"]
        return self

    def creation_bloc_texte_simple(self,orientation,type_icone,sous_type,colonne_texte,nom_CUSTOM,taille_globale_carto="petite"):
        df_bloc_texte_simple = BlocTexteSimple(taille_globale_carto)
        df_bloc_texte_simple.type_icone = type_icone
        df_bloc_texte_simple.sous_type = sous_type
        df_bloc_texte_simple.colonne_texte = colonne_texte
        df_bloc_texte_simple.nom_CUSTOM = nom_CUSTOM
        df_bloc_texte_simple.liste_nom_colonne_a_garder = []
        if orientation=="normal":
            self['dict_boite_maitre_normal']['df_bloc_texte'] = df_bloc_texte_simple 
        if orientation=="orthogonal":
            self['dict_boite_maitre_ortho']['df_bloc_texte'] = df_bloc_texte_simple 
        return self

    def creation_bloc_icone(self,orientation,type_icone,sous_type,colonne_nb_icone,nom_CUSTOM,taille_globale_carto="petite"):
        df_bloc_icone = BlocIcone(taille_globale_carto)
        df_bloc_icone.type_icone = type_icone
        df_bloc_icone.type = 'bloc_icone'
        df_bloc_icone.avancement_max = 4
        df_bloc_icone.sous_type = sous_type
        df_bloc_icone.colonne_nb_icone = colonne_nb_icone
        df_bloc_icone.nom_CUSTOM = nom_CUSTOM
        df_bloc_icone.liste_nom_colonne_a_garder = []
        if orientation=="normal":
            self['dict_boite_maitre_normal']['df_bloc_icone'] = df_bloc_icone 
        if orientation=="orthogonal":
            self['dict_boite_maitre_ortho']['df_bloc_icone'] = df_bloc_icone 
        return self

    def creation_bloc_lignes_multiples(self,orientation,type_icone,sous_type,colonne_texte,nom_CUSTOM,taille_globale_carto="petite"):
        df_bloc_lignes_multiples = BlocLignesMultiples(taille_globale_carto)
        df_bloc_lignes_multiples.type = 'bloc_lignes_multiples'
        df_bloc_lignes_multiples.type_icone = type_icone
        df_bloc_lignes_multiples.avancement_max = 4
        df_bloc_lignes_multiples.sous_type = sous_type
        df_bloc_lignes_multiples.colonne_texte = colonne_texte
        df_bloc_lignes_multiples.nom_CUSTOM = nom_CUSTOM
        df_bloc_lignes_multiples.liste_nom_colonne_a_garder = []
        if orientation=="normal":
            self['dict_boite_maitre_normal']['df_bloc_lignes_multiples'] = df_bloc_lignes_multiples 
        if orientation=="orthogonal":
            self['dict_boite_maitre_ortho']['df_bloc_lignes_multiples'] = df_bloc_lignes_multiples         
        return self